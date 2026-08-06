import openpyxl
from pathlib import Path

path = Path(r'C:\Users\hevma\OneDrive\_H-Queex\Finance\Fiance App by Hev\H-Queex_Financial_Control by Claude V6.0 for App.xlsm')
wb = openpyxl.load_workbook(path, data_only=False)
print('SHEETS', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print('---', name, 'rows', ws.max_row, 'cols', ws.max_column)
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
        rows.append(row)
    for row in rows:
        print(row)
    print()
