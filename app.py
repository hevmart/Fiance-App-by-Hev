from __future__ import annotations

import csv
import json
import time
import threading
from calendar import monthrange
from datetime import date, datetime
from functools import lru_cache
from io import StringIO
from pathlib import Path
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from flask import Flask, Response, redirect, render_template, request, url_for
from openpyxl import load_workbook

app = Flask(__name__)


class WorkbookWriteError(RuntimeError):
    """Raised when the workbook cannot be updated because it is locked or busy."""


@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["Surrogate-Control"] = "no-store"
    return response


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK_NAME = "H-Queex_Financial_Control V7.0.xlsm"
WORKBOOK_PATH = BASE_DIR / DEFAULT_WORKBOOK_NAME
SUBSCRIPTIONS_PATH = BASE_DIR / "subscriptions.json"
ARCHIVE_PATH = BASE_DIR / "archives.json"
AUDIT_LOG_PATH = BASE_DIR / "audit-log.json"
BUSINESS_PROFILE_PATH = BASE_DIR / "business-profile.json"
CHART_OF_ACCOUNTS_PATH = BASE_DIR / "chart-of-accounts.json"
LEDGER_JOURNAL_PATH = BASE_DIR / "ledger-journal.json"
CAPITAL_ASSETS_PATH = BASE_DIR / "capital-assets.json"
PAYROLL_PATH = BASE_DIR / "payroll-register.json"
BANK_STATEMENTS_PATH = BASE_DIR / "bank-statements.json"
INCOME_PATH = BASE_DIR / "income.json"
EXPENSES_PATH = BASE_DIR / "expenses.json"
INVOICES_PATH = BASE_DIR / "invoices.json"
CLIENTS_PATH = BASE_DIR / "clients.json"
SUPPLIERS_PATH = BASE_DIR / "suppliers.json"
SHEET_JSON_PATHS = {
    "Income": INCOME_PATH,
    "Expenses": EXPENSES_PATH,
    "Invoices": INVOICES_PATH,
    "Clients": CLIENTS_PATH,
    "Suppliers": SUPPLIERS_PATH,
}

if not SUBSCRIPTIONS_PATH.exists():
    SUBSCRIPTIONS_PATH.write_text("[]", encoding="utf-8")

_sheet_write_lock = threading.Lock()
SUBSCRIPTION_FREQUENCIES = {"monthly": 1, "quarterly": 3, "yearly": 12}
SUBSCRIPTION_STATUSES = ("active", "paused", "cancelled")
BUSINESS_STRUCTURES = ("sole_trader", "limited_company")
INCOME_PAYMENT_METHODS = {
    "sole_trader": ["Business Bank Account", "Stripe", "PayPal", "Cash", "Proprietor Capital"],
    "limited_company": ["Business Bank Account", "Stripe", "PayPal", "Cash", "Director Loan Account"],
}
EXPENSE_PAYMENT_METHODS = {
    "sole_trader": ["Business Bank", "Credit Card", "Cash", "Proprietor Contribution"],
    "limited_company": ["Business Bank", "Credit Card", "Cash", "Director Contribution"],
}
VAT_RATE_OPTIONS = ["0%", "4.8%", "9%", "13.5%", "23%", "Exempt"]
VAT_TREATMENT_OPTIONS = [
    {"value": "standard", "label": "Standard"},
    {"value": "zero_rated", "label": "Zero-rated"},
    {"value": "exempt", "label": "Exempt"},
    {"value": "reverse_charge", "label": "Reverse charge"},
]
SUPPLY_TYPE_OPTIONS = [
    {"value": "services", "label": "Services"},
    {"value": "goods", "label": "Goods"},
]
INVOICE_STATUS_OPTIONS = ["Draft", "Issued", "Paid", "Partially Paid", "Overdue", "Bad Debt", "Cancelled"]
EXPENSE_STATUS_OPTIONS = ["Pending", "Approved", "Paid", "Auto-posted", "Cancelled"]
EXPENSE_INPUT_VAT_OPTIONS = ["Yes", "No", "Partial"]
EXPENSE_DEDUCTIBILITY_OPTIONS = ["Fully Deductible", "Partially Deductible", "Non-Deductible"]
RECONCILIATION_OPTIONS = ["Reconciled", "Unreconciled"]
YES_NO_OPTIONS = ["Yes", "No"]
PAYROLL_STATUS_OPTIONS = ["Draft", "Approved", "Paid", "Filed"]
RECONCILIATION_MATCH_DAYS = 3
VAT_TURNOVER_THRESHOLDS = {
    "services": {"label": "Services", "annual_limit": 42000.0},
    "goods": {"label": "Goods", "annual_limit": 85000.0},
}
VAT_THRESHOLD_WARNING_RATIO = 0.8
PHASE_POLICY = {
    "sole_trader": {
        "tax_regime": "Income Tax (Form 11)",
        "estimated_tax_rate": 0.20,
        "report_template": "Form 11 outputs",
        "owner_account_label": "Proprietor Capital Account",
        "next_filing_deadline": "Preliminary tax and Form 11 due by 31 October (ROS extension may apply)",
    },
    "limited_company": {
        "tax_regime": "Corporation Tax (CT1)",
        "estimated_tax_rate": 0.125,
        "report_template": "CT1 outputs",
        "owner_account_label": "Director Loan Account",
        "next_filing_deadline": "CT1 due 9 months after accounting year end",
    },
}
DEFAULT_CHART_OF_ACCOUNTS = [
    {"code": "1000", "name": "Cash at Bank", "type": "Asset", "tax_treatment": "n/a", "active": True},
    {"code": "1100", "name": "Accounts Receivable", "type": "Asset", "tax_treatment": "n/a", "active": True},
    {"code": "1200", "name": "Input VAT Control", "type": "Asset", "tax_treatment": "vat", "active": True},
    {"code": "2000", "name": "Accounts Payable", "type": "Liability", "tax_treatment": "n/a", "active": True},
    {"code": "2100", "name": "Output VAT Control", "type": "Liability", "tax_treatment": "vat", "active": True},
    {"code": "2200", "name": "PAYE / USC / PRSI Control", "type": "Liability", "tax_treatment": "payroll", "active": True},
    {"code": "3000", "name": "Owner / Director Account", "type": "Equity", "tax_treatment": "entity", "active": True},
    {"code": "4000", "name": "Consulting / Project Fees", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "4010", "name": "Retainer / Advisory Fees", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "4020", "name": "Service Add-ons", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "4030", "name": "Grant Income", "type": "Income", "tax_treatment": "non-trading", "active": True},
    {"code": "4040", "name": "BTWEA / Welfare Support", "type": "Income", "tax_treatment": "personal-excluded", "active": True},
    {"code": "4900", "name": "Other Income", "type": "Income", "tax_treatment": "trading", "active": True},
    {"code": "5000", "name": "Software and Subscriptions", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5010", "name": "Domain / Hosting / Website", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5020", "name": "Professional Fees", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5030", "name": "Marketing and Advertising", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5040", "name": "Bank and Transaction Fees", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5200", "name": "Equipment and Hardware", "type": "Expense", "tax_treatment": "capital-check", "active": True},
    {"code": "5300", "name": "Salaries and Wages", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5310", "name": "Employer PRSI", "type": "Expense", "tax_treatment": "deductible", "active": True},
    {"code": "5900", "name": "Non-Deductible Items", "type": "Expense", "tax_treatment": "non-deductible", "active": True},
]
INCOME_CATEGORY_ACCOUNT_MAP = {
    "consulting / project fees": "4000",
    "retainer / advisory fees": "4010",
    "service add-ons": "4020",
    "grant income": "4030",
    "btwea / welfare support": "4040",
    "other income": "4900",
}
EXPENSE_CATEGORY_ACCOUNT_MAP = {
    "software and subscriptions": "5000",
    "domain / hosting / website": "5010",
    "professional fees": "5020",
    "marketing and advertising": "5030",
    "bank and transaction fees": "5040",
    "equipment and hardware": "5200",
    "non-deductible items": "5900",
}
EXPENSE_CATEGORY_DEDUCTIBILITY_MAP = {
    "software and subscriptions": "Fully Deductible",
    "domain / hosting / website": "Fully Deductible",
    "professional fees": "Fully Deductible",
    "insurance": "Fully Deductible",
    "marketing and advertising": "Fully Deductible",
    "office supplies and stationery": "Fully Deductible",
    "equipment and hardware": "Partially Deductible",
    "home office expenses": "Partially Deductible",
    "travel and subsistence": "Partially Deductible",
    "bank and transaction fees": "Fully Deductible",
    "training and professional development": "Fully Deductible",
    "salaries and wages": "Fully Deductible",
    "subcontractor fees": "Fully Deductible",
    "non-deductible items": "Non-Deductible",
}
ENTITY_ROUTE_MAP = {
    "income": "income_view",
    "expense": "expenses_view",
    "invoice": "invoices_view",
    "client": "clients_view",
    "supplier": "suppliers_view",
    "subscription": "subscriptions_view",
    "payroll": "payroll_view",
}
WORKBOOK_ENTITY_CONFIG = {
    "income": {"sheet": "Income", "audit_type": "income"},
    "expense": {"sheet": "Expenses", "audit_type": "expense"},
    "invoice": {"sheet": "Invoices", "audit_type": "invoice"},
    "client": {"sheet": "Clients", "audit_type": "client"},
    "supplier": {"sheet": "Suppliers", "audit_type": "supplier"},
}


def _default_business_profile() -> dict[str, Any]:
    return {
        "business_name": "H-Queex",
        "owner_name": "Hevandro Martire",
        "cro_number": "790968",
        "registration_date": "2026-08-04",
        "structure": "sole_trader",
        "vat_registered": True,
        "vat_threshold_basis": "services",
        "transition_date": "",
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }


def _normalize_business_structure(value: Any) -> str:
    structure = str(value or "").strip().lower()
    return structure if structure in BUSINESS_STRUCTURES else "sole_trader"


def _load_business_profile() -> dict[str, Any]:
    if not BUSINESS_PROFILE_PATH.exists():
        return _default_business_profile()

    try:
        payload = json.loads(BUSINESS_PROFILE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return _default_business_profile()

    if not isinstance(payload, dict):
        return _default_business_profile()

    defaults = _default_business_profile()
    return {
        "business_name": str(payload.get("business_name") or defaults["business_name"]).strip(),
        "owner_name": str(payload.get("owner_name") or defaults["owner_name"]).strip(),
        "cro_number": str(payload.get("cro_number") or defaults["cro_number"]).strip(),
        "registration_date": str(payload.get("registration_date") or defaults["registration_date"]).strip(),
        "structure": _normalize_business_structure(payload.get("structure")),
        "vat_registered": bool(payload.get("vat_registered", defaults["vat_registered"])),
        "vat_threshold_basis": _normalize_vat_threshold_basis(payload.get("vat_threshold_basis")),
        "transition_date": str(payload.get("transition_date") or "").strip(),
        "updated_at": str(payload.get("updated_at") or defaults["updated_at"]).strip(),
    }


def _save_business_profile(profile: dict[str, Any]) -> None:
    normalized = {
        "business_name": str(profile.get("business_name") or "H-Queex").strip(),
        "owner_name": str(profile.get("owner_name") or "Hevandro Martire").strip(),
        "cro_number": str(profile.get("cro_number") or "790968").strip(),
        "registration_date": str(profile.get("registration_date") or "2026-08-04").strip(),
        "structure": _normalize_business_structure(profile.get("structure")),
        "vat_registered": bool(profile.get("vat_registered", True)),
        "vat_threshold_basis": _normalize_vat_threshold_basis(profile.get("vat_threshold_basis")),
        "transition_date": str(profile.get("transition_date") or "").strip(),
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    BUSINESS_PROFILE_PATH.write_text(json.dumps(normalized, indent=2), encoding="utf-8")


def _normalize_vat_threshold_basis(value: Any) -> str:
    basis = str(value or "").strip().lower()
    return basis if basis in VAT_TURNOVER_THRESHOLDS else "services"


def _parse_transaction_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if not value:
        return None

    text = str(value).strip()
    if not text:
        return None

    for date_format in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    return None


def _resolve_phase_tag(transaction_date: Any) -> str:
    profile = _load_business_profile()
    structure = _normalize_business_structure(profile.get("structure"))
    transition_date = _parse_transaction_date(profile.get("transition_date"))
    record_date = _parse_transaction_date(transaction_date)

    if structure == "sole_trader":
        return "Phase 1"

    if transition_date is None:
        return "Phase 2"

    if record_date is not None and record_date < transition_date:
        return "Phase 1"
    return "Phase 2"


def _phase_label_for_structure(structure: Any) -> str:
    if _normalize_business_structure(structure) == "limited_company":
        return "Phase 2 - Private Limited Company"
    return "Phase 1 - Sole Trader / Business Name"


def _append_message_to_path(path: str, message: str) -> str:
    separator = "&" if "?" in path else "?"
    return f"{path}{separator}message={quote(message)}"


def _build_phase_policy(summary: dict[str, Any], structure: str) -> dict[str, Any]:
    structure_key = _normalize_business_structure(structure)
    config = PHASE_POLICY.get(structure_key, PHASE_POLICY["sole_trader"])
    net_profit = _coerce_number(summary.get("net_cashflow", 0))
    taxable_profit = max(net_profit, 0.0)
    estimated_tax_due = round(taxable_profit * float(config.get("estimated_tax_rate", 0)), 2)
    return {
        "tax_regime": config.get("tax_regime", "Income Tax (Form 11)"),
        "estimated_tax_rate": float(config.get("estimated_tax_rate", 0.0)),
        "estimated_tax_due": estimated_tax_due,
        "report_template": config.get("report_template", "Form 11 outputs"),
        "owner_account_label": config.get("owner_account_label", "Proprietor Capital Account"),
        "next_filing_deadline": config.get("next_filing_deadline", ""),
        "taxable_profit_basis": taxable_profit,
    }


def _resolve_workbook_path() -> Path:
    candidates = [
        Path(WORKBOOK_PATH),
        BASE_DIR / DEFAULT_WORKBOOK_NAME,
    ]

    matching_workbooks = sorted(
        [
            path
            for path in BASE_DIR.glob("H-Queex_Financial_Control*.xls*")
            if path.is_file() and ".tmp-" not in path.name
        ],
        key=lambda path: path.name,
    )
    candidates.extend(matching_workbooks)

    if not matching_workbooks:
        legacy_candidates = [
            BASE_DIR / "H-Queex_Financial_Control by Claude V6.0 for App.xlsx",
            BASE_DIR / "H-Queex_Financial_Control by Claude V6.0 for App.xlsm",
        ]
        candidates.extend(legacy_candidates)

    for candidate in candidates:
        if candidate.exists():
            return candidate

    searched = ", ".join(str(candidate) for candidate in candidates)
    raise FileNotFoundError(
        f"Could not locate the finance workbook. Looked in: {searched}"
    )


SHEET_HEADERS = {
    "Income": ["Date", "Description", "Client / Source", "Category", "Invoice #", "Amount (€)", "Status", "Total incl. VAT (€)", "Payment Date"],
    "Expenses": [
        "Date (Registered)",
        "Supplier / Payee",
        "Supplier VAT Number",
        "Receipt / Invoice Ref",
        "Title",
        "Description",
        "Category",
        "Base Net Amount (€)",
        "Delivery (€)",
        "Fees (€)",
        "Other Charges (€)",
        "Discount (€)",
        "Net Amount (€)",
        "VAT Rate",
        "VAT Amount (€)",
        "Total (€)",
        "Input VAT Reclaimable",
        "Payment Method",
        "Deductibility Status",
        "Capital Expenditure Flag",
        "Receipt Attached",
        "Bank Reconciliation",
        "Status",
        "Notes",
    ],
    "Invoices": [
        "Invoice #",
        "Issue Date",
        "Due Date",
        "Client Name",
        "Client VAT Number",
        "Client Address",
        "Service / Product",
        "Net (€)",
        "VAT Rate",
        "VAT Amount (€)",
        "Total (€)",
        "Balance Due (€)",
        "Status",
        "Payment Method",
        "Payment Date",
        "Bank Reconciliation",
        "Notes",
    ],
    "AR": ["Client", "Invoice #", "Issue Date", "Due Date", "Total (€)", "Paid (€)", "Balance (€)", "Status"],
    "AP": ["Ref #", "Supplier Name", "Description", "Invoice Date", "Due Date", "Net (€)", "Total (€)", "Paid (€)", "Balance Due (€)", "Status"],
    "VAT": ["Period", "Output VAT — Sales (€)", "Input VAT — Purchases (€)", "Net VAT Due (€)", "VAT Paid (€)", "Balance (€)", "Due Date", "Status"],
    "Clients": ["Client Name", "Contact Person", "Email", "Phone", "Country"],
    "Suppliers": ["Supplier Name", "Contact Person", "Email", "Phone", "Country", "Default VAT Treatment"],
}

HEADER_ALIASES = {
    "Income": {
        "Amount": "Amount (€)",
        "Amount (€)": "Amount (€)",
        "Total": "Amount (€)",
        "Client Source": "Client / Source",
        "Client / Source": "Client / Source",
        "Client": "Client / Source",
        "Invoice Number": "Invoice #",
        "Invoice #": "Invoice #",
    },
    "Expenses": {
        "Amount": "Total (€)",
        "Base Net": "Base Net Amount (€)",
        "Base Net Amount": "Base Net Amount (€)",
        "Base Net Amount (€)": "Base Net Amount (€)",
        "Delivery": "Delivery (€)",
        "Delivery (€)": "Delivery (€)",
        "Fees": "Fees (€)",
        "Fees (€)": "Fees (€)",
        "Other Charges": "Other Charges (€)",
        "Other Charges (€)": "Other Charges (€)",
        "Discount": "Discount (€)",
        "Discount (€)": "Discount (€)",
        "Net Amount": "Net Amount (€)",
        "Net Amount (€)": "Net Amount (€)",
        "Total Amount": "Total (€)",
        "Supplier": "Supplier / Payee",
        "Supplier / Payee": "Supplier / Payee",
        "Supplier VAT": "Supplier VAT Number",
        "Supplier VAT Number": "Supplier VAT Number",
        "Receipt Ref": "Receipt / Invoice Ref",
        "Reference": "Receipt / Invoice Ref",
        "Input VAT": "Input VAT Reclaimable",
        "Deductibility": "Deductibility Status",
        "Capex": "Capital Expenditure Flag",
        "Receipt Attached": "Receipt Attached",
        "Bank Reconciliation": "Bank Reconciliation",
        "Notes": "Notes",
    },
    "Clients": {
        "Name": "Client Name",
        "Client": "Client Name",
        "Client Name": "Client Name",
    },
    "Suppliers": {
        "Name": "Supplier Name",
        "Supplier": "Supplier Name",
        "Supplier Name": "Supplier Name",
    },
    "Invoices": {
        "Invoice Number": "Invoice #",
        "Invoice #": "Invoice #",
        "Client VAT": "Client VAT Number",
        "VAT Number": "Client VAT Number",
        "Client Address": "Client Address",
        "Payment Date": "Payment Date",
        "Bank Reconciliation": "Bank Reconciliation",
        "Notes": "Notes",
    },
}


def _coerce_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("€", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def _coerce_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return round(value, 2)
    return value


def _parse_vat_rate(value: Any) -> float:
    text = str(value or "").strip().lower()
    if not text or text == "exempt":
        return 0.0
    cleaned = text.replace("%", "")
    try:
        return float(cleaned) / 100.0
    except ValueError:
        return 0.0


def _normalize_vat_treatment(value: Any, vat_rate: Any = None) -> str:
    text = str(value or "").strip().lower().replace("-", "_").replace(" ", "_")
    if text in {"standard", "zero_rated", "exempt", "reverse_charge"}:
        return text

    vat_rate_text = str(vat_rate or "").strip().lower()
    if vat_rate_text == "exempt":
        return "exempt"
    if vat_rate_text in {"0%", "0", "0.0", "0.00"}:
        return "zero_rated"
    return "standard"


def _normalize_supply_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"services", "service"}:
        return "services"
    if text in {"goods", "good"}:
        return "goods"
    return "services"


def _normalize_input_vat_reclaimable(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text == "partial":
        return "Partial"
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return "Yes"


def _normalize_deductibility_status(value: Any, category: Any) -> str:
    if str(value or "").strip() in EXPENSE_DEDUCTIBILITY_OPTIONS:
        return str(value).strip()
    category_key = _normalize_category_key(category)
    return EXPENSE_CATEGORY_DEDUCTIBILITY_MAP.get(category_key, "Fully Deductible")


def _normalize_yes_no(value: Any, *, default_yes: bool = False) -> str:
    text = str(value or "").strip().lower()
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return "Yes" if default_yes else "No"


def _normalize_reconciliation(value: Any) -> str:
    text = str(value or "").strip().lower()
    return "Reconciled" if text == "reconciled" else "Unreconciled"


def _normalize_payroll_status(value: Any) -> str:
    text = str(value or "").strip().title()
    return text if text in PAYROLL_STATUS_OPTIONS else "Draft"


def _normalize_payroll_payload(payload: dict[str, Any]) -> dict[str, Any]:
    gross_pay = round(_coerce_number(payload.get("Gross Pay (€)")), 2)
    paye_amount = round(max(_coerce_number(payload.get("PAYE (€)")), 0.0), 2)
    usc_amount = round(max(_coerce_number(payload.get("USC (€)")), 0.0), 2)
    employee_prsi_amount = round(max(_coerce_number(payload.get("Employee PRSI (€)")), 0.0), 2)
    employer_prsi_amount = round(max(_coerce_number(payload.get("Employer PRSI (€)")), 0.0), 2)
    deductions_total = round(paye_amount + usc_amount + employee_prsi_amount, 2)
    net_pay = round(max(gross_pay - deductions_total, 0.0), 2)
    employer_cost = round(gross_pay + employer_prsi_amount, 2)

    payload["Gross Pay (€)"] = f"{gross_pay:.2f}"
    payload["PAYE (€)"] = f"{paye_amount:.2f}"
    payload["USC (€)"] = f"{usc_amount:.2f}"
    payload["Employee PRSI (€)"] = f"{employee_prsi_amount:.2f}"
    payload["Employer PRSI (€)"] = f"{employer_prsi_amount:.2f}"
    payload["Net Pay (€)"] = f"{net_pay:.2f}"
    payload["Employer Cost (€)"] = f"{employer_cost:.2f}"
    payload["Status"] = _normalize_payroll_status(payload.get("Status"))
    payload["Bank Reconciliation"] = _normalize_reconciliation(payload.get("Bank Reconciliation"))
    return payload


def _load_payroll_entries() -> list[dict[str, Any]]:
    records = _load_json_records(PAYROLL_PATH)
    normalized: list[dict[str, Any]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        item = {
            "id": str(record.get("id") or uuid4()),
            "Pay Date": str(record.get("Pay Date") or "").strip(),
            "Payroll Period": str(record.get("Payroll Period") or "").strip(),
            "Employee Name": str(record.get("Employee Name") or "").strip(),
            "Gross Pay (€)": record.get("Gross Pay (€)", "0.00"),
            "PAYE (€)": record.get("PAYE (€)", "0.00"),
            "USC (€)": record.get("USC (€)", "0.00"),
            "Employee PRSI (€)": record.get("Employee PRSI (€)", "0.00"),
            "Employer PRSI (€)": record.get("Employer PRSI (€)", "0.00"),
            "Net Pay (€)": record.get("Net Pay (€)", "0.00"),
            "Employer Cost (€)": record.get("Employer Cost (€)", "0.00"),
            "Status": record.get("Status", "Draft"),
            "Payment Method": str(record.get("Payment Method") or "").strip(),
            "Payment Date": str(record.get("Payment Date") or "").strip(),
            "Bank Reconciliation": record.get("Bank Reconciliation", "Unreconciled"),
            "Notes": str(record.get("Notes") or "").strip(),
            "Phase Tag": str(record.get("Phase Tag") or "").strip(),
            "created_at": str(record.get("created_at") or datetime.now().isoformat(timespec="seconds")),
            "last_updated_at": str(record.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
        }
        _normalize_payroll_payload(item)
        normalized.append(item)
    normalized.sort(key=lambda item: (item.get("Pay Date") or "", item.get("Employee Name") or ""), reverse=True)
    return normalized


def _save_payroll_entries(entries: list[dict[str, Any]]) -> None:
    payload: list[dict[str, Any]] = []
    for entry in entries:
        item = dict(entry)
        _normalize_payroll_payload(item)
        payload.append(item)
    _save_json_records(PAYROLL_PATH, payload)


def _find_payroll_by_id(entries: list[dict[str, Any]], payroll_id: Any) -> dict[str, Any] | None:
    target_id = str(payroll_id or "").strip()
    if not target_id:
        return None
    for entry in entries:
        if str(entry.get("id") or "") == target_id:
            return entry
    return None


def _summarize_payroll_entries(entries: list[dict[str, Any]]) -> dict[str, Any]:
    liabilities_total = 0.0
    for entry in entries:
        liabilities_total += _coerce_number(entry.get("PAYE (€)"))
        liabilities_total += _coerce_number(entry.get("USC (€)"))
        liabilities_total += _coerce_number(entry.get("Employee PRSI (€)"))
        liabilities_total += _coerce_number(entry.get("Employer PRSI (€)"))

    return {
        "count": len(entries),
        "gross_total": round(sum(_coerce_number(entry.get("Gross Pay (€)")) for entry in entries), 2),
        "net_total": round(sum(_coerce_number(entry.get("Net Pay (€)")) for entry in entries), 2),
        "employer_cost_total": round(sum(_coerce_number(entry.get("Employer Cost (€)")) for entry in entries), 2),
        "liabilities_total": round(liabilities_total, 2),
        "paid_count": sum(1 for entry in entries if _normalize_payroll_status(entry.get("Status")) == "Paid"),
    }


def _export_payroll_csv(entries: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "pay_date",
            "payroll_period",
            "employee_name",
            "gross_pay_eur",
            "paye_eur",
            "usc_eur",
            "employee_prsi_eur",
            "employer_prsi_eur",
            "net_pay_eur",
            "employer_cost_eur",
            "status",
            "payment_method",
            "bank_reconciliation",
            "notes",
            "phase_tag",
        ]
    )
    for entry in sorted(entries, key=lambda item: (str(item.get("Pay Date") or ""), str(item.get("Employee Name") or "")), reverse=True):
        writer.writerow(
            [
                entry.get("Pay Date", ""),
                entry.get("Payroll Period", ""),
                entry.get("Employee Name", ""),
                round(_coerce_number(entry.get("Gross Pay (€)")), 2),
                round(_coerce_number(entry.get("PAYE (€)")), 2),
                round(_coerce_number(entry.get("USC (€)")), 2),
                round(_coerce_number(entry.get("Employee PRSI (€)")), 2),
                round(_coerce_number(entry.get("Employer PRSI (€)")), 2),
                round(_coerce_number(entry.get("Net Pay (€)")), 2),
                round(_coerce_number(entry.get("Employer Cost (€)")), 2),
                _normalize_payroll_status(entry.get("Status")),
                entry.get("Payment Method", ""),
                _normalize_reconciliation(entry.get("Bank Reconciliation")),
                entry.get("Notes", ""),
                entry.get("Phase Tag", ""),
            ]
        )
    return buffer.getvalue()


def _is_paid_status(entity_type: str, status: Any) -> bool:
    normalized = str(status or "").strip().lower()
    if entity_type == "expense":
        return normalized in {"paid", "auto-posted", "auto_posted"}
    if entity_type == "invoice":
        return normalized in {"paid", "partially paid", "partially_paid"}
    if entity_type == "payroll":
        return normalized in {"paid", "filed"}
    return False


def _apply_default_payment_date_for_paid(payload: dict[str, Any], entity_type: str, date_key: str) -> bool:
    if not _is_paid_status(entity_type, payload.get("Status")):
        return False
    if str(payload.get("Payment Date") or "").strip():
        return False
    fallback_date = str(payload.get(date_key) or "").strip()
    if _parse_iso_date(fallback_date) is not None:
        payload["Payment Date"] = fallback_date
        return True
    return False


def _parse_bank_statement_date(value: Any) -> str:
    parsed = _parse_iso_date(value)
    if parsed:
        return parsed.isoformat()
    raw = str(value or "").strip()
    if not raw:
        return ""
    for pattern in ["%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"]:
        try:
            return datetime.strptime(raw, pattern).date().isoformat()
        except ValueError:
            continue
    return ""


def _load_bank_statement_lines() -> list[dict[str, Any]]:
    lines = _load_json_records(BANK_STATEMENTS_PATH)
    normalized: list[dict[str, Any]] = []
    for line in lines:
        if not isinstance(line, dict):
            continue
        normalized.append(
            {
                "id": str(line.get("id") or uuid4()),
                "date": _parse_bank_statement_date(line.get("date")),
                "description": str(line.get("description") or "").strip(),
                "reference": str(line.get("reference") or "").strip(),
                "amount_eur": round(_coerce_number(line.get("amount_eur")), 2),
                "balance_eur": round(_coerce_number(line.get("balance_eur")), 2),
                "payment_method": str(line.get("payment_method") or "").strip(),
                "source_filename": str(line.get("source_filename") or "").strip(),
                "uploaded_at": str(line.get("uploaded_at") or ""),
            }
        )
    normalized.sort(key=lambda item: (item.get("date") or "", item.get("id") or ""), reverse=True)
    return normalized


def _save_bank_statement_lines(lines: list[dict[str, Any]]) -> None:
    payload: list[dict[str, Any]] = []
    for line in lines:
        payload.append(
            {
                "id": str(line.get("id") or uuid4()),
                "date": _parse_bank_statement_date(line.get("date")),
                "description": str(line.get("description") or "").strip(),
                "reference": str(line.get("reference") or "").strip(),
                "amount_eur": round(_coerce_number(line.get("amount_eur")), 2),
                "balance_eur": round(_coerce_number(line.get("balance_eur")), 2),
                "payment_method": str(line.get("payment_method") or "").strip(),
                "source_filename": str(line.get("source_filename") or "").strip(),
                "uploaded_at": str(line.get("uploaded_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )
    _save_json_records(BANK_STATEMENTS_PATH, payload)


def _ingest_bank_statement_csv(content: str, *, source_filename: str) -> dict[str, int]:
    existing_lines = _load_bank_statement_lines()
    existing_keys = {
        f"{line.get('date')}|{line.get('description')}|{line.get('reference')}|{round(_coerce_number(line.get('amount_eur')), 2):.2f}"
        for line in existing_lines
    }
    reader = csv.DictReader(StringIO(content))
    imported_count = 0
    skipped_count = 0

    for raw_row in reader:
        if not isinstance(raw_row, dict):
            skipped_count += 1
            continue
        row = {str(key or "").strip().lower(): value for key, value in raw_row.items()}
        date_text = _parse_bank_statement_date(
            row.get("date")
            or row.get("transaction date")
            or row.get("posted date")
            or row.get("value date")
        )
        description = str(row.get("description") or row.get("narrative") or row.get("details") or "").strip()
        reference = str(row.get("reference") or row.get("ref") or row.get("transaction id") or "").strip()
        payment_method = str(row.get("payment method") or row.get("account") or "").strip()

        amount_value = _try_parse_number(row.get("amount"))
        if amount_value is None:
            debit_value = _try_parse_number(row.get("debit") or row.get("withdrawal"))
            credit_value = _try_parse_number(row.get("credit") or row.get("deposit"))
            if debit_value is not None and credit_value is None:
                amount_value = -abs(debit_value)
            elif credit_value is not None and debit_value is None:
                amount_value = abs(credit_value)
            elif credit_value is not None and debit_value is not None:
                amount_value = abs(credit_value) - abs(debit_value)

        if not date_text or amount_value is None or abs(amount_value) < 0.005:
            skipped_count += 1
            continue

        rounded_amount = round(float(amount_value), 2)
        key = f"{date_text}|{description}|{reference}|{rounded_amount:.2f}"
        if key in existing_keys:
            skipped_count += 1
            continue

        existing_lines.append(
            {
                "id": str(uuid4()),
                "date": date_text,
                "description": description,
                "reference": reference,
                "amount_eur": rounded_amount,
                "balance_eur": round(_coerce_number(row.get("balance") or row.get("running balance")), 2),
                "payment_method": payment_method,
                "source_filename": source_filename,
                "uploaded_at": datetime.now().isoformat(timespec="seconds"),
            }
        )
        existing_keys.add(key)
        imported_count += 1

    _save_bank_statement_lines(existing_lines)
    return {"imported_count": imported_count, "skipped_count": skipped_count}


def _match_bank_statement_lines(reconciliation_rows: list[dict[str, Any]], statement_lines: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    candidates = [
        row
        for row in reconciliation_rows
        if bool(row.get("is_paid")) and row.get("bank_reconciliation") == "Unreconciled" and _coerce_number(row.get("amount_eur")) > 0
    ]
    used_candidate_indexes: set[int] = set()
    matched_statement_ids: set[str] = set()

    for row in reconciliation_rows:
        row["statement_match_count"] = 0
        row["statement_match_ids"] = []

    enriched_statement_lines: list[dict[str, Any]] = []
    for statement_line in statement_lines:
        line = dict(statement_line)
        line_date = _parse_iso_date(line.get("date"))
        line_amount = abs(round(_coerce_number(line.get("amount_eur")), 2))
        best_index: int | None = None
        best_score: tuple[int, float] | None = None

        for index, candidate in enumerate(candidates):
            if index in used_candidate_indexes:
                continue
            candidate_amount = round(_coerce_number(candidate.get("amount_eur")), 2)
            if abs(candidate_amount - line_amount) > 0.009:
                continue

            candidate_date = _parse_iso_date(candidate.get("date"))
            if line_date is None or candidate_date is None:
                continue

            date_diff = abs((candidate_date - line_date).days)
            if date_diff > RECONCILIATION_MATCH_DAYS:
                continue

            candidate_method = str(candidate.get("payment_method") or "").strip().lower()
            statement_method = str(line.get("payment_method") or "").strip().lower()
            method_score = 1 if candidate_method and statement_method and candidate_method == statement_method else 0
            score = (method_score, -float(date_diff))
            if best_score is None or score > best_score:
                best_score = score
                best_index = index

        if best_index is not None:
            used_candidate_indexes.add(best_index)
            matched_statement_ids.add(str(line.get("id") or ""))
            matched_row = candidates[best_index]
            matched_row["statement_match_count"] = int(matched_row.get("statement_match_count") or 0) + 1
            current_ids = matched_row.get("statement_match_ids") if isinstance(matched_row.get("statement_match_ids"), list) else []
            current_ids.append(str(line.get("id") or ""))
            matched_row["statement_match_ids"] = current_ids
            line["matched_entity_type"] = matched_row.get("entity_type")
            line["matched_reference"] = matched_row.get("reference")
        else:
            line["matched_entity_type"] = ""
            line["matched_reference"] = ""
        enriched_statement_lines.append(line)

    unmatched_lines = [line for line in enriched_statement_lines if str(line.get("id") or "") not in matched_statement_ids]

    for row in reconciliation_rows:
        if not bool(row.get("is_paid")):
            continue
        if row.get("bank_reconciliation") != "Unreconciled":
            continue
        if int(row.get("statement_match_count") or 0) == 0:
            reasons = row.get("exception_reasons") if isinstance(row.get("exception_reasons"), list) else []
            if "no_bank_statement_match" not in reasons:
                reasons.append("no_bank_statement_match")
            row["exception_reasons"] = reasons

    return enriched_statement_lines, unmatched_lines


def _export_bank_statement_csv(lines: list[dict[str, Any]], *, unmatched_only: bool = False) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "date",
            "description",
            "reference",
            "amount_eur",
            "balance_eur",
            "payment_method",
            "matched_entity_type",
            "matched_reference",
            "source_filename",
            "uploaded_at",
        ]
    )
    for line in lines:
        if unmatched_only and str(line.get("matched_entity_type") or "").strip():
            continue
        writer.writerow(
            [
                line.get("date", ""),
                line.get("description", ""),
                line.get("reference", ""),
                round(_coerce_number(line.get("amount_eur")), 2),
                round(_coerce_number(line.get("balance_eur")), 2),
                line.get("payment_method", ""),
                line.get("matched_entity_type", ""),
                line.get("matched_reference", ""),
                line.get("source_filename", ""),
                line.get("uploaded_at", ""),
            ]
        )
    return buffer.getvalue()


def _build_reconciliation_rows(data: dict[str, Any], payroll_entries: list[dict[str, Any]], *, today: date | None = None) -> list[dict[str, Any]]:
    current_day = today or date.today()
    rows: list[dict[str, Any]] = []

    for expense in data.get("sheets", {}).get("Expenses", []):
        transaction_date = _parse_iso_date(expense.get("Date (Registered)"))
        status = str(expense.get("Status") or "").strip()
        reconciliation = _normalize_reconciliation(expense.get("Bank Reconciliation"))
        is_paid = _is_paid_status("expense", status)
        amount = round(_coerce_number(expense.get("Total (€)")), 2)
        age_days = (current_day - transaction_date).days if transaction_date else None
        reasons: list[str] = []
        if is_paid and reconciliation == "Unreconciled" and age_days is not None and age_days > 7:
            reasons.append("paid_unreconciled_over_7_days")
        if is_paid and not str(expense.get("Payment Method") or "").strip():
            reasons.append("missing_payment_method")
        if is_paid and not str(expense.get("Supplier / Payee") or "").strip():
            reasons.append("missing_counterparty")
        rows.append(
            {
                "entity_type": "expense",
                "row_number": expense.get("__row_number"),
                "payroll_id": "",
                "date": transaction_date.isoformat() if transaction_date else str(expense.get("Date (Registered)") or ""),
                "counterparty": str(expense.get("Supplier / Payee") or ""),
                "reference": str(expense.get("Receipt / Invoice Ref") or expense.get("Title") or ""),
                "amount_eur": amount,
                "status": status,
                "is_paid": is_paid,
                "bank_reconciliation": reconciliation,
                "payment_method": str(expense.get("Payment Method") or ""),
                "age_days": age_days,
                "exception_reasons": reasons,
                "matching_key": "",
                "matching_group_size": 1,
                "statement_match_count": 0,
                "statement_match_ids": [],
            }
        )

    for invoice in data.get("sheets", {}).get("Invoices", []):
        payment_date = _parse_iso_date(invoice.get("Payment Date"))
        issue_date = _parse_iso_date(invoice.get("Issue Date"))
        effective_date = payment_date or issue_date
        status = _normalize_invoice_status(invoice.get("Status"))
        reconciliation = _normalize_reconciliation(invoice.get("Bank Reconciliation"))
        is_paid = _is_paid_status("invoice", status)
        amount = round(_coerce_number(invoice.get("Total (€)")), 2)
        age_days = (current_day - effective_date).days if effective_date else None
        reasons: list[str] = []
        if is_paid and not payment_date:
            reasons.append("missing_payment_date")
        if is_paid and reconciliation == "Unreconciled" and age_days is not None and age_days > 7:
            reasons.append("paid_unreconciled_over_7_days")
        if is_paid and not str(invoice.get("Payment Method") or "").strip():
            reasons.append("missing_payment_method")
        rows.append(
            {
                "entity_type": "invoice",
                "row_number": invoice.get("__row_number"),
                "payroll_id": "",
                "date": effective_date.isoformat() if effective_date else str(invoice.get("Issue Date") or ""),
                "counterparty": str(invoice.get("Client Name") or ""),
                "reference": str(invoice.get("Invoice #") or ""),
                "amount_eur": amount,
                "status": status,
                "is_paid": is_paid,
                "bank_reconciliation": reconciliation,
                "payment_method": str(invoice.get("Payment Method") or ""),
                "age_days": age_days,
                "exception_reasons": reasons,
                "matching_key": "",
                "matching_group_size": 1,
                "statement_match_count": 0,
                "statement_match_ids": [],
            }
        )

    for payroll in payroll_entries:
        pay_date = _parse_iso_date(payroll.get("Pay Date"))
        status = _normalize_payroll_status(payroll.get("Status"))
        reconciliation = _normalize_reconciliation(payroll.get("Bank Reconciliation"))
        is_paid = _is_paid_status("payroll", status)
        amount = round(_coerce_number(payroll.get("Net Pay (€)")), 2)
        age_days = (current_day - pay_date).days if pay_date else None
        reasons: list[str] = []
        if is_paid and reconciliation == "Unreconciled" and age_days is not None and age_days > 7:
            reasons.append("paid_unreconciled_over_7_days")
        if is_paid and not str(payroll.get("Payment Method") or "").strip():
            reasons.append("missing_payment_method")
        rows.append(
            {
                "entity_type": "payroll",
                "row_number": None,
                "payroll_id": str(payroll.get("id") or ""),
                "date": pay_date.isoformat() if pay_date else str(payroll.get("Pay Date") or ""),
                "counterparty": str(payroll.get("Employee Name") or ""),
                "reference": str(payroll.get("Payroll Period") or ""),
                "amount_eur": amount,
                "status": status,
                "is_paid": is_paid,
                "bank_reconciliation": reconciliation,
                "payment_method": str(payroll.get("Payment Method") or ""),
                "age_days": age_days,
                "exception_reasons": reasons,
                "matching_key": "",
                "matching_group_size": 1,
                "statement_match_count": 0,
                "statement_match_ids": [],
            }
        )

    groups: dict[str, int] = {}
    for row in rows:
        payment_method = str(row.get("payment_method") or "").strip().lower()
        date_text = str(row.get("date") or "")
        amount_value = round(_coerce_number(row.get("amount_eur")), 2)
        if not date_text or amount_value <= 0 or not payment_method:
            continue
        key = f"{date_text}|{amount_value:.2f}|{payment_method}"
        row["matching_key"] = key
        groups[key] = groups.get(key, 0) + 1

    for row in rows:
        key = str(row.get("matching_key") or "")
        row["matching_group_size"] = groups.get(key, 1) if key else 1

    rows.sort(key=lambda item: (str(item.get("date") or ""), str(item.get("entity_type") or ""), str(item.get("reference") or "")), reverse=True)
    return rows


def _summarize_reconciliation(rows: list[dict[str, Any]]) -> dict[str, Any]:
    paid_rows = [row for row in rows if bool(row.get("is_paid"))]
    unreconciled_paid_rows = [row for row in paid_rows if row.get("bank_reconciliation") == "Unreconciled"]
    reconciled_paid_rows = [row for row in paid_rows if row.get("bank_reconciliation") == "Reconciled"]
    exception_rows = [row for row in paid_rows if row.get("exception_reasons")]
    return {
        "tracked_rows": len(rows),
        "paid_rows": len(paid_rows),
        "reconciled_paid_rows": len(reconciled_paid_rows),
        "unreconciled_paid_rows": len(unreconciled_paid_rows),
        "exception_rows": len(exception_rows),
        "reconciled_amount_eur": round(sum(_coerce_number(row.get("amount_eur")) for row in reconciled_paid_rows), 2),
        "unreconciled_amount_eur": round(sum(_coerce_number(row.get("amount_eur")) for row in unreconciled_paid_rows), 2),
    }


def _export_reconciliation_csv(rows: list[dict[str, Any]], *, exceptions_only: bool = False) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "entity_type",
            "reference",
            "counterparty",
            "date",
            "amount_eur",
            "status",
            "bank_reconciliation",
            "payment_method",
            "is_paid",
            "age_days",
            "exception_reasons",
            "matching_group_size",
        ]
    )
    for row in rows:
        if exceptions_only and not row.get("exception_reasons"):
            continue
        writer.writerow(
            [
                row.get("entity_type", ""),
                row.get("reference", ""),
                row.get("counterparty", ""),
                row.get("date", ""),
                round(_coerce_number(row.get("amount_eur")), 2),
                row.get("status", ""),
                row.get("bank_reconciliation", ""),
                row.get("payment_method", ""),
                "Yes" if row.get("is_paid") else "No",
                row.get("age_days", ""),
                "|".join(row.get("exception_reasons", [])),
                row.get("matching_group_size", 1),
            ]
        )
    return buffer.getvalue()


def _is_capital_expense(payload: dict[str, Any]) -> bool:
    explicit_flag = str(payload.get("Capital Expenditure Flag") or "").strip().lower()
    if explicit_flag in {"yes", "y", "true", "1"}:
        return True
    total_amount = _coerce_number(payload.get("Total (€)"))
    return total_amount > 1000.0


def _apply_expense_compliance_fields(payload: dict[str, Any]) -> None:
    payload["Input VAT Reclaimable"] = _normalize_input_vat_reclaimable(payload.get("Input VAT Reclaimable"))
    payload["Deductibility Status"] = _normalize_deductibility_status(payload.get("Deductibility Status"), payload.get("Category"))
    payload["Receipt Attached"] = _normalize_yes_no(payload.get("Receipt Attached"), default_yes=False)
    payload["Bank Reconciliation"] = _normalize_reconciliation(payload.get("Bank Reconciliation"))
    payload["Capital Expenditure Flag"] = "Yes" if _is_capital_expense(payload) else "No"


def _load_capital_assets() -> list[dict[str, Any]]:
    return _load_json_records(CAPITAL_ASSETS_PATH)


def _save_capital_assets(assets: list[dict[str, Any]]) -> None:
    _save_json_records(CAPITAL_ASSETS_PATH, assets)


def _upsert_capital_asset_from_expense(payload: dict[str, Any], row_number: int, *, active: bool) -> None:
    assets = _load_capital_assets()
    target_id = f"expense-{row_number}"
    remaining = [asset for asset in assets if str(asset.get("id") or "") != target_id]

    if active:
        total_amount = round(_coerce_number(payload.get("Total (€)")), 2)
        annual_allowance = round(total_amount * 0.125, 2)
        remaining.append(
            {
                "id": target_id,
                "source": "expense",
                "expense_row_number": row_number,
                "acquisition_date": str(payload.get("Date (Registered)") or ""),
                "supplier": str(payload.get("Supplier / Payee") or ""),
                "description": str(payload.get("Description") or payload.get("Title") or ""),
                "category": str(payload.get("Category") or ""),
                "cost_eur": total_amount,
                "allowance_rate": 0.125,
                "allowance_years": 8,
                "annual_allowance_eur": annual_allowance,
                "phase_tag": str(payload.get("Phase Tag") or ""),
                "active": True,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            }
        )

    _save_capital_assets(remaining)


def _summarize_capital_assets(assets: list[dict[str, Any]]) -> dict[str, Any]:
    active_assets = [asset for asset in assets if bool(asset.get("active", True))]
    return {
        "asset_count": len(active_assets),
        "total_cost": round(sum(_coerce_number(asset.get("cost_eur")) for asset in active_assets), 2),
        "annual_allowance_total": round(sum(_coerce_number(asset.get("annual_allowance_eur")) for asset in active_assets), 2),
    }


def _export_capital_allowances_csv(assets: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "asset_id",
        "acquisition_date",
        "supplier",
        "description",
        "category",
        "cost_eur",
        "allowance_rate",
        "allowance_years",
        "annual_allowance_eur",
        "phase_tag",
        "active",
    ])
    for asset in assets:
        writer.writerow([
            asset.get("id", ""),
            asset.get("acquisition_date", ""),
            asset.get("supplier", ""),
            asset.get("description", ""),
            asset.get("category", ""),
            round(_coerce_number(asset.get("cost_eur")), 2),
            asset.get("allowance_rate", 0.125),
            asset.get("allowance_years", 8),
            round(_coerce_number(asset.get("annual_allowance_eur")), 2),
            asset.get("phase_tag", ""),
            "Yes" if bool(asset.get("active", True)) else "No",
        ])
    return buffer.getvalue()


def _normalize_invoice_status(value: Any) -> str:
    status = str(value or "").strip().lower()
    status_map = {
        "draft": "Draft",
        "issued": "Issued",
        "paid": "Paid",
        "partially paid": "Partially Paid",
        "partially_paid": "Partially Paid",
        "overdue": "Overdue",
        "bad debt": "Bad Debt",
        "bad_debt": "Bad Debt",
        "cancelled": "Cancelled",
    }
    return status_map.get(status, "Draft")


def _normalize_expense_status(value: Any) -> str:
    status = str(value or "").strip().lower()
    status_map = {
        "pending": "Pending",
        "approved": "Approved",
        "paid": "Paid",
        "auto-posted": "Auto-posted",
        "auto_posted": "Auto-posted",
        "autoposted": "Auto-posted",
        "cancelled": "Cancelled",
    }
    return status_map.get(status, "Pending")


def _next_invoice_number(issue_date_value: Any, existing_invoices: list[dict[str, Any]]) -> str:
    issue_date = _parse_transaction_date(issue_date_value) or date.today()
    year_token = str(issue_date.year)
    prefix = f"HQ-{year_token}-"
    max_seq = 0
    for row in existing_invoices:
        invoice_number = str(row.get("Invoice #") or "").strip().upper()
        if not invoice_number.startswith(prefix):
            continue
        suffix = invoice_number.replace(prefix, "", 1)
        if suffix.isdigit():
            max_seq = max(max_seq, int(suffix))
    return f"{prefix}{max_seq + 1:03d}"


def _apply_vat_classification(
    payload: dict[str, Any],
    *,
    vat_rate_key: str,
    vat_treatment_key: str = "VAT Treatment",
    supply_type_key: str = "Supply Type",
) -> None:
    payload[vat_treatment_key] = _normalize_vat_treatment(payload.get(vat_treatment_key), payload.get(vat_rate_key))
    payload[supply_type_key] = _normalize_supply_type(payload.get(supply_type_key))


def _is_vat_registered() -> bool:
    profile = _load_business_profile()
    return bool(profile.get("vat_registered", True))


def _calculate_threshold_status(turnover: float, annual_threshold: float) -> tuple[str, str, float, float, float]:
    warning_threshold = round(annual_threshold * VAT_THRESHOLD_WARNING_RATIO, 2)
    remaining_before_limit = round(max(annual_threshold - turnover, 0.0), 2)
    progress_pct = round((turnover / annual_threshold) * 100, 1) if annual_threshold > 0 else 0.0

    if annual_threshold <= 0:
        return "normal", "No VAT turnover threshold configured.", warning_threshold, remaining_before_limit, progress_pct
    if turnover >= annual_threshold:
        return "exceeded", "Annual turnover has exceeded the selected VAT registration threshold.", warning_threshold, remaining_before_limit, progress_pct
    if turnover >= warning_threshold:
        return "warning", "Annual turnover is above 80% of the selected VAT registration threshold.", warning_threshold, remaining_before_limit, progress_pct
    return "normal", "Annual turnover is below the VAT registration warning threshold.", warning_threshold, remaining_before_limit, progress_pct


def _collect_turnover_streams(income_rows: list[dict[str, Any]], invoice_rows: list[dict[str, Any]]) -> dict[str, float]:
    income_stream_totals = {"services": 0.0, "goods": 0.0}
    invoice_stream_totals = {"services": 0.0, "goods": 0.0}

    for row in income_rows:
        supply_type = _normalize_supply_type(row.get("Supply Type"))
        amount = _coerce_number(row.get("Total incl. VAT (€)", row.get("Amount (€)", 0)))
        income_stream_totals[supply_type] += amount

    for row in invoice_rows:
        supply_type = _normalize_supply_type(row.get("Supply Type"))
        amount = _coerce_number(row.get("Total (€)", row.get("Amount (€)", 0)))
        invoice_stream_totals[supply_type] += amount

    return {
        "services": round(max(income_stream_totals["services"], invoice_stream_totals["services"]), 2),
        "goods": round(max(income_stream_totals["goods"], invoice_stream_totals["goods"]), 2),
    }


def _compute_vat_threshold_summary(income_rows: list[dict[str, Any]], invoice_rows: list[dict[str, Any]], basis: str) -> dict[str, Any]:
    normalized_basis = _normalize_vat_threshold_basis(basis)
    threshold_config = VAT_TURNOVER_THRESHOLDS.get(normalized_basis, VAT_TURNOVER_THRESHOLDS["services"])
    annual_threshold = float(threshold_config.get("annual_limit") or 0.0)

    cash_turnover = round(sum(_coerce_number(row.get("Total incl. VAT (€)", row.get("Amount (€)", 0))) for row in income_rows), 2)
    invoiced_turnover = round(sum(_coerce_number(row.get("Total (€)", row.get("Amount (€)", 0))) for row in invoice_rows), 2)

    stream_turnovers = _collect_turnover_streams(income_rows, invoice_rows)
    taxable_turnover = round(stream_turnovers.get(normalized_basis, 0.0), 2)
    status, message, warning_threshold, remaining_before_limit, progress_pct = _calculate_threshold_status(taxable_turnover, annual_threshold)

    stream_trackers: list[dict[str, Any]] = []
    for stream_basis, stream_config in VAT_TURNOVER_THRESHOLDS.items():
        stream_annual_threshold = float(stream_config.get("annual_limit") or 0.0)
        stream_turnover = round(stream_turnovers.get(stream_basis, 0.0), 2)
        stream_status, _, stream_warning_threshold, stream_remaining_before_limit, stream_progress_pct = _calculate_threshold_status(stream_turnover, stream_annual_threshold)
        stream_trackers.append(
            {
                "basis": stream_basis,
                "basis_label": str(stream_config.get("label") or stream_basis.title()),
                "taxable_turnover": stream_turnover,
                "annual_threshold": round(stream_annual_threshold, 2),
                "warning_threshold": stream_warning_threshold,
                "remaining_before_limit": stream_remaining_before_limit,
                "progress_pct": stream_progress_pct,
                "status": stream_status,
                "is_selected": stream_basis == normalized_basis,
            }
        )

    return {
        "basis": normalized_basis,
        "basis_label": str(threshold_config.get("label") or normalized_basis.title()),
        "annual_threshold": round(annual_threshold, 2),
        "warning_threshold": warning_threshold,
        "warning_ratio_pct": int(VAT_THRESHOLD_WARNING_RATIO * 100),
        "taxable_turnover": taxable_turnover,
        "cash_turnover": cash_turnover,
        "invoiced_turnover": invoiced_turnover,
        "remaining_before_limit": remaining_before_limit,
        "progress_pct": progress_pct,
        "status": status,
        "message": message,
        "stream_trackers": stream_trackers,
    }


def _normalize_vat_fields(payload: dict[str, Any], *, net_key: str, total_key: str, vat_rate_key: str, vat_amount_key: str, vat_registered: bool) -> None:
    raw_net = _coerce_number(payload.get(net_key))
    raw_total = _coerce_number(payload.get(total_key))
    rate_value = payload.get(vat_rate_key) or "0%"
    rate_ratio = _parse_vat_rate(rate_value)
    explicit_vat = _coerce_number(payload.get(vat_amount_key))

    if raw_total <= 0 and raw_net > 0:
        raw_total = raw_net
    if raw_net <= 0 and raw_total > 0:
        raw_net = raw_total

    if not vat_registered:
        vat_amount = 0.0
        if raw_total <= 0:
            raw_total = raw_net
        raw_net = raw_total
        rate_value = "0%"
    else:
        vat_amount = 0.0
        if explicit_vat > 0:
            vat_amount = explicit_vat
        elif raw_total > 0 and raw_total >= raw_net:
            vat_amount = raw_total - raw_net
        elif raw_net > 0 and rate_ratio > 0:
            vat_amount = raw_net * rate_ratio

        if raw_total <= 0 and raw_net > 0:
            raw_total = raw_net + vat_amount
        if raw_net <= 0 and raw_total > 0:
            raw_net = max(raw_total - vat_amount, 0.0)

    payload[net_key] = f"{round(raw_net, 2):.2f}"
    payload[total_key] = f"{round(raw_total, 2):.2f}"
    payload[vat_rate_key] = str(rate_value)
    payload[vat_amount_key] = f"{round(vat_amount, 2):.2f}"


def _apply_expense_amount_breakdown(payload: dict[str, Any]) -> None:
    base_net = _coerce_number(payload.get("Base Net Amount (€)"))
    delivery = _coerce_number(payload.get("Delivery (€)"))
    fees = _coerce_number(payload.get("Fees (€)"))
    other_charges = _coerce_number(payload.get("Other Charges (€)"))
    discount = _coerce_number(payload.get("Discount (€)"))

    subtotal_before_discount = base_net + delivery + fees + other_charges
    taxable_net = max(subtotal_before_discount - discount, 0.0)

    payload["Base Net Amount (€)"] = f"{round(base_net, 2):.2f}"
    payload["Delivery (€)"] = f"{round(delivery, 2):.2f}"
    payload["Fees (€)"] = f"{round(fees, 2):.2f}"
    payload["Other Charges (€)"] = f"{round(other_charges, 2):.2f}"
    payload["Discount (€)"] = f"{round(discount, 2):.2f}"
    payload["Net Amount (€)"] = f"{round(taxable_net, 2):.2f}"


def _format_currency(value: float) -> str:
    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        numeric_value = _coerce_number(value)
    return f"€{numeric_value:,.2f}"


def _normalize_header_name(name: Any, sheet_name: str) -> str:
    if name is None:
        return ""
    text = str(name).strip()
    if not text:
        return ""
    aliases = HEADER_ALIASES.get(sheet_name, {})
    return aliases.get(text, text)


def _get_header_row(ws, sheet_name: str) -> list[Any]:
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if _is_header_row([value for value in row if value is not None], sheet_name):
            return [value for value in row if value is not None]
    return list(SHEET_HEADERS.get(sheet_name, []))


def _is_header_row(values: list[Any], sheet_name: str) -> bool:
    expected = SHEET_HEADERS.get(sheet_name, [])
    if not expected:
        return False

    normalized = [_normalize_header_name(value, sheet_name) for value in values]
    matches = sum(1 for value in normalized if value in expected)

    if matches >= 3:
        return True

    if sheet_name == "Income" and any("Income Register" in str(value) for value in values):
        return False

    return False


def _find_header_row_number(ws, sheet_name: str) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=12):
        values = [cell.value for cell in row if cell.value is not None]
        if _is_header_row(values, sheet_name):
            return row[0].row
    return None


def _save_workbook_atomic(wb, resolved_path: Path) -> None:
    temp_path = resolved_path.with_name(f"{resolved_path.stem}.tmp-{uuid4().hex}{resolved_path.suffix}")
    try:
        wb.save(temp_path)

        last_error: PermissionError | None = None
        for _ in range(10):
            try:
                temp_path.replace(resolved_path)
                return
            except PermissionError as exc:
                last_error = exc
                time.sleep(0.2)

        raise WorkbookWriteError("Workbook is busy. Close Excel for this file and retry.") from last_error
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def _load_sheet_records_raw(sheet_name: str) -> list[dict[str, Any]]:
    path = SHEET_JSON_PATHS[sheet_name]
    return _load_json_records(path)


def _save_sheet_records_raw(sheet_name: str, records: list[dict[str, Any]]) -> None:
    path = SHEET_JSON_PATHS[sheet_name]
    _save_json_records(path, records)


def _load_sheet_rows_with_row_numbers(sheet_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, record in enumerate(_load_sheet_records_raw(sheet_name), start=1):
        row = dict(record) if isinstance(record, dict) else {}
        row["__row_number"] = index
        rows.append(row)
    return rows


def _append_row_to_sheet(sheet_name: str, values: dict[str, Any]) -> int:
    with _sheet_write_lock:
        records = _load_sheet_records_raw(sheet_name)
        record = {k: v for k, v in values.items() if k != "__row_number"}
        records.append(record)
        _save_sheet_records_raw(sheet_name, records)
        return len(records)


def _delete_row_from_sheet(sheet_name: str, row_number: int) -> None:
    with _sheet_write_lock:
        records = _load_sheet_records_raw(sheet_name)
        if row_number < 1 or row_number > len(records):
            raise ValueError(f"Invalid row number for {sheet_name}: {row_number}")
        del records[row_number - 1]
        _save_sheet_records_raw(sheet_name, records)


def _update_row_in_sheet(sheet_name: str, row_number: int, values: dict[str, Any]) -> None:
    with _sheet_write_lock:
        records = _load_sheet_records_raw(sheet_name)
        if row_number < 1 or row_number > len(records):
            raise ValueError(f"Invalid row number for {sheet_name}: {row_number}")
        records[row_number - 1] = {k: v for k, v in values.items() if k != "__row_number"}
        _save_sheet_records_raw(sheet_name, records)


def _load_json_records(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        records = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    return records if isinstance(records, list) else []


def _save_json_records(path: Path, records: list[dict[str, Any]]) -> None:
    path.write_text(json.dumps(records, indent=2), encoding="utf-8")


def _append_json_record(path: Path, record: dict[str, Any]) -> None:
    records = _load_json_records(path)
    records.append(record)
    _save_json_records(path, records)


def _pop_json_record(path: Path, record_id: str) -> dict[str, Any] | None:
    records = _load_json_records(path)
    remaining: list[dict[str, Any]] = []
    matched: dict[str, Any] | None = None
    for record in records:
        if matched is None and str(record.get("id")) == record_id:
            matched = record
            continue
        remaining.append(record)
    if matched is not None:
        _save_json_records(path, remaining)
    return matched


def _find_json_record(path: Path, record_id: str) -> dict[str, Any] | None:
    for record in _load_json_records(path):
        if str(record.get("id")) == record_id:
            return record
    return None


def _load_chart_of_accounts() -> list[dict[str, Any]]:
    accounts = _load_json_records(CHART_OF_ACCOUNTS_PATH)
    if accounts:
        return accounts
    return [dict(account) for account in DEFAULT_CHART_OF_ACCOUNTS]


def _save_chart_of_accounts(accounts: list[dict[str, Any]]) -> None:
    _save_json_records(CHART_OF_ACCOUNTS_PATH, accounts)


def _ensure_chart_of_accounts() -> list[dict[str, Any]]:
    accounts = _load_chart_of_accounts()
    if not CHART_OF_ACCOUNTS_PATH.exists():
        _save_chart_of_accounts(accounts)
    return accounts


def _normalize_category_key(value: Any) -> str:
    return str(value or "").strip().lower()


def _find_account_by_code(accounts: list[dict[str, Any]], code: str) -> dict[str, Any] | None:
    for account in accounts:
        if str(account.get("code") or "") == code:
            return account
    return None


def _resolve_account_for_entity(entity_type: str, record: dict[str, Any], accounts: list[dict[str, Any]]) -> dict[str, Any]:
    if entity_type == "income":
        category_key = _normalize_category_key(record.get("Category"))
        account_code = INCOME_CATEGORY_ACCOUNT_MAP.get(category_key, "4900")
    elif entity_type == "expense":
        category_key = _normalize_category_key(record.get("Category"))
        account_code = EXPENSE_CATEGORY_ACCOUNT_MAP.get(category_key, "5000")
    elif entity_type == "invoice":
        account_code = "4000"
    elif entity_type == "payroll":
        account_code = "5300"
    else:
        account_code = "4900"

    account = _find_account_by_code(accounts, account_code)
    if account is not None:
        return account
    return {"code": account_code, "name": "Unmapped Account", "tax_treatment": "review"}


def _extract_transaction_amount(entity_type: str, record: dict[str, Any]) -> float:
    components = _extract_amount_components(entity_type, record)
    return components["total"]


def _extract_amount_components(entity_type: str, record: dict[str, Any]) -> dict[str, float]:
    if entity_type == "income":
        net_amount = _coerce_number(record.get("Amount (€)", 0))
        total_amount = _coerce_number(record.get("Total incl. VAT (€)", net_amount))
        if total_amount <= 0:
            total_amount = net_amount
        explicit_vat_amount = _coerce_number(record.get("VAT Amount (€)", 0))
        vat_amount = explicit_vat_amount if explicit_vat_amount > 0 else max(total_amount - net_amount, 0.0)
        return {
            "net": round(net_amount, 2),
            "vat": round(vat_amount, 2),
            "total": round(total_amount, 2),
        }

    if entity_type == "expense":
        net_amount = _coerce_number(record.get("Net Amount (€)", 0))
        total_amount = _coerce_number(record.get("Total (€)", net_amount))
        if total_amount <= 0:
            total_amount = net_amount
        explicit_vat_amount = _coerce_number(record.get("VAT Amount (€)", 0))
        vat_amount = explicit_vat_amount if explicit_vat_amount > 0 else max(total_amount - net_amount, 0.0)
        return {
            "net": round(net_amount, 2),
            "vat": round(vat_amount, 2),
            "total": round(total_amount, 2),
        }
    if entity_type == "invoice":
        net_amount = _coerce_number(record.get("Net (€)", 0))
        total_amount = _coerce_number(record.get("Total (€)", net_amount))
        if total_amount <= 0:
            total_amount = net_amount
        explicit_vat_amount = _coerce_number(record.get("VAT Amount (€)", 0))
        vat_amount = explicit_vat_amount if explicit_vat_amount > 0 else max(total_amount - net_amount, 0.0)
        return {
            "net": round(net_amount, 2),
            "vat": round(vat_amount, 2),
            "total": round(total_amount, 2),
        }
    if entity_type == "payroll":
        gross_pay = _coerce_number(record.get("Gross Pay (€)", 0))
        employer_prsi = _coerce_number(record.get("Employer PRSI (€)", 0))
        employer_cost = _coerce_number(record.get("Employer Cost (€)", gross_pay + employer_prsi))
        return {
            "net": round(gross_pay, 2),
            "vat": 0.0,
            "total": round(employer_cost, 2),
        }
    return {"net": 0.0, "vat": 0.0, "total": 0.0}


def _extract_transaction_date(entity_type: str, record: dict[str, Any]) -> str:
    if entity_type == "income":
        return str(record.get("Date") or "")
    if entity_type == "expense":
        return str(record.get("Date (Registered)") or "")
    if entity_type == "invoice":
        return str(record.get("Issue Date") or "")
    if entity_type == "payroll":
        return str(record.get("Pay Date") or "")
    return ""


def _extract_transaction_description(entity_type: str, record: dict[str, Any]) -> str:
    if entity_type == "income":
        return str(record.get("Description") or record.get("Client / Source") or "Income")
    if entity_type == "expense":
        return str(record.get("Description") or record.get("Title") or "Expense")
    if entity_type == "invoice":
        return str(record.get("Service / Product") or record.get("Invoice #") or "Invoice")
    if entity_type == "payroll":
        return str(record.get("Employee Name") or "Payroll")
    return str(record.get("title") or record.get("Description") or entity_type.title())


def _build_journal_lines(action: str, entity_type: str, account_code: str, total_amount: float, net_amount: float, vat_amount: float, *, record: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    if total_amount <= 0:
        return []

    if entity_type == "income":
        if action in {"archive", "cancel"}:
            if vat_amount > 0:
                return [
                    {"account_code": account_code, "debit": net_amount, "credit": 0.0},
                    {"account_code": "2100", "debit": vat_amount, "credit": 0.0},
                    {"account_code": "1000", "debit": 0.0, "credit": total_amount},
                ]
            return [
                {"account_code": account_code, "debit": total_amount, "credit": 0.0},
                {"account_code": "1000", "debit": 0.0, "credit": total_amount},
            ]
        if vat_amount > 0:
            return [
                {"account_code": "1000", "debit": total_amount, "credit": 0.0},
                {"account_code": account_code, "debit": 0.0, "credit": net_amount},
                {"account_code": "2100", "debit": 0.0, "credit": vat_amount},
            ]
        return [
            {"account_code": "1000", "debit": total_amount, "credit": 0.0},
            {"account_code": account_code, "debit": 0.0, "credit": total_amount},
        ]

    if entity_type == "expense":
        if action in {"archive", "cancel"}:
            if vat_amount > 0:
                return [
                    {"account_code": "1000", "debit": total_amount, "credit": 0.0},
                    {"account_code": account_code, "debit": 0.0, "credit": net_amount},
                    {"account_code": "1200", "debit": 0.0, "credit": vat_amount},
                ]
            return [
                {"account_code": "1000", "debit": total_amount, "credit": 0.0},
                {"account_code": account_code, "debit": 0.0, "credit": total_amount},
            ]
        if vat_amount > 0:
            return [
                {"account_code": account_code, "debit": net_amount, "credit": 0.0},
                {"account_code": "1200", "debit": vat_amount, "credit": 0.0},
                {"account_code": "1000", "debit": 0.0, "credit": total_amount},
            ]
        return [
            {"account_code": account_code, "debit": total_amount, "credit": 0.0},
            {"account_code": "1000", "debit": 0.0, "credit": total_amount},
        ]

    if entity_type == "invoice":
        if action in {"archive", "cancel"}:
            if vat_amount > 0:
                return [
                    {"account_code": account_code, "debit": net_amount, "credit": 0.0},
                    {"account_code": "2100", "debit": vat_amount, "credit": 0.0},
                    {"account_code": "1100", "debit": 0.0, "credit": total_amount},
                ]
            return [
                {"account_code": account_code, "debit": total_amount, "credit": 0.0},
                {"account_code": "1100", "debit": 0.0, "credit": total_amount},
            ]
        if vat_amount > 0:
            return [
                {"account_code": "1100", "debit": total_amount, "credit": 0.0},
                {"account_code": account_code, "debit": 0.0, "credit": net_amount},
                {"account_code": "2100", "debit": 0.0, "credit": vat_amount},
            ]
        return [
            {"account_code": "1100", "debit": total_amount, "credit": 0.0},
            {"account_code": account_code, "debit": 0.0, "credit": total_amount},
        ]

    if entity_type == "payroll":
        payroll_record = record or {}
        gross_pay = round(_coerce_number(payroll_record.get("Gross Pay (€)", net_amount)), 2)
        paye_amount = round(_coerce_number(payroll_record.get("PAYE (€)", 0)), 2)
        usc_amount = round(_coerce_number(payroll_record.get("USC (€)", 0)), 2)
        employee_prsi_amount = round(_coerce_number(payroll_record.get("Employee PRSI (€)", 0)), 2)
        employer_prsi_amount = round(_coerce_number(payroll_record.get("Employer PRSI (€)", 0)), 2)
        net_pay = round(_coerce_number(payroll_record.get("Net Pay (€)", gross_pay - paye_amount - usc_amount - employee_prsi_amount)), 2)
        liabilities = round(paye_amount + usc_amount + employee_prsi_amount + employer_prsi_amount, 2)

        if action in {"archive", "cancel"}:
            lines: list[dict[str, Any]] = [{"account_code": "1000", "debit": net_pay, "credit": 0.0}]
            if liabilities > 0:
                lines.append({"account_code": "2200", "debit": liabilities, "credit": 0.0})
            lines.append({"account_code": account_code, "debit": 0.0, "credit": gross_pay})
            if employer_prsi_amount > 0:
                lines.append({"account_code": "5310", "debit": 0.0, "credit": employer_prsi_amount})
            return lines

        lines = [{"account_code": account_code, "debit": gross_pay, "credit": 0.0}]
        if employer_prsi_amount > 0:
            lines.append({"account_code": "5310", "debit": employer_prsi_amount, "credit": 0.0})
        if liabilities > 0:
            lines.append({"account_code": "2200", "debit": 0.0, "credit": liabilities})
        lines.append({"account_code": "1000", "debit": 0.0, "credit": net_pay})
        return lines

    return []


def _find_account_name(accounts: list[dict[str, Any]], account_code: str) -> str:
    account = _find_account_by_code(accounts, account_code)
    if account is None:
        return "Unmapped Account"
    return str(account.get("name") or "Unmapped Account")


def _compute_trial_balance(ledger_entries: list[dict[str, Any]], accounts: list[dict[str, Any]]) -> dict[str, Any]:
    balances: dict[str, dict[str, Any]] = {}

    def ensure_row(code: str) -> dict[str, Any]:
        if code not in balances:
            balances[code] = {
                "account_code": code,
                "account_name": _find_account_name(accounts, code),
                "debit": 0.0,
                "credit": 0.0,
            }
        return balances[code]

    for entry in ledger_entries:
        lines = entry.get("journal_lines") if isinstance(entry.get("journal_lines"), list) else []
        if not lines:
            account_code = str(entry.get("account_code") or "")
            amount = _coerce_number(entry.get("amount_eur"))
            if not account_code or amount <= 0:
                continue
            fallback_lines = _build_journal_lines(
                str(entry.get("action") or "create"),
                str(entry.get("entity_type") or ""),
                account_code,
                amount,
                amount,
                0.0,
            )
            for line in fallback_lines:
                row = ensure_row(str(line.get("account_code") or ""))
                row["debit"] += _coerce_number(line.get("debit"))
                row["credit"] += _coerce_number(line.get("credit"))
            continue

        for line in lines:
            code = str(line.get("account_code") or "")
            if not code:
                continue
            row = ensure_row(code)
            row["debit"] += _coerce_number(line.get("debit"))
            row["credit"] += _coerce_number(line.get("credit"))

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for code in sorted(balances.keys()):
        row = balances[code]
        row["debit"] = round(row["debit"], 2)
        row["credit"] = round(row["credit"], 2)
        row["net"] = round(row["debit"] - row["credit"], 2)
        total_debit += row["debit"]
        total_credit += row["credit"]
        rows.append(row)

    total_debit = round(total_debit, 2)
    total_credit = round(total_credit, 2)
    return {
        "rows": rows,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "difference": round(total_debit - total_credit, 2),
        "is_balanced": abs(total_debit - total_credit) < 0.005,
    }


def _load_ledger_entries() -> list[dict[str, Any]]:
    return _load_json_records(LEDGER_JOURNAL_PATH)


def _record_ledger_entry(action: str, entity_type: str, record: dict[str, Any], *, source: str, row_number: int | None = None) -> None:
    if entity_type not in {"income", "expense", "invoice", "payroll"}:
        return

    accounts = _ensure_chart_of_accounts()
    account = _resolve_account_for_entity(entity_type, record, accounts)
    account_code = str(account.get("code") or "")
    amount_components = _extract_amount_components(entity_type, record)
    net_amount = round(amount_components["net"], 2)
    vat_amount = round(amount_components["vat"], 2)
    amount = round(amount_components["total"], 2)
    if not _is_vat_registered():
        vat_amount = 0.0
        net_amount = amount
    journal_lines = _build_journal_lines(action, entity_type, account_code, amount, net_amount, vat_amount, record=record)
    entry = {
        "id": str(uuid4()),
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "action": action,
        "entity_type": entity_type,
        "source": source,
        "row_number": row_number,
        "transaction_date": _extract_transaction_date(entity_type, record),
        "description": _extract_transaction_description(entity_type, record),
        "category": str(record.get("Category") or record.get("Payroll Period") or ""),
        "account_code": account_code,
        "account_name": str(account.get("name") or ""),
        "tax_treatment": str(account.get("tax_treatment") or ""),
        "vat_rate": str(record.get("VAT Rate") or "0%"),
        "vat_treatment": _normalize_vat_treatment(record.get("VAT Treatment"), record.get("VAT Rate")),
        "supply_type": _normalize_supply_type(record.get("Supply Type")),
        "net_amount_eur": net_amount,
        "vat_amount_eur": vat_amount,
        "total_amount_eur": amount,
        "amount_eur": amount,
        "journal_lines": journal_lines,
        "debit_total": round(sum(_coerce_number(line.get("debit")) for line in journal_lines), 2),
        "credit_total": round(sum(_coerce_number(line.get("credit")) for line in journal_lines), 2),
        "entry_balanced": abs(
            sum(_coerce_number(line.get("debit")) for line in journal_lines)
            - sum(_coerce_number(line.get("credit")) for line in journal_lines)
        )
        < 0.005,
        "phase_tag": str(record.get("Phase Tag") or ""),
    }
    _append_json_record(LEDGER_JOURNAL_PATH, entry)


def _vat_period_bounds(target_date: date) -> tuple[date, date, str]:
    if target_date.month in (1, 2):
        start_month = 1
    elif target_date.month in (3, 4):
        start_month = 3
    elif target_date.month in (5, 6):
        start_month = 5
    elif target_date.month in (7, 8):
        start_month = 7
    elif target_date.month in (9, 10):
        start_month = 9
    else:
        start_month = 11

    end_month = start_month + 1
    start_date = date(target_date.year, start_month, 1)
    end_day = monthrange(target_date.year, end_month)[1]
    end_date = date(target_date.year, end_month, end_day)
    label = f"{start_date.strftime('%b')} - {end_date.strftime('%b %Y')}"
    return start_date, end_date, label


def _compute_vat_control_summary(ledger_entries: list[dict[str, Any]]) -> dict[str, Any]:
    current_day = date.today()
    period_start, period_end, period_label = _vat_period_bounds(current_day)
    if period_end.month == 12:
        due_year = period_end.year + 1
        due_month = 1
    else:
        due_year = period_end.year
        due_month = period_end.month + 1
    due_date = date(due_year, due_month, 23).isoformat()
    summary = {
        "period_label": period_label,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "t1_output_vat": 0.0,
        "t2_input_vat": 0.0,
        "t3_net_vat": 0.0,
        "t4_refund": 0.0,
        "due_date": due_date,
        "zero_rated_sales": 0.0,
        "exempt_sales": 0.0,
        "reverse_charge_purchases": 0.0,
        "treatment_notes": "",
    }

    if not _is_vat_registered():
        return summary

    for entry in ledger_entries:
        transaction_date = _parse_transaction_date(entry.get("transaction_date"))
        if transaction_date is None or transaction_date < period_start or transaction_date > period_end:
            continue

        lines = entry.get("journal_lines") if isinstance(entry.get("journal_lines"), list) else []
        for line in lines:
            code = str(line.get("account_code") or "")
            debit = _coerce_number(line.get("debit"))
            credit = _coerce_number(line.get("credit"))
            if code == "2100":
                summary["t1_output_vat"] += max(credit - debit, 0.0)
            elif code == "1200":
                summary["t2_input_vat"] += max(debit - credit, 0.0)

        vat_treatment = _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate"))
        total_amount = _coerce_number(entry.get("total_amount_eur"))
        entity_type = str(entry.get("entity_type") or "")
        if entity_type in {"income", "invoice"} and vat_treatment == "zero_rated":
            summary["zero_rated_sales"] += total_amount
        elif entity_type in {"income", "invoice"} and vat_treatment == "exempt":
            summary["exempt_sales"] += total_amount
        elif entity_type == "expense" and vat_treatment == "reverse_charge":
            summary["reverse_charge_purchases"] += total_amount

    summary["t1_output_vat"] = round(summary["t1_output_vat"], 2)
    summary["t2_input_vat"] = round(summary["t2_input_vat"], 2)
    summary["t3_net_vat"] = round(max(summary["t1_output_vat"] - summary["t2_input_vat"], 0.0), 2)
    summary["t4_refund"] = round(max(summary["t2_input_vat"] - summary["t1_output_vat"], 0.0), 2)
    summary["zero_rated_sales"] = round(summary["zero_rated_sales"], 2)
    summary["exempt_sales"] = round(summary["exempt_sales"], 2)
    summary["reverse_charge_purchases"] = round(summary["reverse_charge_purchases"], 2)
    summary["treatment_notes"] = (
        f"Zero-rated sales €{summary['zero_rated_sales']:.2f}; "
        f"Exempt sales €{summary['exempt_sales']:.2f}; "
        f"Reverse-charge purchases €{summary['reverse_charge_purchases']:.2f}"
    )
    return summary


def _export_trial_balance_csv(trial_balance: dict[str, Any]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["account_code", "account_name", "debit_eur", "credit_eur", "net_eur"])
    for row in trial_balance.get("rows", []):
        writer.writerow([
            row.get("account_code", ""),
            row.get("account_name", ""),
            row.get("debit", 0),
            row.get("credit", 0),
            row.get("net", 0),
        ])
    writer.writerow([
        "TOTAL",
        "",
        trial_balance.get("total_debit", 0),
        trial_balance.get("total_credit", 0),
        trial_balance.get("difference", 0),
    ])
    return buffer.getvalue()


def _export_vat3_csv(vat_summary: dict[str, Any]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["VAT3_Period", "T1", "T2", "T3", "T4", "Due_Date", "ZeroRatedSales", "ExemptSales", "ReverseChargePurchases", "Treatment_Notes"])
    writer.writerow([
        vat_summary.get("period_label", ""),
        vat_summary.get("t1_output_vat", 0),
        vat_summary.get("t2_input_vat", 0),
        vat_summary.get("t3_net_vat", 0),
        vat_summary.get("t4_refund", 0),
        vat_summary.get("due_date", ""),
        vat_summary.get("zero_rated_sales", 0),
        vat_summary.get("exempt_sales", 0),
        vat_summary.get("reverse_charge_purchases", 0),
        vat_summary.get("treatment_notes", ""),
    ])
    return buffer.getvalue()


def _entry_vat_anomaly_flags(entry: dict[str, Any]) -> list[str]:
    vat_treatment = _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate"))
    vat_amount = _coerce_number(entry.get("vat_amount_eur"))
    total_amount = _coerce_number(entry.get("total_amount_eur"))
    vat_rate_ratio = _parse_vat_rate(entry.get("vat_rate"))
    flags: list[str] = []

    if vat_treatment in {"zero_rated", "exempt"} and vat_amount > 0.009:
        flags.append("non_zero_vat_with_zero_or_exempt_treatment")
    if vat_treatment == "reverse_charge" and vat_amount > 0.009:
        flags.append("reverse_charge_should_not_post_local_vat_amount")
    if vat_treatment == "standard" and total_amount > 0 and vat_rate_ratio > 0 and vat_amount <= 0.009:
        flags.append("missing_vat_amount_for_standard_rate")

    return flags


def _detect_vat_anomalies(ledger_entries: list[dict[str, Any]], *, limit: int = 20) -> list[dict[str, Any]]:
    anomalies: list[dict[str, Any]] = []
    for entry in sorted(ledger_entries, key=lambda item: str(item.get("timestamp") or ""), reverse=True):
        flags = _entry_vat_anomaly_flags(entry)
        if not flags:
            continue
        anomalies.append(
            {
                "timestamp": str(entry.get("timestamp") or ""),
                "entity_type": str(entry.get("entity_type") or ""),
                "description": str(entry.get("description") or ""),
                "transaction_date": str(entry.get("transaction_date") or ""),
                "amount_eur": round(_coerce_number(entry.get("amount_eur")), 2),
                "vat_rate": str(entry.get("vat_rate") or "0%"),
                "vat_treatment": _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate")),
                "supply_type": _normalize_supply_type(entry.get("supply_type")),
                "flags": flags,
            }
        )
        if len(anomalies) >= limit:
            break
    return anomalies


def _export_ledger_journal_csv(ledger_entries: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "timestamp",
            "action",
            "entity_type",
            "transaction_date",
            "description",
            "account_code",
            "account_name",
            "amount_eur",
            "debit_total",
            "credit_total",
            "phase_tag",
            "vat_rate",
            "vat_treatment",
            "supply_type",
            "vat_amount_eur",
            "net_amount_eur",
            "total_amount_eur",
            "anomaly_flags",
        ]
    )

    for entry in sorted(ledger_entries, key=lambda item: str(item.get("timestamp") or ""), reverse=True):
        flags = _entry_vat_anomaly_flags(entry)
        writer.writerow(
            [
                entry.get("timestamp", ""),
                entry.get("action", ""),
                entry.get("entity_type", ""),
                entry.get("transaction_date", ""),
                entry.get("description", ""),
                entry.get("account_code", ""),
                entry.get("account_name", ""),
                round(_coerce_number(entry.get("amount_eur")), 2),
                round(_coerce_number(entry.get("debit_total")), 2),
                round(_coerce_number(entry.get("credit_total")), 2),
                entry.get("phase_tag", ""),
                str(entry.get("vat_rate") or "0%"),
                _normalize_vat_treatment(entry.get("vat_treatment"), entry.get("vat_rate")),
                _normalize_supply_type(entry.get("supply_type")),
                round(_coerce_number(entry.get("vat_amount_eur")), 2),
                round(_coerce_number(entry.get("net_amount_eur")), 2),
                round(_coerce_number(entry.get("total_amount_eur")), 2),
                "|".join(flags),
            ]
        )
    return buffer.getvalue()


def _summarize_chart_of_accounts(accounts: list[dict[str, Any]]) -> dict[str, Any]:
    summary = {
        "total_accounts": len(accounts),
        "income_accounts": 0,
        "expense_accounts": 0,
        "asset_accounts": 0,
        "liability_accounts": 0,
        "equity_accounts": 0,
    }
    for account in accounts:
        account_type = str(account.get("type") or "").strip().lower()
        if account_type == "income":
            summary["income_accounts"] += 1
        elif account_type == "expense":
            summary["expense_accounts"] += 1
        elif account_type == "asset":
            summary["asset_accounts"] += 1
        elif account_type == "liability":
            summary["liability_accounts"] += 1
        elif account_type == "equity":
            summary["equity_accounts"] += 1
    return summary


def _record_audit(action: str, entity_type: str, details: dict[str, Any]) -> None:
    _append_json_record(
        AUDIT_LOG_PATH,
        {
            "id": str(uuid4()),
            "timestamp": datetime.now().isoformat(timespec="seconds"),
            "action": action,
            "entity_type": entity_type,
            "details": details,
        },
    )


def _archive_record(entity_type: str, record: dict[str, Any], *, source: str) -> None:
    sanitized_record = {key: value for key, value in record.items() if not key.startswith("__")}
    archive_entry = {
        "id": str(uuid4()),
        "archived_at": datetime.now().isoformat(timespec="seconds"),
        "entity_type": entity_type,
        "source": source,
        "record": sanitized_record,
    }
    _append_json_record(ARCHIVE_PATH, archive_entry)
    _record_audit("archive", entity_type, {"source": source, "record": sanitized_record})


def _load_archives() -> list[dict[str, Any]]:
    return _load_json_records(ARCHIVE_PATH)


def _load_audit_entries() -> list[dict[str, Any]]:
    return _load_json_records(AUDIT_LOG_PATH)


def _parse_validation_query(raw: Any) -> dict[str, str]:
    if not raw:
        return {}
    try:
        parsed = json.loads(str(raw))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(key): str(value) for key, value in parsed.items()}


def _parse_form_query(raw: Any) -> dict[str, Any]:
    if not raw:
        return {}
    try:
        parsed = json.loads(str(raw))
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _build_validation_message(errors: dict[str, str]) -> str:
    return f"Validation: {'; '.join(errors.values())}"


def _redirect_with_form_errors(route_name: str, form_data: dict[str, Any], validation_errors: dict[str, str], **extra_params: Any):
    return redirect(
        url_for(
            route_name,
            message=_build_validation_message(validation_errors),
            validation_errors=json.dumps(validation_errors),
            form_data=json.dumps(form_data),
            **extra_params,
        )
    )


def _validate_positive_amount(value: Any, field_name: str, label: str, errors: dict[str, str], *, allow_zero: bool = False) -> None:
    if value in (None, ""):
        return
    numeric_value = _coerce_number(value)
    if numeric_value < 0 or (not allow_zero and numeric_value == 0):
        comparator = "zero or greater" if allow_zero else "greater than zero"
        errors[field_name] = f"{label} must be {comparator}"


def _validate_required_text(value: Any, field_name: str, label: str, errors: dict[str, str]) -> None:
    if not str(value or "").strip():
        errors[field_name] = f"{label} is required"


def _validate_income_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Date"), "date", "Income date", errors)
    _validate_required_text(payload.get("Description"), "description", "Income description", errors)
    _validate_positive_amount(payload.get("Amount (€)"), "amount", "Income amount", errors)
    return errors


def _validate_expense_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Date (Registered)"), "date", "Expense date", errors)
    if not str(payload.get("Title") or "").strip() and not str(payload.get("Description") or "").strip():
        errors["title"] = "Expense title or description is required"
        errors["description"] = "Expense title or description is required"
    _validate_positive_amount(payload.get("Net Amount (€)"), "net_amount", "Expense net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Total (€)"), "total_amount", "Expense total amount", errors)
    _validate_positive_amount(payload.get("Base Net Amount (€)"), "base_net_amount", "Expense base net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Delivery (€)"), "delivery_amount", "Expense delivery amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Fees (€)"), "fees_amount", "Expense fees amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Other Charges (€)"), "other_charges_amount", "Expense other charges amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Discount (€)"), "discount_amount", "Expense discount amount", errors, allow_zero=True)
    subtotal_before_discount = (
        _coerce_number(payload.get("Base Net Amount (€)"))
        + _coerce_number(payload.get("Delivery (€)"))
        + _coerce_number(payload.get("Fees (€)"))
        + _coerce_number(payload.get("Other Charges (€)"))
    )
    if _coerce_number(payload.get("Discount (€)")) > subtotal_before_discount:
        errors["discount_amount"] = "Expense discount cannot exceed base net plus add-on charges"
    input_vat_reclaimable = _normalize_input_vat_reclaimable(payload.get("Input VAT Reclaimable"))
    if _is_vat_registered() and input_vat_reclaimable in {"Yes", "Partial"} and not str(payload.get("Supplier VAT Number") or "").strip():
        errors["supplier_vat_number"] = "Supplier VAT number is required when claiming input VAT"
    if str(payload.get("Deductibility Status") or "").strip() not in EXPENSE_DEDUCTIBILITY_OPTIONS:
        errors["deductibility_status"] = "Deductibility status is invalid"
    if str(payload.get("Receipt Attached") or "").strip() not in YES_NO_OPTIONS:
        errors["receipt_attached"] = "Receipt attached must be Yes or No"
    if str(payload.get("Bank Reconciliation") or "").strip() not in RECONCILIATION_OPTIONS:
        errors["bank_reconciliation"] = "Bank reconciliation must be Reconciled or Unreconciled"
    status = _normalize_expense_status(payload.get("Status"))
    if status not in EXPENSE_STATUS_OPTIONS:
        errors["status"] = "Expense status is invalid"
    if _is_paid_status("expense", status) and not str(payload.get("Payment Method") or "").strip():
        errors["payment_method"] = "Payment method is required when status is Paid"
    if _coerce_number(payload.get("Net Amount (€)")) > _coerce_number(payload.get("Total (€)")) and payload.get("Net Amount (€)") not in (None, ""):
        errors["total_amount"] = "Expense total amount must be at least the net amount"
    return errors


def _validate_invoice_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Invoice #"), "invoice_number", "Invoice number", errors)
    _validate_required_text(payload.get("Client Name"), "client_name", "Invoice client", errors)
    _validate_positive_amount(payload.get("Net (€)"), "net_amount", "Invoice net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Total (€)"), "total_amount", "Invoice total amount", errors)
    _validate_positive_amount(payload.get("Balance Due (€)"), "balance_due", "Invoice balance due", errors, allow_zero=True)
    issue_date = _parse_iso_date(payload.get("Issue Date"))
    due_date = _parse_iso_date(payload.get("Due Date"))
    if issue_date and due_date and due_date < issue_date:
        errors["due_date"] = "Invoice due date cannot be before the issue date"
    total_amount = _coerce_number(payload.get("Total (€)"))
    balance_due = _coerce_number(payload.get("Balance Due (€)"))
    if total_amount and balance_due > total_amount:
        errors["balance_due"] = "Invoice balance due cannot exceed the total amount"
    status = _normalize_invoice_status(payload.get("Status"))
    if status not in INVOICE_STATUS_OPTIONS:
        errors["status"] = "Invoice status is invalid"
    if _is_paid_status("invoice", status) and not str(payload.get("Payment Method") or "").strip():
        errors["payment_method"] = "Payment method is required when status is Paid"
    if _is_paid_status("invoice", status) and _parse_iso_date(payload.get("Payment Date")) is None:
        errors["payment_date"] = "Payment date is required when status is Paid"
    if not str(payload.get("Client VAT Number") or "").strip() and _is_vat_registered():
        errors["client_vat_number"] = "Client VAT number is required for VAT invoices"
    if not str(payload.get("Client Address") or "").strip() and _is_vat_registered():
        errors["client_address"] = "Client address is required for VAT invoices"
    return errors


def _validate_client_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Client Name"), "client_name", "Client name", errors)
    return errors


def _validate_supplier_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Supplier Name"), "supplier_name", "Supplier name", errors)
    return errors


def _validate_subscription_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("title"), "title", "Subscription name", errors)
    _validate_positive_amount(payload.get("net_amount"), "net_amount", "Subscription net amount", errors, allow_zero=True)
    _validate_positive_amount(payload.get("total_amount"), "total_amount", "Subscription total amount", errors)
    frequency = str(payload.get("frequency") or "").strip().lower()
    if frequency not in SUBSCRIPTION_FREQUENCIES:
        errors["frequency"] = "Subscription frequency is invalid"
    status = str(payload.get("status") or "").strip().lower()
    if status not in SUBSCRIPTION_STATUSES:
        errors["status"] = "Subscription status is invalid"
    start_date = _parse_iso_date(payload.get("start_date"))
    next_charge_date = _parse_iso_date(payload.get("next_charge_date") or payload.get("start_date"))
    end_date = _parse_iso_date(payload.get("end_date"))
    if start_date is None:
        errors["start_date"] = "Subscription start date is required"
    if next_charge_date is None:
        errors["next_charge_date"] = "Subscription next charge date is required"
    if start_date and next_charge_date and next_charge_date < start_date:
        errors["next_charge_date"] = "Subscription next charge date cannot be before the start date"
    if start_date and end_date and end_date < start_date:
        errors["end_date"] = "Subscription end date cannot be before the start date"
    return errors


def _validate_payroll_payload(payload: dict[str, Any]) -> dict[str, str]:
    errors: dict[str, str] = {}
    _validate_required_text(payload.get("Pay Date"), "pay_date", "Payroll pay date", errors)
    _validate_required_text(payload.get("Employee Name"), "employee_name", "Employee name", errors)
    _validate_positive_amount(payload.get("Gross Pay (€)"), "gross_pay", "Gross pay", errors)
    _validate_positive_amount(payload.get("PAYE (€)"), "paye", "PAYE", errors, allow_zero=True)
    _validate_positive_amount(payload.get("USC (€)"), "usc", "USC", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Employee PRSI (€)"), "employee_prsi", "Employee PRSI", errors, allow_zero=True)
    _validate_positive_amount(payload.get("Employer PRSI (€)"), "employer_prsi", "Employer PRSI", errors, allow_zero=True)

    gross_pay = _coerce_number(payload.get("Gross Pay (€)"))
    deductions = (
        _coerce_number(payload.get("PAYE (€)"))
        + _coerce_number(payload.get("USC (€)"))
        + _coerce_number(payload.get("Employee PRSI (€)"))
    )
    if deductions > gross_pay + 0.0001:
        errors["gross_pay"] = "Gross pay must be at least employee deductions total"

    pay_date = _parse_iso_date(payload.get("Pay Date"))
    payment_date = _parse_iso_date(payload.get("Payment Date"))
    if pay_date and payment_date and payment_date < pay_date:
        errors["payment_date"] = "Payment date cannot be before pay date"

    status = _normalize_payroll_status(payload.get("Status"))
    if status not in PAYROLL_STATUS_OPTIONS:
        errors["status"] = "Payroll status is invalid"
    if _is_paid_status("payroll", status) and not str(payload.get("Payment Method") or "").strip():
        errors["payment_method"] = "Payment method is required when status is Paid"
    if _is_paid_status("payroll", status) and payment_date is None:
        errors["payment_date"] = "Payment date is required when status is Paid"
    if str(payload.get("Bank Reconciliation") or "").strip() not in RECONCILIATION_OPTIONS:
        errors["bank_reconciliation"] = "Bank reconciliation must be Reconciled or Unreconciled"

    return errors


def _build_validation_state(active_tab: str) -> tuple[dict[str, str], dict[str, Any]]:
    expected_tab = request.args.get("validation_tab")
    if expected_tab != active_tab:
        return {}, {}
    return _parse_validation_query(request.args.get("validation_errors")), _parse_form_query(request.args.get("form_data"))


def _build_workbook_form_data(payload: dict[str, Any], field_map: dict[str, str]) -> dict[str, Any]:
    return {form_field: payload.get(sheet_field, "") for form_field, sheet_field in field_map.items()}


def _restore_workbook_archive(entity_type: str, archive_entry: dict[str, Any]) -> None:
    config = WORKBOOK_ENTITY_CONFIG[entity_type]
    row_number = _append_row_to_sheet(config["sheet"], archive_entry.get("record", {}))
    load_finance_data.cache_clear()
    _record_audit("restore", config["audit_type"], {"archive_id": archive_entry.get("id"), "row_number": row_number, "record": archive_entry.get("record", {})})
    _record_ledger_entry("restore", entity_type, archive_entry.get("record", {}), source="archive", row_number=row_number)


def _restore_subscription_archive(archive_entry: dict[str, Any]) -> None:
    subscriptions = _load_subscriptions()
    record = dict(archive_entry.get("record", {}))
    if not record.get("id"):
        record["id"] = str(uuid4())
    record["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    subscriptions.append(record)
    _save_subscriptions(subscriptions)
    _record_audit("restore", "subscription", {"archive_id": archive_entry.get("id"), "subscription_id": record.get("id"), "record": record})


def _restore_payroll_archive(archive_entry: dict[str, Any]) -> None:
    payroll_entries = _load_payroll_entries()
    record = dict(archive_entry.get("record", {}))
    if not record.get("id"):
        record["id"] = str(uuid4())
    record["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
    _normalize_payroll_payload(record)
    payroll_entries.append(record)
    _save_payroll_entries(payroll_entries)
    _record_audit("restore", "payroll", {"archive_id": archive_entry.get("id"), "payroll_id": record.get("id"), "record": record})
    _record_ledger_entry("restore", "payroll", record, source="archive", row_number=None)


def _collect_select_options(data: dict[str, Any], sheet_name: str, field_name: str) -> list[str]:
    rows = data.get("sheets", {}).get(sheet_name, [])
    options: list[str] = []
    for row in rows:
        value = row.get(field_name)
        if not value:
            continue
        value_str = str(value).strip()
        if value_str and value_str not in options:
            options.append(value_str)
    return options


def _parse_iso_date(value: Any) -> date | None:
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").date()
    except ValueError:
        return None


def _add_months(base_date: date, months: int) -> date:
    total_month = (base_date.month - 1) + months
    year = base_date.year + (total_month // 12)
    month = (total_month % 12) + 1
    day = min(base_date.day, monthrange(year, month)[1])
    return date(year, month, day)


def _load_subscriptions() -> list[dict[str, Any]]:
    if not SUBSCRIPTIONS_PATH.exists():
        return []

    try:
        records = json.loads(SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []

    today_iso = date.today().isoformat()
    subscriptions: list[dict[str, Any]] = []
    for record in records if isinstance(records, list) else []:
        if not isinstance(record, dict):
            continue
        frequency = str(record.get("frequency") or "monthly").strip().lower()
        status = str(record.get("status") or "active").strip().lower()
        start_date = _parse_iso_date(record.get("start_date"))
        next_charge_date = _parse_iso_date(record.get("next_charge_date"))
        last_posted_date = _parse_iso_date(record.get("last_posted_date"))
        end_date = _parse_iso_date(record.get("end_date"))
        normalized_start = (start_date or next_charge_date or date.today()).isoformat()
        subscriptions.append(
            {
                "id": str(record.get("id") or uuid4()),
                "title": str(record.get("title") or "").strip(),
                "description": str(record.get("description") or "").strip(),
                "supplier": str(record.get("supplier") or "").strip(),
                "category": str(record.get("category") or "").strip(),
                "net_amount": round(_coerce_number(record.get("net_amount")), 2),
                "total_amount": round(_coerce_number(record.get("total_amount")), 2),
                "frequency": frequency if frequency in SUBSCRIPTION_FREQUENCIES else "monthly",
                "start_date": normalized_start,
                "next_charge_date": (next_charge_date or start_date or date.today()).isoformat(),
                "last_posted_date": last_posted_date.isoformat() if last_posted_date else "",
                "end_date": end_date.isoformat() if end_date else "",
                "status": status if status in SUBSCRIPTION_STATUSES else "active",
                "notes": str(record.get("notes") or "").strip(),
                "created_at": str(record.get("created_at") or datetime.now().isoformat(timespec="seconds")),
                "last_updated_at": str(record.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )

    return subscriptions


def _save_subscriptions(subscriptions: list[dict[str, Any]]) -> None:
    payload = []
    for subscription in subscriptions:
        payload.append(
            {
                "id": subscription.get("id") or str(uuid4()),
                "title": str(subscription.get("title") or "").strip(),
                "description": str(subscription.get("description") or "").strip(),
                "supplier": str(subscription.get("supplier") or "").strip(),
                "category": str(subscription.get("category") or "").strip(),
                "net_amount": round(_coerce_number(subscription.get("net_amount")), 2),
                "total_amount": round(_coerce_number(subscription.get("total_amount")), 2),
                "frequency": subscription.get("frequency") if subscription.get("frequency") in SUBSCRIPTION_FREQUENCIES else "monthly",
                "start_date": (_parse_iso_date(subscription.get("start_date")) or date.today()).isoformat(),
                "next_charge_date": (_parse_iso_date(subscription.get("next_charge_date")) or date.today()).isoformat(),
                "last_posted_date": (_parse_iso_date(subscription.get("last_posted_date")) or None).isoformat() if _parse_iso_date(subscription.get("last_posted_date")) else "",
                "end_date": (_parse_iso_date(subscription.get("end_date")) or None).isoformat() if _parse_iso_date(subscription.get("end_date")) else "",
                "status": subscription.get("status") if subscription.get("status") in SUBSCRIPTION_STATUSES else "active",
                "notes": str(subscription.get("notes") or "").strip(),
                "created_at": str(subscription.get("created_at") or datetime.now().isoformat(timespec="seconds")),
                "last_updated_at": str(subscription.get("last_updated_at") or datetime.now().isoformat(timespec="seconds")),
            }
        )

    SUBSCRIPTIONS_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _build_subscription_expense_payload(subscription: dict[str, Any], charge_date: date) -> dict[str, Any]:
    title = subscription.get("title") or "Subscription"
    description = subscription.get("description") or title
    return {
        "Date (Registered)": charge_date.isoformat(),
        "Title": title,
        "Description": f"{description} (Subscription charge)",
        "Supplier / Payee": subscription.get("supplier") or title,
        "Category": subscription.get("category") or "Subscription",
        "Net Amount (€)": subscription.get("net_amount") or subscription.get("total_amount") or 0,
        "Total (€)": subscription.get("total_amount") or subscription.get("net_amount") or 0,
        "Input VAT Reclaimable": "No",
        "Deductibility Status": _normalize_deductibility_status("", subscription.get("category") or ""),
        "Capital Expenditure Flag": "No",
        "Receipt Attached": "No",
        "Bank Reconciliation": "Unreconciled",
        "Status": "Auto-posted",
        "Phase Tag": _resolve_phase_tag(charge_date.isoformat()),
    }


def _sync_subscriptions_to_expenses(today: date | None = None) -> dict[str, int]:
    current_day = today or date.today()
    subscriptions = _load_subscriptions()
    posted_count = 0
    changed = False

    for subscription in subscriptions:
        if subscription.get("status") != "active":
            continue

        next_charge = _parse_iso_date(subscription.get("next_charge_date")) or _parse_iso_date(subscription.get("start_date"))
        if next_charge is None:
            continue

        end_date = _parse_iso_date(subscription.get("end_date"))
        frequency_months = SUBSCRIPTION_FREQUENCIES.get(str(subscription.get("frequency")), 1)

        while next_charge <= current_day and (end_date is None or next_charge <= end_date):
            payload = _build_subscription_expense_payload(subscription, next_charge)
            row_number = _append_row_to_sheet("Expenses", payload)
            subscription["last_posted_date"] = next_charge.isoformat()
            subscription["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
            next_charge = _add_months(next_charge, frequency_months)
            subscription["next_charge_date"] = next_charge.isoformat()
            posted_count += 1
            changed = True
            _record_ledger_entry("sync_post", "expense", payload, source="subscription", row_number=row_number)

    if changed:
        _save_subscriptions(subscriptions)
        load_finance_data.cache_clear()

    active_count = sum(1 for subscription in subscriptions if subscription.get("status") == "active")
    return {"posted_count": posted_count, "active_count": active_count}


def _build_subscription_rows(subscriptions: list[dict[str, Any]], today: date | None = None) -> list[dict[str, Any]]:
    current_day = today or date.today()
    rows: list[dict[str, Any]] = []
    for subscription in subscriptions:
        next_charge = _parse_iso_date(subscription.get("next_charge_date"))
        status = str(subscription.get("status") or "active")
        days_until = (next_charge - current_day).days if next_charge else None
        if status != "active":
            due_label = status.title()
        elif days_until is None:
            due_label = "Unscheduled"
        elif days_until < 0:
            due_label = "Overdue"
        elif days_until == 0:
            due_label = "Due today"
        elif days_until <= 7:
            due_label = "Due soon"
        else:
            due_label = "Scheduled"

        rows.append(
            {
                **subscription,
                "next_charge_date": next_charge.isoformat() if next_charge else "",
                "last_posted_date": (_parse_iso_date(subscription.get("last_posted_date")) or None).isoformat() if _parse_iso_date(subscription.get("last_posted_date")) else "",
                "end_date": (_parse_iso_date(subscription.get("end_date")) or None).isoformat() if _parse_iso_date(subscription.get("end_date")) else "",
                "days_until": days_until,
                "due_label": due_label,
                "monthly_equivalent": round((_coerce_number(subscription.get("total_amount")) or _coerce_number(subscription.get("net_amount"))) / SUBSCRIPTION_FREQUENCIES.get(str(subscription.get("frequency")), 1), 2),
            }
        )

    rows.sort(key=lambda row: (row.get("status") != "active", row.get("next_charge_date") or "9999-12-31", row.get("title") or ""))
    return rows


def _summarize_subscriptions(subscription_rows: list[dict[str, Any]], today: date | None = None) -> dict[str, Any]:
    current_day = today or date.today()
    active_rows = [row for row in subscription_rows if row.get("status") == "active"]
    due_rows = [row for row in active_rows if row.get("next_charge_date") and (_parse_iso_date(row.get("next_charge_date")) or current_day) <= current_day]
    upcoming_rows = [row for row in active_rows if row.get("next_charge_date") and 0 <= ((_parse_iso_date(row.get("next_charge_date")) or current_day) - current_day).days <= 30]
    monthly_commitment = sum(_coerce_number(row.get("monthly_equivalent")) for row in active_rows)
    return {
        "active_count": len(active_rows),
        "due_count": len(due_rows),
        "upcoming_count": len(upcoming_rows),
        "monthly_commitment": monthly_commitment,
    }


def _build_chart_data(summary: dict[str, Any]) -> dict[str, float]:
    income_total = _coerce_number(summary.get("income_total", 0))
    expense_total = _coerce_number(summary.get("expense_total", 0))
    net_cashflow = _coerce_number(summary.get("net_cashflow", income_total - expense_total))
    max_chart_value = max(abs(income_total), abs(expense_total), 1.0)
    return {
        "income": income_total,
        "expense": expense_total,
        "net": net_cashflow,
        "income_width": min(100, abs(income_total) / max_chart_value * 100),
        "expense_width": min(100, abs(expense_total) / max_chart_value * 100),
        "net_width": min(100, abs(net_cashflow) / max_chart_value * 100),
    }


def _parse_row_number(value: Any) -> int | None:
    try:
        row_number = int(str(value).strip())
    except (TypeError, ValueError):
        return None
    return row_number if row_number >= 1 else None


def _find_row_by_number(rows: list[dict[str, Any]], row_number: Any) -> dict[str, Any] | None:
    resolved_row_number = _parse_row_number(row_number)
    if resolved_row_number is None:
        return None
    for row in rows:
        if row.get("__row_number") == resolved_row_number:
            return row
    return None


def _find_subscription_by_id(subscriptions: list[dict[str, Any]], subscription_id: Any) -> dict[str, Any] | None:
    resolved_id = str(subscription_id or "").strip()
    if not resolved_id:
        return None
    for subscription in subscriptions:
        if str(subscription.get("id")) == resolved_id:
            return subscription
    return None


def _find_sheet_row_or_raise(sheet_name: str, row_number: int) -> dict[str, Any]:
    data = load_finance_data()
    row = _find_row_by_number(data.get("sheets", {}).get(sheet_name, []), row_number)
    if row is None:
        raise ValueError(f"Could not find {sheet_name} row {row_number}")
    return row


def _try_parse_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("€", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _values_equivalent(left: Any, right: Any) -> bool:
    left_number = _try_parse_number(left)
    right_number = _try_parse_number(right)
    if left_number is not None and right_number is not None:
        return abs(left_number - right_number) < 0.000001

    left_text = str(left or "").strip()
    right_text = str(right or "").strip()
    return left_text == right_text


def _records_match(left: dict[str, Any], right: dict[str, Any], headers: list[str]) -> bool:
    for header in headers:
        if not _values_equivalent(left.get(header, ""), right.get(header, "")):
            return False
    return True


def _find_restore_conflict(entity_type: str, archive_entry: dict[str, Any]) -> dict[str, Any] | None:
    record = archive_entry.get("record", {}) if isinstance(archive_entry.get("record"), dict) else {}
    if entity_type == "subscription":
        for subscription in _load_subscriptions():
            if subscription.get("id") == record.get("id"):
                return subscription
            if all(
                str(subscription.get(field, "") or "").strip() == str(record.get(field, "") or "").strip()
                for field in ["title", "supplier", "frequency", "next_charge_date", "status"]
            ):
                return subscription
        return None

    if entity_type == "payroll":
        for payroll_entry in _load_payroll_entries():
            if payroll_entry.get("id") == record.get("id"):
                return payroll_entry
            if all(
                str(payroll_entry.get(field, "") or "").strip() == str(record.get(field, "") or "").strip()
                for field in ["Pay Date", "Employee Name", "Gross Pay (€)"]
            ):
                return payroll_entry
        return None

    config = WORKBOOK_ENTITY_CONFIG.get(entity_type)
    if not config:
        return None

    data = load_finance_data()
    rows = data.get("sheets", {}).get(config["sheet"], [])
    headers = SHEET_HEADERS.get(config["sheet"], [])
    for row in rows:
        if _records_match(row, record, headers):
            return row
    return None


def _export_audit_entries_csv(entries: list[dict[str, Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["timestamp", "action", "entity_type", "details_json"])
    for entry in entries:
        writer.writerow([
            entry.get("timestamp", ""),
            entry.get("action", ""),
            entry.get("entity_type", ""),
            json.dumps(entry.get("details", {}), ensure_ascii=True),
        ])
    return buffer.getvalue()


def _summarize_archives(archive_records: list[dict[str, Any]]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for record in archive_records:
        entity_type = str(record.get("entity_type") or "other")
        summary[entity_type] = summary.get(entity_type, 0) + 1
    return summary


def _build_sync_message(sync_result: dict[str, int]) -> str | None:
    posted_count = sync_result.get("posted_count", 0)
    if posted_count <= 0:
        return None
    if posted_count == 1:
        return "1 subscription charge was posted to expenses automatically."
    return f"{posted_count} subscription charges were posted to expenses automatically."


def _build_page_context(
    page_title: str,
    active_tab: str,
    data: dict[str, Any],
    *,
    income: list[dict[str, Any]] | None = None,
    expenses: list[dict[str, Any]] | None = None,
    invoices: list[dict[str, Any]] | None = None,
    clients: list[dict[str, Any]] | None = None,
    suppliers: list[dict[str, Any]] | None = None,
    subscriptions: list[dict[str, Any]] | None = None,
    payroll: list[dict[str, Any]] | None = None,
    subscription_summary: dict[str, Any] | None = None,
    payroll_summary: dict[str, Any] | None = None,
    chart_data: dict[str, float] | None = None,
    editing_income: dict[str, Any] | None = None,
    editing_expense: dict[str, Any] | None = None,
    editing_invoice: dict[str, Any] | None = None,
    editing_client: dict[str, Any] | None = None,
    editing_supplier: dict[str, Any] | None = None,
    editing_subscription: dict[str, Any] | None = None,
    editing_payroll: dict[str, Any] | None = None,
    income_form: dict[str, Any] | None = None,
    expense_form: dict[str, Any] | None = None,
    invoice_form: dict[str, Any] | None = None,
    client_form: dict[str, Any] | None = None,
    supplier_form: dict[str, Any] | None = None,
    subscription_form: dict[str, Any] | None = None,
    payroll_form: dict[str, Any] | None = None,
    validation_errors: dict[str, str] | None = None,
    archived_records: list[dict[str, Any]] | None = None,
    archive_summary: dict[str, int] | None = None,
    audit_entries: list[dict[str, Any]] | None = None,
    message: str | None = None,
    sync_message: str | None = None,
    error: str | None = None,
) -> dict[str, Any]:
    try:
        workbook_path = _resolve_workbook_path()
    except FileNotFoundError:
        workbook_path = WORKBOOK_PATH
    summary = data.get("summary", {})
    version = request.args.get("v", "20260729")
    business_profile = _load_business_profile()
    structure = _normalize_business_structure(business_profile.get("structure"))
    vat_threshold_basis = _normalize_vat_threshold_basis(business_profile.get("vat_threshold_basis"))
    income_rows_for_metrics = income if income is not None else data.get("sheets", {}).get("Income", [])
    invoice_rows_for_metrics = invoices if invoices is not None else data.get("sheets", {}).get("Invoices", [])
    phase_policy = _build_phase_policy(summary, structure)
    chart_of_accounts = _ensure_chart_of_accounts()
    ledger_entries = _load_ledger_entries()
    ledger_entries_sorted = sorted(ledger_entries, key=lambda item: str(item.get("timestamp") or ""), reverse=True)
    trial_balance = _compute_trial_balance(ledger_entries, chart_of_accounts)
    vat_control_summary = _compute_vat_control_summary(ledger_entries)
    vat_threshold_summary = _compute_vat_threshold_summary(income_rows_for_metrics, invoice_rows_for_metrics, vat_threshold_basis)
    vat_anomalies = _detect_vat_anomalies(ledger_entries)
    capital_assets = _load_capital_assets()
    capital_summary = _summarize_capital_assets(capital_assets)
    payroll_entries = payroll if payroll is not None else _load_payroll_entries()
    resolved_payroll_summary = payroll_summary or _summarize_payroll_entries(payroll_entries)
    bank_statement_lines = _load_bank_statement_lines()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    matched_statement_lines, unmatched_statement_lines = _match_bank_statement_lines(reconciliation_rows, bank_statement_lines)
    reconciliation_summary = _summarize_reconciliation(reconciliation_rows)
    reconciliation_summary["statement_line_count"] = len(matched_statement_lines)
    reconciliation_summary["unmatched_statement_count"] = len(unmatched_statement_lines)
    reconciliation_exceptions = [row for row in reconciliation_rows if row.get("exception_reasons")]
    return {
        "page_title": page_title,
        "summary": summary,
        "income": income or [],
        "expenses": expenses or [],
        "invoices": invoices or [],
        "clients": clients or [],
        "suppliers": suppliers or [],
        "subscriptions": subscriptions or [],
        "payroll": payroll_entries,
        "subscription_summary": subscription_summary or {"active_count": 0, "due_count": 0, "upcoming_count": 0, "monthly_commitment": 0.0},
        "payroll_summary": resolved_payroll_summary,
        "chart_data": chart_data or _build_chart_data(summary),
        "format_currency": _format_currency,
        "version": version,
        "active_tab": active_tab,
        "editing_income": editing_income,
        "editing_expense": editing_expense,
        "editing_invoice": editing_invoice,
        "editing_client": editing_client,
        "editing_supplier": editing_supplier,
        "editing_subscription": editing_subscription,
        "editing_payroll": editing_payroll,
        "income_form": income_form or {},
        "expense_form": expense_form or {},
        "invoice_form": invoice_form or {},
        "client_form": client_form or {},
        "supplier_form": supplier_form or {},
        "subscription_form": subscription_form or {},
        "payroll_form": payroll_form or {},
        "validation_errors": validation_errors or {},
        "archived_records": archived_records or [],
        "archive_summary": archive_summary or {},
        "audit_entries": audit_entries or [],
        "today_iso": date.today().isoformat(),
        "message": message,
        "sync_message": sync_message,
        "error": error,
        "income_clients": _collect_select_options(data, "Income", "Client / Source"),
        "income_categories": _collect_select_options(data, "Income", "Category"),
        "expense_suppliers": _collect_select_options(data, "Expenses", "Supplier / Payee"),
        "expense_categories": _collect_select_options(data, "Expenses", "Category"),
        "client_names": _collect_select_options(data, "Clients", "Client Name"),
        "supplier_names": _collect_select_options(data, "Suppliers", "Supplier Name"),
        "subscription_frequencies": list(SUBSCRIPTION_FREQUENCIES.keys()),
        "subscription_statuses": list(SUBSCRIPTION_STATUSES),
        "workbook_path": str(workbook_path.name),
        "workbook_status": "connected" if workbook_path.exists() else "missing",
        "business_profile": business_profile,
        "business_structure": structure,
        "vat_registered": bool(business_profile.get("vat_registered", True)),
        "vat_threshold_options": [
            {"value": key, "label": value["label"]}
            for key, value in VAT_TURNOVER_THRESHOLDS.items()
        ],
        "vat_threshold_summary": vat_threshold_summary,
        "phase_label": _phase_label_for_structure(structure),
        "phase_policy": phase_policy,
        "income_payment_methods": INCOME_PAYMENT_METHODS.get(structure, INCOME_PAYMENT_METHODS["sole_trader"]),
        "expense_payment_methods": EXPENSE_PAYMENT_METHODS.get(structure, EXPENSE_PAYMENT_METHODS["sole_trader"]),
        "vat_rate_options": VAT_RATE_OPTIONS,
        "vat_treatment_options": VAT_TREATMENT_OPTIONS,
        "supply_type_options": SUPPLY_TYPE_OPTIONS,
        "invoice_status_options": INVOICE_STATUS_OPTIONS,
        "expense_status_options": EXPENSE_STATUS_OPTIONS,
        "expense_input_vat_options": EXPENSE_INPUT_VAT_OPTIONS,
        "expense_deductibility_options": EXPENSE_DEDUCTIBILITY_OPTIONS,
        "reconciliation_options": RECONCILIATION_OPTIONS,
        "yes_no_options": YES_NO_OPTIONS,
        "payroll_status_options": PAYROLL_STATUS_OPTIONS,
        "chart_of_accounts": chart_of_accounts,
        "coa_summary": _summarize_chart_of_accounts(chart_of_accounts),
        "ledger_entries": ledger_entries_sorted[:80],
        "vat_anomalies": vat_anomalies,
        "vat_anomaly_count": len(vat_anomalies),
        "capital_summary": capital_summary,
        "capital_assets": capital_assets,
        "reconciliation_rows": reconciliation_rows,
        "reconciliation_summary": reconciliation_summary,
        "reconciliation_exceptions": reconciliation_exceptions,
        "bank_statement_lines": matched_statement_lines,
        "unmatched_statement_lines": unmatched_statement_lines,
        "ledger_summary": {
            "entries_count": len(ledger_entries),
            "posted_total": round(sum(_coerce_number(entry.get("amount_eur")) for entry in ledger_entries), 2),
            "debit_total": trial_balance["total_debit"],
            "credit_total": trial_balance["total_credit"],
        },
        "trial_balance": trial_balance,
        "vat_control_summary": vat_control_summary,
    }


@lru_cache(maxsize=1)
def load_finance_data() -> dict[str, Any]:
    sheets: dict[str, Any] = {}
    for sheet_name in ["Income", "Expenses", "Invoices", "Clients", "Suppliers"]:
        sheets[sheet_name] = _load_sheet_rows_with_row_numbers(sheet_name)

    income_total = sum(_coerce_number(row.get("Amount (€)", row.get("Total incl. VAT (€)", 0))) for row in sheets["Income"])
    expense_total = sum(_coerce_number(row.get("Total (€)", row.get("Net Amount (€)", 0))) for row in sheets["Expenses"])
    invoice_balance = sum(_coerce_number(row.get("Balance Due (€)", row.get("Balance (€)", 0))) for row in sheets["Invoices"])
    ap_balance = sum(
        _coerce_number(row.get("Total (€)", 0))
        for row in sheets["Expenses"]
        if not _is_paid_status("expense", row.get("Status"))
    )
    vat_balance = _compute_vat_control_summary(_load_ledger_entries()).get("t3_net_vat", 0.0)

    try:
        resolved_workbook_name = _resolve_workbook_path().name
    except FileNotFoundError:
        resolved_workbook_name = None

    return {
        "sheets": sheets,
        "workbook_path": resolved_workbook_name,
        "summary": {
            "income_total": income_total,
            "expense_total": expense_total,
            "net_cashflow": income_total - expense_total,
            "invoice_balance": invoice_balance,
            "ap_balance": ap_balance,
            "vat_balance": vat_balance,
        },
    }


def _read_workbook_sheet_rows(wb, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb:
        return []
    ws = wb[sheet_name]
    rows: list[dict[str, Any]] = []
    headers: list[Any] = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        cleaned = [_coerce_value(v) for v in row]
        if not header_found and _is_header_row(cleaned, sheet_name):
            headers = [_normalize_header_name(v, sheet_name) for v in cleaned]
            header_found = True
            continue
        if not header_found:
            continue
        if not any(v not in (None, "") for v in cleaned):
            continue
        rows.append(dict(zip(headers, cleaned)))
    return rows


def _migrate_transaction_sheets_from_workbook() -> None:
    """One-time seed of income/expenses/invoices/clients/suppliers JSON files from the
    legacy xlsm workbook, if present. Runs once at process start; after the JSON files
    exist, the workbook is never read for normal operation again."""
    missing_sheet_names = [name for name, path in SHEET_JSON_PATHS.items() if not path.exists()]
    if not missing_sheet_names:
        return

    try:
        resolved_path = _resolve_workbook_path()
    except FileNotFoundError:
        resolved_path = None

    seeded: dict[str, list[dict[str, Any]]] = {name: [] for name in missing_sheet_names}
    if resolved_path is not None and resolved_path.exists():
        try:
            wb = load_workbook(resolved_path, data_only=True, read_only=True)
            try:
                for sheet_name in missing_sheet_names:
                    seeded[sheet_name] = _read_workbook_sheet_rows(wb, sheet_name)
            finally:
                wb.close()
        except Exception:
            # Migration is best-effort; fall back to empty JSON files for any sheet
            # we couldn't read rather than blocking startup.
            pass

    for sheet_name in missing_sheet_names:
        _save_json_records(SHEET_JSON_PATHS[sheet_name], seeded[sheet_name])


@app.route("/")
def index():
    sync_result = _sync_subscriptions_to_expenses()
    data = load_finance_data()
    subscription_rows = _build_subscription_rows(_load_subscriptions())
    subscription_summary = _summarize_subscriptions(subscription_rows)
    if "error" in data:
        return render_template("index.html", **_build_page_context("Finance App", "dashboard", {}, error=data["error"], subscriptions=subscription_rows, subscription_summary=subscription_summary, sync_message=_build_sync_message(sync_result)))

    summary = data["summary"]
    income = data["sheets"].get("Income", [])[:8]
    expenses = data["sheets"].get("Expenses", [])[:8]
    invoices = data["sheets"].get("Invoices", [])[:8]
    clients = data["sheets"].get("Clients", [])[:8]
    suppliers = data["sheets"].get("Suppliers", [])[:8]
    return render_template(
        "index.html",
        **_build_page_context(
            "Financial Control Centre",
            "dashboard",
            data,
            income=income,
            expenses=expenses,
            invoices=invoices,
            clients=clients,
            suppliers=suppliers,
            subscriptions=subscription_rows,
            subscription_summary=subscription_summary,
            chart_data=_build_chart_data(summary),
            message=request.args.get("message"),
            sync_message=_build_sync_message(sync_result),
        ),
    )


@app.route("/healthz")
def healthz():
    return Response("ok", mimetype="text/plain")


@app.route("/income")
def income_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Income", [])
    validation_errors, income_form = _build_validation_state("income")
    editing_income = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Income", "income", data, income=rows, editing_income=editing_income, income_form=income_form, validation_errors=validation_errors, message=request.args.get("message")))


@app.route("/expenses")
def expenses_view():
    sync_result = _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Expenses", [])
    subscription_rows = _build_subscription_rows(_load_subscriptions())
    validation_errors, expense_form = _build_validation_state("expenses")
    editing_expense = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Expenses", "expenses", data, expenses=rows, subscription_summary=_summarize_subscriptions(subscription_rows), editing_expense=editing_expense, expense_form=expense_form, validation_errors=validation_errors, message=request.args.get("message"), sync_message=_build_sync_message(sync_result)))


@app.route("/invoices")
def invoices_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Invoices", [])
    validation_errors, invoice_form = _build_validation_state("invoices")
    editing_invoice = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Invoices", "invoices", data, invoices=rows, editing_invoice=editing_invoice, invoice_form=invoice_form, validation_errors=validation_errors, message=request.args.get("message")))


@app.route("/clients")
def clients_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Clients", [])
    validation_errors, client_form = _build_validation_state("clients")
    editing_client = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Clients", "clients", data, clients=rows, editing_client=editing_client, client_form=client_form, validation_errors=validation_errors, message=request.args.get("message")))


@app.route("/suppliers")
def suppliers_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    rows = data["sheets"].get("Suppliers", [])
    validation_errors, supplier_form = _build_validation_state("suppliers")
    editing_supplier = _find_row_by_number(rows, request.args.get("edit_row"))
    return render_template("index.html", **_build_page_context("Suppliers", "suppliers", data, suppliers=rows, editing_supplier=editing_supplier, supplier_form=supplier_form, validation_errors=validation_errors, message=request.args.get("message")))


@app.route("/subscriptions")
def subscriptions_view():
    sync_result = _sync_subscriptions_to_expenses()
    data = load_finance_data()
    subscription_rows = _build_subscription_rows(_load_subscriptions())
    subscription_summary = _summarize_subscriptions(subscription_rows)
    validation_errors, subscription_form = _build_validation_state("subscriptions")
    editing_subscription = _find_subscription_by_id(subscription_rows, request.args.get("edit_id"))
    return render_template(
        "index.html",
        **_build_page_context(
            "Subscriptions",
            "subscriptions",
            data if "error" not in data else {},
            subscriptions=subscription_rows,
            subscription_summary=subscription_summary,
            editing_subscription=editing_subscription,
            subscription_form=subscription_form,
            validation_errors=validation_errors,
            message=request.args.get("message"),
            sync_message=_build_sync_message(sync_result),
            error=data.get("error"),
        ),
    )


@app.route("/payroll")
def payroll_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    validation_errors, payroll_form = _build_validation_state("payroll")
    editing_payroll = _find_payroll_by_id(payroll_entries, request.args.get("edit_id"))
    return render_template(
        "index.html",
        **_build_page_context(
            "Payroll",
            "payroll",
            data if "error" not in data else {},
            payroll=payroll_entries,
            payroll_summary=_summarize_payroll_entries(payroll_entries),
            editing_payroll=editing_payroll,
            payroll_form=payroll_form,
            validation_errors=validation_errors,
            message=request.args.get("message"),
            error=data.get("error"),
        ),
    )


@app.route("/archive")
def archive_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    archive_records = sorted(_load_archives(), key=lambda record: str(record.get("archived_at") or ""), reverse=True)
    audit_entries = sorted(_load_audit_entries(), key=lambda record: str(record.get("timestamp") or ""), reverse=True)
    return render_template(
        "index.html",
        **_build_page_context(
            "Archive & Audit",
            "archive",
            data if "error" not in data else {},
            archived_records=archive_records[:50],
            archive_summary=_summarize_archives(archive_records),
            audit_entries=audit_entries[:50],
            error=data.get("error"),
            message=request.args.get("message"),
        ),
    )


@app.route("/ledger")
def ledger_view():
    _sync_subscriptions_to_expenses()
    data = load_finance_data()
    return render_template(
        "index.html",
        **_build_page_context(
            "Chart of Accounts & Ledger",
            "ledger",
            data if "error" not in data else {},
            error=data.get("error"),
            message=request.args.get("message"),
        ),
    )


@app.route("/ledger/trial-balance.csv")
def export_trial_balance_csv():
    accounts = _ensure_chart_of_accounts()
    ledger_entries = _load_ledger_entries()
    trial_balance = _compute_trial_balance(ledger_entries, accounts)
    csv_text = _export_trial_balance_csv(trial_balance)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=trial-balance.csv"},
    )


@app.route("/capital-allowances/export.csv")
def export_capital_allowances_csv():
    assets = _load_capital_assets()
    csv_text = _export_capital_allowances_csv(assets)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=capital-allowances.csv"},
    )


@app.route("/ledger/journal.csv")
def export_ledger_journal_csv():
    ledger_entries = _load_ledger_entries()
    csv_text = _export_ledger_journal_csv(ledger_entries)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=ledger-journal.csv"},
    )


@app.route("/vat3/export.csv")
def export_vat3_csv():
    vat_summary = _compute_vat_control_summary(_load_ledger_entries())
    csv_text = _export_vat3_csv(vat_summary)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=vat3-summary.csv"},
    )


@app.route("/audit/export.csv")
def export_audit_csv():
    entries = sorted(_load_audit_entries(), key=lambda record: str(record.get("timestamp") or ""), reverse=True)
    csv_text = _export_audit_entries_csv(entries)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=audit-log.csv"},
    )


@app.route("/payroll/export.csv")
def export_payroll_csv():
    entries = _load_payroll_entries()
    csv_text = _export_payroll_csv(entries)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=payroll-register.csv"},
    )


@app.route("/reconciliation/export.csv")
def export_reconciliation_csv():
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    rows = _build_reconciliation_rows(data, payroll_entries)
    _match_bank_statement_lines(rows, _load_bank_statement_lines())
    csv_text = _export_reconciliation_csv(rows)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=reconciliation-queue.csv"},
    )


@app.route("/reconciliation/exceptions.csv")
def export_reconciliation_exceptions_csv():
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    rows = _build_reconciliation_rows(data, payroll_entries)
    _match_bank_statement_lines(rows, _load_bank_statement_lines())
    csv_text = _export_reconciliation_csv(rows, exceptions_only=True)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=reconciliation-exceptions.csv"},
    )


@app.route("/reconciliation/bank-statements.csv")
def export_bank_statements_csv():
    statement_lines = _load_bank_statement_lines()
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    matched_lines, _ = _match_bank_statement_lines(reconciliation_rows, statement_lines)
    csv_text = _export_bank_statement_csv(matched_lines)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=bank-statements.csv"},
    )


@app.route("/reconciliation/unmatched-bank-statements.csv")
def export_unmatched_bank_statements_csv():
    statement_lines = _load_bank_statement_lines()
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    matched_lines, _ = _match_bank_statement_lines(reconciliation_rows, statement_lines)
    csv_text = _export_bank_statement_csv(matched_lines, unmatched_only=True)
    return Response(
        csv_text,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=unmatched-bank-statements.csv"},
    )


@app.route("/archive/restore", methods=["POST"])
def restore_archive_record():
    archive_id = str(request.form.get("archive_id") or "").strip()
    force_restore = str(request.form.get("force_restore") or "").strip().lower() in {"1", "true", "yes"}
    archive_entry = _find_json_record(ARCHIVE_PATH, archive_id)
    if archive_entry is None:
        return redirect(url_for("archive_view", message="Archive record could not be restored"))

    entity_type = str(archive_entry.get("entity_type") or "").strip()
    conflict = _find_restore_conflict(entity_type, archive_entry)
    if conflict is not None and not force_restore:
        _record_audit("restore_conflict", entity_type or "archive", {"archive_id": archive_id, "record": archive_entry.get("record", {}), "conflict": conflict})
        return redirect(url_for("archive_view", message=f"Restore conflict detected for {entity_type}. Use Force Restore if you want to restore it anyway."))

    archive_entry = _pop_json_record(ARCHIVE_PATH, archive_id)
    if archive_entry is None:
        return redirect(url_for("archive_view", message="Archive record could not be restored"))

    if entity_type == "subscription":
        _restore_subscription_archive(archive_entry)
    elif entity_type == "payroll":
        _restore_payroll_archive(archive_entry)
    elif entity_type in WORKBOOK_ENTITY_CONFIG:
        _restore_workbook_archive(entity_type, archive_entry)
    else:
        _append_json_record(ARCHIVE_PATH, archive_entry)
        return redirect(url_for("archive_view", message="Archive record type is not supported for restore"))

    return redirect(url_for("archive_view", message=f"{entity_type.title()} restored"))


@app.route("/refresh", methods=["POST"])
def refresh_workbook():
    load_finance_data.cache_clear()
    next_page = request.args.get("next", "/")
    return redirect(next_page or "/")


@app.route("/business-structure/update", methods=["POST"])
def update_business_structure():
    structure = _normalize_business_structure(request.form.get("structure"))
    transition_date = str(request.form.get("transition_date") or "").strip()
    vat_registered = str(request.form.get("vat_registered") or "").strip().lower() in {"1", "true", "yes", "on"}
    vat_threshold_basis = _normalize_vat_threshold_basis(request.form.get("vat_threshold_basis"))
    if transition_date and _parse_transaction_date(transition_date) is None:
        next_page = str(request.args.get("next") or "/")
        return redirect(_append_message_to_path(next_page, "Invalid transition date format"))

    profile = _load_business_profile()
    previous_profile = dict(profile)
    profile["structure"] = structure
    profile["transition_date"] = transition_date
    profile["vat_registered"] = vat_registered
    profile["vat_threshold_basis"] = vat_threshold_basis
    _save_business_profile(profile)
    _record_audit(
        "business_structure_update",
        "settings",
        {
            "from": {
                "structure": previous_profile.get("structure"),
                "transition_date": previous_profile.get("transition_date"),
                "vat_registered": bool(previous_profile.get("vat_registered", False)),
                "vat_threshold_basis": _normalize_vat_threshold_basis(previous_profile.get("vat_threshold_basis")),
            },
            "to": {
                "structure": structure,
                "transition_date": transition_date,
                "vat_registered": vat_registered,
                "vat_threshold_basis": vat_threshold_basis,
            },
        },
    )

    next_page = str(request.args.get("next") or "/")
    return redirect(_append_message_to_path(next_page, "Business structure updated"))


@app.route("/subscriptions/sync", methods=["POST"])
def sync_subscriptions():
    sync_result = _sync_subscriptions_to_expenses()
    posted_count = sync_result.get("posted_count", 0)
    if posted_count == 1:
        message = "1 subscription charge posted to expenses"
    elif posted_count > 1:
        message = f"{posted_count} subscription charges posted to expenses"
    else:
        message = "No subscription charges were due"
    _record_audit("sync", "subscriptions", {"posted_count": posted_count})
    return redirect(url_for("subscriptions_view", message=message))


@app.route("/reconciliation/import-statement", methods=["POST"])
def import_bank_statement():
    return_to = str(request.form.get("return_to") or "/ledger")
    statement_file = request.files.get("statement_file")
    if statement_file is None or not str(statement_file.filename or "").strip():
        return redirect(_append_message_to_path(return_to, "Bank statement import failed: file is required"))

    try:
        content = statement_file.read().decode("utf-8-sig", errors="ignore")
    except OSError:
        return redirect(_append_message_to_path(return_to, "Bank statement import failed: could not read file"))

    result = _ingest_bank_statement_csv(content, source_filename=str(statement_file.filename or "statement.csv"))
    imported_count = result.get("imported_count", 0)
    skipped_count = result.get("skipped_count", 0)
    _record_audit(
        "import",
        "bank_statement",
        {
            "source_filename": str(statement_file.filename or "statement.csv"),
            "imported_count": imported_count,
            "skipped_count": skipped_count,
        },
    )
    return redirect(_append_message_to_path(return_to, f"Bank statement imported: {imported_count} new lines, {skipped_count} skipped"))


@app.route("/reconciliation/apply-suggested", methods=["POST"])
def apply_suggested_reconciliation():
    return_to = str(request.form.get("return_to") or "/ledger")
    data = load_finance_data()
    payroll_entries = _load_payroll_entries()
    reconciliation_rows = _build_reconciliation_rows(data, payroll_entries)
    _match_bank_statement_lines(reconciliation_rows, _load_bank_statement_lines())

    duplicate_key_counts: dict[tuple[str, float, str], int] = {}
    for item in reconciliation_rows:
        if not bool(item.get("is_paid")):
            continue
        if item.get("bank_reconciliation") != "Unreconciled":
            continue
        item_key = (
            str(item.get("date") or ""),
            round(_coerce_number(item.get("amount_eur")), 2),
            str(item.get("payment_method") or "").strip().lower(),
        )
        if not item_key[0] or item_key[1] <= 0:
            continue
        duplicate_key_counts[item_key] = duplicate_key_counts.get(item_key, 0) + 1

    expenses_marked = 0
    invoices_marked = 0
    payroll_marked = 0
    workbook_changed = False
    payroll_changed = False

    for row in reconciliation_rows:
        if not bool(row.get("is_paid")):
            continue
        if row.get("bank_reconciliation") == "Reconciled":
            continue
        if int(row.get("statement_match_count") or 0) <= 0:
            continue
        row_key = (
            str(row.get("date") or ""),
            round(_coerce_number(row.get("amount_eur")), 2),
            str(row.get("payment_method") or "").strip().lower(),
        )
        if duplicate_key_counts.get(row_key, 0) > 1:
            continue
        # Batch apply should only reconcile unambiguous suggestions.
        if int(row.get("matching_group_size") or 1) > 1:
            continue
        if int(row.get("statement_match_count") or 0) != 1:
            continue

        entity_type = str(row.get("entity_type") or "")
        if entity_type in {"expense", "invoice"}:
            row_number = _parse_row_number(row.get("row_number"))
            if row_number is None:
                continue
            sheet_name = "Expenses" if entity_type == "expense" else "Invoices"
            record = _find_sheet_row_or_raise(sheet_name, row_number)
            if _normalize_reconciliation(record.get("Bank Reconciliation")) == "Reconciled":
                continue
            record["Bank Reconciliation"] = "Reconciled"
            _update_row_in_sheet(sheet_name, row_number, record)
            workbook_changed = True
            if entity_type == "expense":
                expenses_marked += 1
            else:
                invoices_marked += 1
            continue

        if entity_type == "payroll":
            payroll_id = str(row.get("payroll_id") or "")
            payroll_entry = _find_payroll_by_id(payroll_entries, payroll_id)
            if payroll_entry is None:
                continue
            if _normalize_reconciliation(payroll_entry.get("Bank Reconciliation")) == "Reconciled":
                continue
            payroll_entry["Bank Reconciliation"] = "Reconciled"
            payroll_entry["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
            payroll_changed = True
            payroll_marked += 1

    if workbook_changed:
        load_finance_data.cache_clear()
    if payroll_changed:
        _save_payroll_entries(payroll_entries)

    total_marked = expenses_marked + invoices_marked + payroll_marked
    _record_audit(
        "reconcile_batch",
        "reconciliation",
        {
            "expenses_marked": expenses_marked,
            "invoices_marked": invoices_marked,
            "payroll_marked": payroll_marked,
            "total_marked": total_marked,
        },
    )
    if total_marked == 0:
        return redirect(_append_message_to_path(return_to, "No suggested matches were available to apply"))
    return redirect(_append_message_to_path(return_to, f"Applied {total_marked} suggested reconciliation matches"))


@app.route("/reconciliation/mark", methods=["POST"])
def mark_reconciliation_status():
    entity_type = str(request.form.get("entity_type") or "").strip().lower()
    target_status = _normalize_reconciliation(request.form.get("bank_reconciliation"))
    return_to = str(request.form.get("return_to") or "/ledger")

    def _reconciled_mark_block_reason(record: dict[str, Any], entity: str) -> str | None:
        if target_status != "Reconciled":
            return None
        if not _is_paid_status(entity, record.get("Status")):
            return "Record must have a paid status before it can be reconciled"
        has_payment_method_column = "Payment Method" in record
        has_payment_date_column = "Payment Date" in record
        if has_payment_method_column and not str(record.get("Payment Method") or "").strip():
            return "Payment method is required before a paid record can be reconciled"
        if entity in {"invoice", "payroll"} and has_payment_date_column and _parse_iso_date(record.get("Payment Date")) is None:
            source_key = "Issue Date" if entity == "invoice" else "Pay Date"
            fallback_date = str(record.get(source_key) or "").strip()
            if _parse_iso_date(fallback_date) is None:
                return "Payment date is required before a paid record can be reconciled"
            record["Payment Date"] = fallback_date
        return None

    if entity_type in {"expense", "invoice"}:
        row_number = _parse_row_number(request.form.get("row_number"))
        if row_number is None:
            return redirect(_append_message_to_path(return_to, "Reconciliation update failed: missing row number"))
        sheet_name = "Expenses" if entity_type == "expense" else "Invoices"
        row = _find_sheet_row_or_raise(sheet_name, row_number)
        block_reason = _reconciled_mark_block_reason(row, entity_type)
        if block_reason:
            return redirect(_append_message_to_path(return_to, f"Reconciliation update failed: {block_reason}"))
        row["Bank Reconciliation"] = target_status
        _update_row_in_sheet(sheet_name, row_number, row)
        load_finance_data.cache_clear()
        _record_audit("reconcile", entity_type, {"row_number": row_number, "bank_reconciliation": target_status})
        return redirect(_append_message_to_path(return_to, f"{entity_type.title()} reconciliation updated"))

    if entity_type == "payroll":
        payroll_id = str(request.form.get("payroll_id") or "").strip()
        payroll_entries = _load_payroll_entries()
        payroll_entry = _find_payroll_by_id(payroll_entries, payroll_id)
        if payroll_entry is None:
            return redirect(_append_message_to_path(return_to, "Reconciliation update failed: payroll entry not found"))
        block_reason = _reconciled_mark_block_reason(payroll_entry, "payroll")
        if block_reason:
            return redirect(_append_message_to_path(return_to, f"Reconciliation update failed: {block_reason}"))
        payroll_entry["Bank Reconciliation"] = target_status
        payroll_entry["last_updated_at"] = datetime.now().isoformat(timespec="seconds")
        _save_payroll_entries(payroll_entries)
        _record_audit("reconcile", "payroll", {"payroll_id": payroll_id, "bank_reconciliation": target_status})
        return redirect(_append_message_to_path(return_to, "Payroll reconciliation updated"))

    return redirect(_append_message_to_path(return_to, "Reconciliation update failed: unsupported entity type"))


@app.route("/payroll/add", methods=["POST"])
def add_payroll():
    payload = {
        "id": str(uuid4()),
        "Pay Date": request.form.get("pay_date", ""),
        "Payroll Period": request.form.get("payroll_period", ""),
        "Employee Name": request.form.get("employee_name", ""),
        "Gross Pay (€)": request.form.get("gross_pay", ""),
        "PAYE (€)": request.form.get("paye", ""),
        "USC (€)": request.form.get("usc", ""),
        "Employee PRSI (€)": request.form.get("employee_prsi", ""),
        "Employer PRSI (€)": request.form.get("employer_prsi", ""),
        "Net Pay (€)": request.form.get("net_pay", ""),
        "Employer Cost (€)": request.form.get("employer_cost", ""),
        "Status": request.form.get("status", "Draft"),
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("pay_date", "")),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "last_updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    _normalize_payroll_payload(payload)
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "payroll", "Pay Date")
    validation_errors = _validate_payroll_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "payroll_view",
            {
                "pay_date": request.form.get("pay_date", ""),
                "payroll_period": request.form.get("payroll_period", ""),
                "employee_name": request.form.get("employee_name", ""),
                "gross_pay": request.form.get("gross_pay", ""),
                "paye": request.form.get("paye", ""),
                "usc": request.form.get("usc", ""),
                "employee_prsi": request.form.get("employee_prsi", ""),
                "employer_prsi": request.form.get("employer_prsi", ""),
                "status": request.form.get("status", "Draft"),
                "payment_method": request.form.get("payment_method", ""),
                "payment_date": request.form.get("payment_date", ""),
                "bank_reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
                "notes": request.form.get("notes", ""),
            },
            validation_errors,
            validation_tab="payroll",
        )

    payroll_entries = _load_payroll_entries()
    payroll_entries.append(payload)
    _save_payroll_entries(payroll_entries)
    _record_audit("create", "payroll", {"payroll_id": payload["id"], "record": payload})
    _record_ledger_entry("create", "payroll", payload, source="payroll", row_number=None)
    message = "Payroll entry added"
    if payment_date_autofilled:
        message = "Payroll entry added (payment date auto-filled from pay date)"
    return redirect(url_for("payroll_view", message=message))


@app.route("/payroll/update", methods=["POST"])
def update_payroll():
    payroll_id = str(request.form.get("payroll_id") or "").strip()
    payroll_entries = _load_payroll_entries()
    existing = _find_payroll_by_id(payroll_entries, payroll_id)
    if existing is None:
        return redirect(url_for("payroll_view", message="Payroll entry could not be updated"))

    payload = {
        "id": existing.get("id") or payroll_id,
        "Pay Date": request.form.get("pay_date", existing.get("Pay Date", "")),
        "Payroll Period": request.form.get("payroll_period", existing.get("Payroll Period", "")),
        "Employee Name": request.form.get("employee_name", existing.get("Employee Name", "")),
        "Gross Pay (€)": request.form.get("gross_pay", existing.get("Gross Pay (€)", "")),
        "PAYE (€)": request.form.get("paye", existing.get("PAYE (€)", "")),
        "USC (€)": request.form.get("usc", existing.get("USC (€)", "")),
        "Employee PRSI (€)": request.form.get("employee_prsi", existing.get("Employee PRSI (€)", "")),
        "Employer PRSI (€)": request.form.get("employer_prsi", existing.get("Employer PRSI (€)", "")),
        "Net Pay (€)": request.form.get("net_pay", existing.get("Net Pay (€)", "")),
        "Employer Cost (€)": request.form.get("employer_cost", existing.get("Employer Cost (€)", "")),
        "Status": request.form.get("status", existing.get("Status", "Draft")),
        "Payment Method": request.form.get("payment_method", existing.get("Payment Method", "")),
        "Payment Date": request.form.get("payment_date", existing.get("Payment Date", "")),
        "Bank Reconciliation": request.form.get("bank_reconciliation", existing.get("Bank Reconciliation", "Unreconciled")),
        "Notes": request.form.get("notes", existing.get("Notes", "")),
        "Phase Tag": _resolve_phase_tag(request.form.get("pay_date", existing.get("Pay Date", ""))),
        "created_at": str(existing.get("created_at") or datetime.now().isoformat(timespec="seconds")),
        "last_updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    _normalize_payroll_payload(payload)
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "payroll", "Pay Date")
    validation_errors = _validate_payroll_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "payroll_view",
            {
                "pay_date": payload.get("Pay Date", ""),
                "payroll_period": payload.get("Payroll Period", ""),
                "employee_name": payload.get("Employee Name", ""),
                "gross_pay": payload.get("Gross Pay (€)", ""),
                "paye": payload.get("PAYE (€)", ""),
                "usc": payload.get("USC (€)", ""),
                "employee_prsi": payload.get("Employee PRSI (€)", ""),
                "employer_prsi": payload.get("Employer PRSI (€)", ""),
                "status": payload.get("Status", "Draft"),
                "payment_method": payload.get("Payment Method", ""),
                "payment_date": payload.get("Payment Date", ""),
                "bank_reconciliation": payload.get("Bank Reconciliation", "Unreconciled"),
                "notes": payload.get("Notes", ""),
            },
            validation_errors,
            validation_tab="payroll",
            edit_id=payroll_id,
        )

    existing.update(payload)
    _save_payroll_entries(payroll_entries)
    _record_audit("update", "payroll", {"payroll_id": payroll_id, "record": existing})
    _record_ledger_entry("update", "payroll", existing, source="payroll", row_number=None)
    message = "Payroll entry updated"
    if payment_date_autofilled:
        message = "Payroll entry updated (payment date auto-filled from pay date)"
    return redirect(url_for("payroll_view", message=message))


@app.route("/payroll/delete", methods=["POST"])
def delete_payroll():
    payroll_id = str(request.form.get("payroll_id") or "").strip()
    payroll_entries = _load_payroll_entries()
    existing = _find_payroll_by_id(payroll_entries, payroll_id)
    if existing is None:
        return redirect(url_for("payroll_view", message="Payroll entry could not be removed"))

    _archive_record("payroll", existing, source="payroll")
    remaining = [entry for entry in payroll_entries if str(entry.get("id") or "") != payroll_id]
    _save_payroll_entries(remaining)
    _record_ledger_entry("archive", "payroll", existing, source="payroll", row_number=None)
    return redirect(url_for("payroll_view", message="Payroll entry archived"))


@app.route("/income/add", methods=["POST"])
def add_income():
    payload = {
        "Date": request.form.get("date", ""),
        "Description": request.form.get("description", ""),
        "Client / Source": request.form.get("client_source", ""),
        "Category": request.form.get("category", ""),
        "Invoice #": request.form.get("invoice_number", ""),
        "Amount (€)": request.form.get("amount", ""),
        "Total incl. VAT (€)": request.form.get("total_incl_vat", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Status": request.form.get("status", ""),
        "Payment Method": request.form.get("payment_method", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    _normalize_vat_fields(
        payload,
        net_key="Amount (€)",
        total_key="Total incl. VAT (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    validation_errors = _validate_income_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "income_view",
            _build_workbook_form_data(payload, {
                "date": "Date",
                "description": "Description",
                "client_source": "Client / Source",
                "category": "Category",
                "invoice_number": "Invoice #",
                "amount": "Amount (€)",
                "total_incl_vat": "Total incl. VAT (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "status": "Status",
                "payment_method": "Payment Method",
            }),
            validation_errors,
            validation_tab="income",
        )
    row_number = _append_row_to_sheet("Income", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "income", {"row_number": row_number, "record": payload})
    _record_ledger_entry("create", "income", payload, source="workbook", row_number=row_number)
    return redirect(url_for("income_view", message="Income entry added"))


@app.route("/income/update", methods=["POST"])
def update_income():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("income_view", message="Income entry could not be updated"))

    payload = {
        "Date": request.form.get("date", ""),
        "Description": request.form.get("description", ""),
        "Client / Source": request.form.get("client_source", ""),
        "Category": request.form.get("category", ""),
        "Invoice #": request.form.get("invoice_number", ""),
        "Amount (€)": request.form.get("amount", ""),
        "Total incl. VAT (€)": request.form.get("total_incl_vat", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Status": request.form.get("status", ""),
        "Payment Method": request.form.get("payment_method", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    _normalize_vat_fields(
        payload,
        net_key="Amount (€)",
        total_key="Total incl. VAT (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    validation_errors = _validate_income_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "income_view",
            _build_workbook_form_data(payload, {
                "date": "Date",
                "description": "Description",
                "client_source": "Client / Source",
                "category": "Category",
                "invoice_number": "Invoice #",
                "amount": "Amount (€)",
                "total_incl_vat": "Total incl. VAT (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "status": "Status",
                "payment_method": "Payment Method",
            }),
            validation_errors,
            validation_tab="income",
            edit_row=row_number,
        )
    _update_row_in_sheet("Income", row_number, payload)
    load_finance_data.cache_clear()
    _record_audit("update", "income", {"row_number": row_number, "record": payload})
    _record_ledger_entry("update", "income", payload, source="workbook", row_number=row_number)
    return redirect(url_for("income_view", message="Income entry updated"))


@app.route("/income/delete", methods=["POST"])
def delete_income():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("income_view", message="Income entry could not be removed"))

    row = _find_sheet_row_or_raise("Income", row_number)
    _archive_record("income", row, source="workbook")
    _delete_row_from_sheet("Income", row_number)
    load_finance_data.cache_clear()
    _record_ledger_entry("archive", "income", row, source="workbook", row_number=row_number)
    return redirect(url_for("income_view", message="Income entry archived"))


@app.route("/expenses/add", methods=["POST"])
def add_expense():
    payload = {
        "Date (Registered)": request.form.get("date", ""),
        "Title": request.form.get("title", ""),
        "Description": request.form.get("description", ""),
        "Supplier / Payee": request.form.get("supplier", ""),
        "Supplier VAT Number": request.form.get("supplier_vat_number", ""),
        "Receipt / Invoice Ref": request.form.get("receipt_reference", ""),
        "Category": request.form.get("category", ""),
        "Base Net Amount (€)": request.form.get("base_net_amount", request.form.get("net_amount", "")),
        "Delivery (€)": request.form.get("delivery_amount", ""),
        "Fees (€)": request.form.get("fees_amount", ""),
        "Other Charges (€)": request.form.get("other_charges_amount", ""),
        "Discount (€)": request.form.get("discount_amount", ""),
        "Net Amount (€)": request.form.get("net_amount", ""),
        "Total (€)": request.form.get("total_amount", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Input VAT Reclaimable": request.form.get("input_vat_reclaimable", "Yes"),
        "Deductibility Status": request.form.get("deductibility_status", ""),
        "Capital Expenditure Flag": request.form.get("capital_expenditure_flag", ""),
        "Receipt Attached": request.form.get("receipt_attached", "No"),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Status": _normalize_expense_status(request.form.get("status", "Pending")),
        "Payment Method": request.form.get("payment_method", ""),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    _form_total = payload.get("Total (€)", "")
    _form_vat = payload.get("VAT Amount (€)", "")
    _apply_expense_amount_breakdown(payload)
    payload["Total (€)"] = _form_total
    payload["VAT Amount (€)"] = _form_vat
    _normalize_vat_fields(
        payload,
        net_key="Net Amount (€)",
        total_key="Total (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    _apply_expense_compliance_fields(payload)
    validation_errors = _validate_expense_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "expenses_view",
            _build_workbook_form_data(payload, {
                "date": "Date (Registered)",
                "title": "Title",
                "description": "Description",
                "supplier": "Supplier / Payee",
                "supplier_vat_number": "Supplier VAT Number",
                "receipt_reference": "Receipt / Invoice Ref",
                "category": "Category",
                "base_net_amount": "Base Net Amount (€)",
                "delivery_amount": "Delivery (€)",
                "fees_amount": "Fees (€)",
                "other_charges_amount": "Other Charges (€)",
                "discount_amount": "Discount (€)",
                "net_amount": "Net Amount (€)",
                "total_amount": "Total (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "input_vat_reclaimable": "Input VAT Reclaimable",
                "deductibility_status": "Deductibility Status",
                "capital_expenditure_flag": "Capital Expenditure Flag",
                "receipt_attached": "Receipt Attached",
                "bank_reconciliation": "Bank Reconciliation",
                "status": "Status",
                "payment_method": "Payment Method",
                "notes": "Notes",
            }),
            validation_errors,
            validation_tab="expenses",
        )
    row_number = _append_row_to_sheet("Expenses", payload)
    _upsert_capital_asset_from_expense(payload, row_number, active=payload.get("Capital Expenditure Flag") == "Yes")
    load_finance_data.cache_clear()
    _record_audit("create", "expense", {"row_number": row_number, "record": payload})
    _record_ledger_entry("create", "expense", payload, source="workbook", row_number=row_number)
    return redirect(url_for("expenses_view", message="Expense entry added"))


@app.route("/expenses/update", methods=["POST"])
def update_expense():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("expenses_view", message="Expense could not be updated"))

    payload = {
        "Date (Registered)": request.form.get("date", ""),
        "Title": request.form.get("title", ""),
        "Description": request.form.get("description", ""),
        "Supplier / Payee": request.form.get("supplier", ""),
        "Supplier VAT Number": request.form.get("supplier_vat_number", ""),
        "Receipt / Invoice Ref": request.form.get("receipt_reference", ""),
        "Category": request.form.get("category", ""),
        "Base Net Amount (€)": request.form.get("base_net_amount", request.form.get("net_amount", "")),
        "Delivery (€)": request.form.get("delivery_amount", ""),
        "Fees (€)": request.form.get("fees_amount", ""),
        "Other Charges (€)": request.form.get("other_charges_amount", ""),
        "Discount (€)": request.form.get("discount_amount", ""),
        "Net Amount (€)": request.form.get("net_amount", ""),
        "Total (€)": request.form.get("total_amount", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Input VAT Reclaimable": request.form.get("input_vat_reclaimable", "Yes"),
        "Deductibility Status": request.form.get("deductibility_status", ""),
        "Capital Expenditure Flag": request.form.get("capital_expenditure_flag", ""),
        "Receipt Attached": request.form.get("receipt_attached", "No"),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Status": _normalize_expense_status(request.form.get("status", "Pending")),
        "Payment Method": request.form.get("payment_method", ""),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("date", "")),
    }
    _form_total = payload.get("Total (€)", "")
    _form_vat = payload.get("VAT Amount (€)", "")
    _apply_expense_amount_breakdown(payload)
    payload["Total (€)"] = _form_total
    payload["VAT Amount (€)"] = _form_vat
    _normalize_vat_fields(
        payload,
        net_key="Net Amount (€)",
        total_key="Total (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    _apply_expense_compliance_fields(payload)
    validation_errors = _validate_expense_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "expenses_view",
            _build_workbook_form_data(payload, {
                "date": "Date (Registered)",
                "title": "Title",
                "description": "Description",
                "supplier": "Supplier / Payee",
                "supplier_vat_number": "Supplier VAT Number",
                "receipt_reference": "Receipt / Invoice Ref",
                "category": "Category",
                "base_net_amount": "Base Net Amount (€)",
                "delivery_amount": "Delivery (€)",
                "fees_amount": "Fees (€)",
                "other_charges_amount": "Other Charges (€)",
                "discount_amount": "Discount (€)",
                "net_amount": "Net Amount (€)",
                "total_amount": "Total (€)",
                "vat_rate": "VAT Rate",
                "vat_amount": "VAT Amount (€)",
                "vat_treatment": "VAT Treatment",
                "supply_type": "Supply Type",
                "input_vat_reclaimable": "Input VAT Reclaimable",
                "deductibility_status": "Deductibility Status",
                "capital_expenditure_flag": "Capital Expenditure Flag",
                "receipt_attached": "Receipt Attached",
                "bank_reconciliation": "Bank Reconciliation",
                "status": "Status",
                "payment_method": "Payment Method",
                "notes": "Notes",
            }),
            validation_errors,
            validation_tab="expenses",
            edit_row=row_number,
        )
    _update_row_in_sheet("Expenses", row_number, payload)
    _upsert_capital_asset_from_expense(payload, row_number, active=payload.get("Capital Expenditure Flag") == "Yes")
    load_finance_data.cache_clear()

    _record_audit("update", "expense", {"row_number": row_number, "record": payload})
    _record_ledger_entry("update", "expense", payload, source="workbook", row_number=row_number)
    return Response(
        """
<!doctype html>
<html lang=\"en\">
    <head>
        <meta charset=\"utf-8\">
        <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
        <title>Expense Updated</title>
        <style>
            body { font-family: Segoe UI, Arial, sans-serif; margin: 28px; color: #0f172a; }
            .ok { color: #065f46; font-weight: 600; margin-bottom: 10px; }
            a { color: #1d4ed8; text-decoration: none; }
            a:hover { text-decoration: underline; }
        </style>
    </head>
    <body>
        <div class=\"ok\">Expense updated in app view.</div>
        <div><a href=\"/expenses\">Return to Expenses</a></div>
    </body>
</html>
        """,
        mimetype="text/html",
    )


@app.route("/expenses/delete", methods=["POST"])
def delete_expense():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("expenses_view", message="Expense could not be removed"))

    row = _find_sheet_row_or_raise("Expenses", row_number)
    try:
        _archive_record("expense", row, source="workbook")
        _delete_row_from_sheet("Expenses", row_number)
        _upsert_capital_asset_from_expense(row, row_number, active=False)
        load_finance_data.cache_clear()
        _record_ledger_entry("archive", "expense", row, source="workbook", row_number=row_number)
        return redirect(url_for("expenses_view", message="Expense archived"))
    except WorkbookWriteError as exc:
        return redirect(url_for("expenses_view", message=str(exc)))


@app.route("/invoices/add", methods=["POST"])
def add_invoice():
    existing_invoices = load_finance_data().get("sheets", {}).get("Invoices", [])
    generated_invoice_number = _next_invoice_number(request.form.get("issue_date", ""), existing_invoices)
    payload = {
        "Invoice #": generated_invoice_number,
        "Issue Date": request.form.get("issue_date", ""),
        "Due Date": request.form.get("due_date", ""),
        "Client Name": request.form.get("client_name", ""),
        "Client VAT Number": request.form.get("client_vat_number", ""),
        "Client Address": request.form.get("client_address", ""),
        "Service / Product": request.form.get("service_product", ""),
        "Net (€)": request.form.get("net_amount", ""),
        "Total (€)": request.form.get("total_amount", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Balance Due (€)": request.form.get("balance_due", ""),
        "Status": _normalize_invoice_status(request.form.get("status", "Draft")),
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("issue_date", "")),
    }
    _normalize_vat_fields(
        payload,
        net_key="Net (€)",
        total_key="Total (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "invoice", "Issue Date")
    validation_errors = _validate_invoice_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "invoices_view",
            {
                "invoice_number": request.form.get("invoice_number", ""),
                "issue_date": request.form.get("issue_date", ""),
                "due_date": request.form.get("due_date", ""),
                "client_name": request.form.get("client_name", ""),
                "client_vat_number": request.form.get("client_vat_number", ""),
                "client_address": request.form.get("client_address", ""),
                "service_product": request.form.get("service_product", ""),
                "net_amount": request.form.get("net_amount", ""),
                "total_amount": request.form.get("total_amount", ""),
                "vat_rate": request.form.get("vat_rate", "0%"),
                "vat_amount": request.form.get("vat_amount", ""),
                "vat_treatment": request.form.get("vat_treatment", "standard"),
                "supply_type": request.form.get("supply_type", "services"),
                "balance_due": request.form.get("balance_due", ""),
                "status": _normalize_invoice_status(request.form.get("status", "Draft")),
                "payment_method": request.form.get("payment_method", ""),
                "payment_date": request.form.get("payment_date", ""),
                "bank_reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
                "notes": request.form.get("notes", ""),
            },
            validation_errors,
            validation_tab="invoices",
        )
    row_number = _append_row_to_sheet("Invoices", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "invoice", {"row_number": row_number, "record": payload})
    _record_ledger_entry("create", "invoice", payload, source="workbook", row_number=row_number)
    message = "Invoice added"
    if payment_date_autofilled:
        message = "Invoice added (payment date auto-filled from issue date)"
    return redirect(url_for("invoices_view", message=message))


@app.route("/invoices/update", methods=["POST"])
def update_invoice():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("invoices_view", message="Invoice could not be updated"))

    current_row = _find_sheet_row_or_raise("Invoices", row_number)
    payload = {
        "Invoice #": str(current_row.get("Invoice #") or request.form.get("invoice_number", "")),
        "Issue Date": request.form.get("issue_date", ""),
        "Due Date": request.form.get("due_date", ""),
        "Client Name": request.form.get("client_name", ""),
        "Client VAT Number": request.form.get("client_vat_number", ""),
        "Client Address": request.form.get("client_address", ""),
        "Service / Product": request.form.get("service_product", ""),
        "Net (€)": request.form.get("net_amount", ""),
        "Total (€)": request.form.get("total_amount", ""),
        "VAT Rate": request.form.get("vat_rate", "0%"),
        "VAT Amount (€)": request.form.get("vat_amount", ""),
        "VAT Treatment": request.form.get("vat_treatment", "standard"),
        "Supply Type": request.form.get("supply_type", "services"),
        "Balance Due (€)": request.form.get("balance_due", ""),
        "Status": _normalize_invoice_status(request.form.get("status", "Draft")),
        "Payment Method": request.form.get("payment_method", ""),
        "Payment Date": request.form.get("payment_date", ""),
        "Bank Reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
        "Notes": request.form.get("notes", ""),
        "Phase Tag": _resolve_phase_tag(request.form.get("issue_date", "")),
    }
    _normalize_vat_fields(
        payload,
        net_key="Net (€)",
        total_key="Total (€)",
        vat_rate_key="VAT Rate",
        vat_amount_key="VAT Amount (€)",
        vat_registered=_is_vat_registered(),
    )
    _apply_vat_classification(payload, vat_rate_key="VAT Rate")
    payment_date_autofilled = _apply_default_payment_date_for_paid(payload, "invoice", "Issue Date")
    validation_errors = _validate_invoice_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "invoices_view",
            {
                "invoice_number": request.form.get("invoice_number", ""),
                "issue_date": request.form.get("issue_date", ""),
                "due_date": request.form.get("due_date", ""),
                "client_name": request.form.get("client_name", ""),
                "client_vat_number": request.form.get("client_vat_number", ""),
                "client_address": request.form.get("client_address", ""),
                "service_product": request.form.get("service_product", ""),
                "net_amount": request.form.get("net_amount", ""),
                "total_amount": request.form.get("total_amount", ""),
                "vat_rate": request.form.get("vat_rate", "0%"),
                "vat_amount": request.form.get("vat_amount", ""),
                "vat_treatment": request.form.get("vat_treatment", "standard"),
                "supply_type": request.form.get("supply_type", "services"),
                "balance_due": request.form.get("balance_due", ""),
                "status": _normalize_invoice_status(request.form.get("status", "Draft")),
                "payment_method": request.form.get("payment_method", ""),
                "payment_date": request.form.get("payment_date", ""),
                "bank_reconciliation": request.form.get("bank_reconciliation", "Unreconciled"),
                "notes": request.form.get("notes", ""),
            },
            validation_errors,
            validation_tab="invoices",
            edit_row=row_number,
        )
    _update_row_in_sheet("Invoices", row_number, payload)
    load_finance_data.cache_clear()
    _record_audit("update", "invoice", {"row_number": row_number, "record": payload})
    _record_ledger_entry("update", "invoice", payload, source="workbook", row_number=row_number)
    message = "Invoice updated"
    if payment_date_autofilled:
        message = "Invoice updated (payment date auto-filled from issue date)"
    return redirect(url_for("invoices_view", message=message))


@app.route("/invoices/delete", methods=["POST"])
def delete_invoice():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("invoices_view", message="Invoice could not be removed"))

    row = _find_sheet_row_or_raise("Invoices", row_number)
    if _normalize_invoice_status(row.get("Status")) != "Cancelled":
        row["Status"] = "Cancelled"
    _update_row_in_sheet("Invoices", row_number, row)
    load_finance_data.cache_clear()
    _record_audit("cancel", "invoice", {"row_number": row_number, "record": row})
    _record_ledger_entry("cancel", "invoice", row, source="workbook", row_number=row_number)
    return redirect(url_for("invoices_view", message="Invoice cancelled and retained for audit trail"))


@app.route("/subscriptions/add", methods=["POST"])
def add_subscription():
    start_date_value = request.form.get("start_date", "")
    start_date = _parse_iso_date(start_date_value) or date.today()
    subscription = {
        "id": str(uuid4()),
        "title": request.form.get("title", "").strip(),
        "description": request.form.get("description", "").strip(),
        "supplier": request.form.get("supplier", "").strip(),
        "category": request.form.get("category", "").strip(),
        "net_amount": request.form.get("net_amount", ""),
        "total_amount": request.form.get("total_amount", ""),
        "frequency": request.form.get("frequency", "monthly").strip().lower(),
        "start_date": start_date.isoformat(),
        "next_charge_date": start_date.isoformat(),
        "last_posted_date": "",
        "end_date": request.form.get("end_date", "").strip(),
        "status": request.form.get("status", "active").strip().lower(),
        "notes": request.form.get("notes", "").strip(),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "last_updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    subscription["next_charge_date"] = request.form.get("next_charge_date", "").strip() or start_date.isoformat()
    validation_errors = _validate_subscription_payload(subscription)
    if validation_errors:
        return _redirect_with_form_errors(
            "subscriptions_view",
            {
                "title": request.form.get("title", "").strip(),
                "description": request.form.get("description", "").strip(),
                "supplier": request.form.get("supplier", "").strip(),
                "category": request.form.get("category", "").strip(),
                "frequency": request.form.get("frequency", "monthly").strip().lower(),
                "start_date": subscription["start_date"],
                "next_charge_date": subscription["next_charge_date"],
                "end_date": request.form.get("end_date", "").strip(),
                "net_amount": request.form.get("net_amount", ""),
                "total_amount": request.form.get("total_amount", ""),
                "status": request.form.get("status", "active").strip().lower(),
                "notes": request.form.get("notes", "").strip(),
            },
            validation_errors,
            validation_tab="subscriptions",
        )
    subscriptions = _load_subscriptions()
    subscriptions.append(subscription)
    _save_subscriptions(subscriptions)
    _record_audit("create", "subscription", {"subscription_id": subscription["id"], "record": subscription})
    return redirect(url_for("subscriptions_view", message="Subscription added"))


@app.route("/subscriptions/update", methods=["POST"])
def update_subscription():
    subscription_id = request.form.get("subscription_id", "").strip()
    subscriptions = _load_subscriptions()
    existing = _find_subscription_by_id(subscriptions, subscription_id)
    if existing is None:
        return redirect(url_for("subscriptions_view", message="Subscription could not be updated"))

    start_date = _parse_iso_date(request.form.get("start_date", "")) or _parse_iso_date(existing.get("start_date")) or date.today()
    next_charge_date = _parse_iso_date(request.form.get("next_charge_date", "")) or _parse_iso_date(existing.get("next_charge_date")) or start_date
    end_date = _parse_iso_date(request.form.get("end_date", ""))

    payload = {
        "title": request.form.get("title", "").strip(),
        "description": request.form.get("description", "").strip(),
        "supplier": request.form.get("supplier", "").strip(),
        "category": request.form.get("category", "").strip(),
        "net_amount": request.form.get("net_amount", ""),
        "total_amount": request.form.get("total_amount", ""),
        "frequency": request.form.get("frequency", "monthly").strip().lower(),
        "start_date": start_date.isoformat(),
        "next_charge_date": next_charge_date.isoformat(),
        "end_date": end_date.isoformat() if end_date else "",
        "status": request.form.get("status", "active").strip().lower(),
        "notes": request.form.get("notes", "").strip(),
    }
    validation_errors = _validate_subscription_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "subscriptions_view",
            {
                "title": payload["title"],
                "description": payload["description"],
                "supplier": payload["supplier"],
                "category": payload["category"],
                "frequency": payload["frequency"],
                "start_date": payload["start_date"],
                "next_charge_date": payload["next_charge_date"],
                "end_date": payload["end_date"],
                "net_amount": payload["net_amount"],
                "total_amount": payload["total_amount"],
                "status": payload["status"],
                "notes": payload["notes"],
            },
            validation_errors,
            validation_tab="subscriptions",
            edit_id=subscription_id,
        )

    existing.update(
        {
            **payload,
            "last_updated_at": datetime.now().isoformat(timespec="seconds"),
        }
    )
    _save_subscriptions(subscriptions)
    _record_audit("update", "subscription", {"subscription_id": subscription_id, "record": existing})
    return redirect(url_for("subscriptions_view", message="Subscription updated"))


@app.route("/subscriptions/delete", methods=["POST"])
def delete_subscription():
    subscription_id = request.form.get("subscription_id", "").strip()
    subscriptions = _load_subscriptions()
    existing = _find_subscription_by_id(subscriptions, subscription_id)
    if existing is None:
        return redirect(url_for("subscriptions_view", message="Subscription could not be removed"))

    _archive_record("subscription", existing, source="subscriptions")
    remaining = [subscription for subscription in subscriptions if str(subscription.get("id")) != subscription_id]
    _save_subscriptions(remaining)
    return redirect(url_for("subscriptions_view", message="Subscription archived"))


@app.route("/clients/add", methods=["POST"])
def add_client():
    payload = {
        "Client Name": request.form.get("client_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
    }
    validation_errors = _validate_client_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "clients_view",
            {
                "client_name": request.form.get("client_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
            },
            validation_errors,
            validation_tab="clients",
        )
    row_number = _append_row_to_sheet("Clients", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "client", {"row_number": row_number, "record": payload})
    return redirect(url_for("clients_view", message="Client added"))


@app.route("/clients/update", methods=["POST"])
def update_client():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("clients_view", message="Client could not be updated"))

    payload = {
        "Client Name": request.form.get("client_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
    }
    validation_errors = _validate_client_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "clients_view",
            {
                "client_name": request.form.get("client_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
            },
            validation_errors,
            validation_tab="clients",
            edit_row=row_number,
        )
    _update_row_in_sheet("Clients", row_number, payload)
    load_finance_data.cache_clear()
    _record_audit("update", "client", {"row_number": row_number, "record": payload})
    return redirect(url_for("clients_view", message="Client updated"))


@app.route("/clients/delete", methods=["POST"])
def delete_client():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("clients_view", message="Client could not be removed"))

    row = _find_sheet_row_or_raise("Clients", row_number)
    _archive_record("client", row, source="workbook")
    _delete_row_from_sheet("Clients", row_number)
    load_finance_data.cache_clear()
    return redirect(url_for("clients_view", message="Client archived"))


@app.route("/suppliers/add", methods=["POST"])
def add_supplier():
    payload = {
        "Supplier Name": request.form.get("supplier_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
        "Default VAT Treatment": request.form.get("default_vat_treatment", ""),
    }
    validation_errors = _validate_supplier_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "suppliers_view",
            {
                "supplier_name": request.form.get("supplier_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
                "default_vat_treatment": request.form.get("default_vat_treatment", ""),
            },
            validation_errors,
            validation_tab="suppliers",
        )
    row_number = _append_row_to_sheet("Suppliers", payload)
    load_finance_data.cache_clear()
    _record_audit("create", "supplier", {"row_number": row_number, "record": payload})
    return redirect(url_for("suppliers_view", message="Supplier added"))


@app.route("/suppliers/update", methods=["POST"])
def update_supplier():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("suppliers_view", message="Supplier could not be updated"))

    payload = {
        "Supplier Name": request.form.get("supplier_name", ""),
        "Contact Person": request.form.get("contact_person", ""),
        "Email": request.form.get("email", ""),
        "Phone": request.form.get("phone", ""),
        "Country": request.form.get("country", ""),
        "Default VAT Treatment": request.form.get("default_vat_treatment", ""),
    }
    validation_errors = _validate_supplier_payload(payload)
    if validation_errors:
        return _redirect_with_form_errors(
            "suppliers_view",
            {
                "supplier_name": request.form.get("supplier_name", ""),
                "contact_person": request.form.get("contact_person", ""),
                "email": request.form.get("email", ""),
                "phone": request.form.get("phone", ""),
                "country": request.form.get("country", ""),
                "default_vat_treatment": request.form.get("default_vat_treatment", ""),
            },
            validation_errors,
            validation_tab="suppliers",
            edit_row=row_number,
        )
    _update_row_in_sheet("Suppliers", row_number, payload)
    load_finance_data.cache_clear()
    _record_audit("update", "supplier", {"row_number": row_number, "record": payload})
    return redirect(url_for("suppliers_view", message="Supplier updated"))


@app.route("/suppliers/delete", methods=["POST"])
def delete_supplier():
    row_number = _parse_row_number(request.form.get("row_number"))
    if row_number is None:
        return redirect(url_for("suppliers_view", message="Supplier could not be removed"))

    row = _find_sheet_row_or_raise("Suppliers", row_number)
    _archive_record("supplier", row, source="workbook")
    _delete_row_from_sheet("Suppliers", row_number)
    load_finance_data.cache_clear()
    return redirect(url_for("suppliers_view", message="Supplier archived"))


@app.route("/export/xlsm", methods=["POST"])
def export_xlsm():
    next_page = str(request.args.get("next") or "/")
    try:
        resolved_path = _resolve_workbook_path()
    except FileNotFoundError as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: {exc}"))

    try:
        wb = load_workbook(resolved_path, data_only=False, keep_links=False)
    except Exception as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: could not open the workbook ({exc})"))

    try:
        for sheet_name in SHEET_JSON_PATHS:
            if sheet_name not in wb:
                continue
            ws = wb[sheet_name]
            header_row_number = _find_header_row_number(ws, sheet_name)
            if header_row_number is None:
                continue
            headers = _get_header_row(ws, sheet_name)
            normalized_headers = [_normalize_header_name(header, sheet_name) for header in headers]

            if ws.max_row > header_row_number:
                ws.delete_rows(header_row_number + 1, ws.max_row - header_row_number)

            for record in _load_sheet_records_raw(sheet_name):
                row_values = [record.get(header, record.get(str(header), "")) for header in normalized_headers]
                ws.append(row_values)

        _save_workbook_atomic(wb, resolved_path)
    except WorkbookWriteError as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: {exc}"))
    except Exception as exc:
        return redirect(_append_message_to_path(next_page, f"Export failed: {exc}"))
    finally:
        wb.close()

    return redirect(_append_message_to_path(next_page, "Exported current data to Excel workbook"))


_migrate_transaction_sheets_from_workbook()


if __name__ == "__main__":
    app.run(
        debug=False,
        use_reloader=False,
        host="127.0.0.1",
        port=5000,
    )
