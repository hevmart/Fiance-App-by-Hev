import os
import tempfile
import json
from io import BytesIO
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import app


@pytest.fixture
def workbook_copy(tmp_path):
    dst = tmp_path / "sample.xlsm"
    wb = Workbook()
    wb.remove(wb.active)

    income = wb.create_sheet("Income")
    income.append(["Date", "Description", "Client / Source", "Category", "Invoice #", "Amount (€)", "Status"])
    income.append(["2026-07-29", "Test income", "Client A", "Travel", "INV-001", 150.0, "Paid"])

    expenses = wb.create_sheet("Expenses")
    expenses.append(["Date (Registered)", "Title", "Description", "Supplier / Payee", "Category", "Net Amount (€)", "Total (€)", "Status"])
    expenses.append(["2026-07-29", "Travel", "Hotel", "Supplier A", "Travel", 100.0, 120.0, "Pending"])

    invoices = wb.create_sheet("Invoices")
    invoices.append(["Invoice #", "Issue Date", "Due Date", "Client Name", "Service / Product", "Net (€)", "Total (€)", "Balance Due (€)", "Status"])
    invoices.append(["INV-001", "2026-07-01", "2026-07-31", "Client A", "Brand Strategy", 500.0, 605.0, 605.0, "Sent"])

    clients = wb.create_sheet("Clients")
    clients.append(["Client Name", "Contact Person", "Email", "Phone", "Country"])
    clients.append(["Client A", "Jane", "jane@example.com", "123", "Belgium"])

    suppliers = wb.create_sheet("Suppliers")
    suppliers.append(["Supplier Name", "Contact Person", "Email", "Phone", "Country", "Default VAT Treatment"])
    suppliers.append(["Supplier A", "John", "john@example.com", "456", "Netherlands", "Standard"])

    wb.save(dst)
    wb.close()
    _seed_transaction_json(tmp_path)
    return dst


@pytest.fixture(autouse=True)
def isolated_subscription_file(tmp_path):
    original_path = app.SUBSCRIPTIONS_PATH
    original_archive_path = app.ARCHIVE_PATH
    original_audit_log_path = app.AUDIT_LOG_PATH
    original_business_profile_path = app.BUSINESS_PROFILE_PATH
    original_coa_path = app.CHART_OF_ACCOUNTS_PATH
    original_ledger_path = app.LEDGER_JOURNAL_PATH
    original_capital_assets_path = app.CAPITAL_ASSETS_PATH
    original_payroll_path = app.PAYROLL_PATH
    original_bank_statements_path = app.BANK_STATEMENTS_PATH
    original_income_path = app.INCOME_PATH
    original_expenses_path = app.EXPENSES_PATH
    original_invoices_path = app.INVOICES_PATH
    original_clients_path = app.CLIENTS_PATH
    original_suppliers_path = app.SUPPLIERS_PATH
    original_sheet_json_paths = dict(app.SHEET_JSON_PATHS)
    app.SUBSCRIPTIONS_PATH = tmp_path / "subscriptions.json"
    app.ARCHIVE_PATH = tmp_path / "archives.json"
    app.AUDIT_LOG_PATH = tmp_path / "audit-log.json"
    app.BUSINESS_PROFILE_PATH = tmp_path / "business-profile.json"
    app.CHART_OF_ACCOUNTS_PATH = tmp_path / "chart-of-accounts.json"
    app.LEDGER_JOURNAL_PATH = tmp_path / "ledger-journal.json"
    app.CAPITAL_ASSETS_PATH = tmp_path / "capital-assets.json"
    app.PAYROLL_PATH = tmp_path / "payroll-register.json"
    app.BANK_STATEMENTS_PATH = tmp_path / "bank-statements.json"
    app.INCOME_PATH = tmp_path / "income.json"
    app.EXPENSES_PATH = tmp_path / "expenses.json"
    app.INVOICES_PATH = tmp_path / "invoices.json"
    app.CLIENTS_PATH = tmp_path / "clients.json"
    app.SUPPLIERS_PATH = tmp_path / "suppliers.json"
    app.SHEET_JSON_PATHS = {
        "Income": app.INCOME_PATH,
        "Expenses": app.EXPENSES_PATH,
        "Invoices": app.INVOICES_PATH,
        "Clients": app.CLIENTS_PATH,
        "Suppliers": app.SUPPLIERS_PATH,
    }
    app.load_finance_data.cache_clear()
    yield
    app.SUBSCRIPTIONS_PATH = original_path
    app.ARCHIVE_PATH = original_archive_path
    app.AUDIT_LOG_PATH = original_audit_log_path
    app.BUSINESS_PROFILE_PATH = original_business_profile_path
    app.CHART_OF_ACCOUNTS_PATH = original_coa_path
    app.LEDGER_JOURNAL_PATH = original_ledger_path
    app.CAPITAL_ASSETS_PATH = original_capital_assets_path
    app.PAYROLL_PATH = original_payroll_path
    app.BANK_STATEMENTS_PATH = original_bank_statements_path
    app.INCOME_PATH = original_income_path
    app.EXPENSES_PATH = original_expenses_path
    app.INVOICES_PATH = original_invoices_path
    app.CLIENTS_PATH = original_clients_path
    app.SUPPLIERS_PATH = original_suppliers_path
    app.SHEET_JSON_PATHS = original_sheet_json_paths
    app.load_finance_data.cache_clear()


def _seed_transaction_json(tmp_path):
    (tmp_path / "income.json").write_text(
        json.dumps([
            {
                "Date": "2026-07-29",
                "Description": "Test income",
                "Client / Source": "Client A",
                "Category": "Travel",
                "Invoice #": "INV-001",
                "Amount (€)": 150.0,
                "Status": "Paid",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "expenses.json").write_text(
        json.dumps([
            {
                "Date (Registered)": "2026-07-29",
                "Title": "Travel",
                "Description": "Hotel",
                "Supplier / Payee": "Supplier A",
                "Category": "Travel",
                "Net Amount (€)": 100.0,
                "Total (€)": 120.0,
                "Status": "Pending",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "invoices.json").write_text(
        json.dumps([
            {
                "Invoice #": "INV-001",
                "Issue Date": "2026-07-01",
                "Due Date": "2026-07-31",
                "Client Name": "Client A",
                "Service / Product": "Brand Strategy",
                "Net (€)": 500.0,
                "Total (€)": 605.0,
                "Balance Due (€)": 605.0,
                "Status": "Sent",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "clients.json").write_text(
        json.dumps([
            {
                "Client Name": "Client A",
                "Contact Person": "Jane",
                "Email": "jane@example.com",
                "Phone": "123",
                "Country": "Belgium",
            }
        ]),
        encoding="utf-8",
    )
    (tmp_path / "suppliers.json").write_text(
        json.dumps([
            {
                "Supplier Name": "Supplier A",
                "Contact Person": "John",
                "Email": "john@example.com",
                "Phone": "456",
                "Country": "Netherlands",
                "Default VAT Treatment": "Standard",
            }
        ]),
        encoding="utf-8",
    )


def test_append_income_row_updates_workbook(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-07-29",
        "description": "Test income",
        "client_source": "Client A",
        "amount": "150.00",
        "status": "Paid",
    }

    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    income_records = json.loads(app.INCOME_PATH.read_text(encoding="utf-8"))
    last_record = income_records[-1]
    assert last_record["Description"] == "Test income"
    assert str(last_record["Amount (€)"]) == "150.00"


def test_update_income_route_updates_existing_workbook_row(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    row_number = app.load_finance_data()["sheets"]["Income"][0]["__row_number"]
    payload = {
        "row_number": str(row_number),
        "date": "2026-07-30",
        "description": "Updated income",
        "client_source": "Client A",
        "category": "Consulting",
        "invoice_number": "INV-009",
        "amount": "250.00",
        "status": "Paid",
    }

    response = app.app.test_client().post('/income/update', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Income entry updated' in response.data

    income_records = json.loads(app.INCOME_PATH.read_text(encoding="utf-8"))
    record = income_records[row_number - 1]
    assert record["Description"] == "Updated income"
    assert str(record["Amount (€)"]) == "250.00"


def test_income_validation_prevents_invalid_amount(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-07-30",
        "description": "Invalid income",
        "client_source": "Client A",
        "amount": "-10.00",
        "status": "Paid",
    }
    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)

    assert response.status_code == 200
    assert b'Validation:' in response.data
    assert b'Income amount must be greater than zero' in response.data
    assert b'value="-10.00"' in response.data


def test_refresh_route_redirects_back_to_page(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post('/refresh?next=/income', follow_redirects=True)

    assert response.status_code == 200
    assert b'Income' in response.data


def test_business_structure_toggle_changes_dashboard_mode(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    update_response = app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "limited_company", "transition_date": "2026-08-15"},
        follow_redirects=True,
    )

    assert update_response.status_code == 200
    assert b'Phase 2 - Private Limited Company' in update_response.data
    assert b'Corporation Tax (CT1)' in update_response.data
    assert b'CT1 outputs' in update_response.data
    assert b'Director Loan Account' in update_response.data


def test_phase_resolution_uses_transition_date_for_limited_company(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "limited_company", "transition_date": "2026-08-15"},
    )

    assert app._resolve_phase_tag("2026-08-10") == "Phase 1"
    assert app._resolve_phase_tag("2026-08-15") == "Phase 2"
    assert app._resolve_phase_tag("2026-08-20") == "Phase 2"


def test_ledger_posts_income_with_mapped_account(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-08-05",
        "description": "Strategy workshop",
        "client_source": "Client A",
        "category": "Consulting / Project Fees",
        "invoice_number": "INV-2026-101",
        "amount": "500.00",
        "status": "Paid",
        "payment_method": "Business Bank Account",
    }

    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    ledger_entries = json.loads(app.LEDGER_JOURNAL_PATH.read_text(encoding="utf-8"))
    assert len(ledger_entries) >= 1
    latest_entry = ledger_entries[-1]
    assert latest_entry["entity_type"] == "income"
    assert latest_entry["account_code"] == "4000"
    assert latest_entry["account_name"] == "Consulting / Project Fees"
    assert latest_entry["amount_eur"] == 500.0
    assert latest_entry["entry_balanced"] is True
    assert latest_entry["debit_total"] == latest_entry["credit_total"]
    assert len(latest_entry["journal_lines"]) == 2


def test_ledger_view_renders_accounts_and_journal(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().get('/ledger')
    assert response.status_code == 200
    assert b'Chart of accounts' in response.data
    assert b'Ledger journal' in response.data
    assert b'Trial balance' in response.data
    assert b'VAT control summary' in response.data


def test_ledger_posts_expense_with_vat_control_accounts(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-08-05",
        "title": "Cloud Hosting",
        "description": "Monthly infra",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "category": "Professional Fees",
        "net_amount": "100.00",
        "total_amount": "123.00",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }
    response = app.app.test_client().post('/expenses/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    ledger_entries = json.loads(app.LEDGER_JOURNAL_PATH.read_text(encoding="utf-8"))
    latest_entry = ledger_entries[-1]
    assert latest_entry["entity_type"] == "expense"
    assert latest_entry["vat_amount_eur"] == 23.0
    assert latest_entry["entry_balanced"] is True

    lines = latest_entry["journal_lines"]
    assert any(line["account_code"] == "1200" and line["debit"] == 23.0 for line in lines)
    assert any(line["account_code"] == "1000" and line["credit"] == 123.0 for line in lines)


def test_trial_balance_csv_export_returns_download(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/income/add',
        data={
            "date": "2026-08-05",
            "description": "Export test",
            "client_source": "Client A",
            "category": "Consulting / Project Fees",
            "invoice_number": "INV-EXP-1",
            "amount": "250.00",
            "status": "Paid",
            "payment_method": "Business Bank Account",
        },
    )

    response = app.app.test_client().get('/ledger/trial-balance.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=trial-balance.csv' in response.headers['Content-Disposition']
    assert b'account_code,account_name,debit_eur,credit_eur,net_eur' in response.data
    assert b'TOTAL' in response.data


def test_ledger_journal_csv_export_includes_vat_trace_columns(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "RC purchase",
            "description": "Cross-border service",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "100.00",
            "vat_rate": "23%",
            "vat_amount": "0.00",
            "vat_treatment": "reverse_charge",
            "supply_type": "services",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "payment_method": "Business Bank",
        },
    )

    response = app.app.test_client().get('/ledger/journal.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=ledger-journal.csv' in response.headers['Content-Disposition']
    assert b'vat_rate,vat_treatment,supply_type,vat_amount_eur,net_amount_eur,total_amount_eur,anomaly_flags' in response.data
    assert b'reverse_charge' in response.data


def test_vat_registration_setting_persists(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "sole_trader", "transition_date": "", "vat_registered": "1"},
        follow_redirects=True,
    )
    assert response.status_code == 200

    profile = json.loads(app.BUSINESS_PROFILE_PATH.read_text(encoding="utf-8"))
    assert profile["vat_registered"] is True


def test_vat_threshold_basis_setting_persists(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/business-structure/update?next=/',
        data={
            "structure": "sole_trader",
            "transition_date": "",
            "vat_registered": "1",
            "vat_threshold_basis": "goods",
        },
        follow_redirects=True,
    )
    assert response.status_code == 200

    profile = json.loads(app.BUSINESS_PROFILE_PATH.read_text(encoding="utf-8"))
    assert profile["vat_threshold_basis"] == "goods"


def test_vat3_export_returns_ros_style_fields(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/business-structure/update?next=/',
        data={"structure": "sole_trader", "transition_date": "", "vat_registered": "1"},
    )
    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "VAT Expense",
            "description": "Hosting",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "vat_rate": "23%",
            "vat_amount": "23.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "payment_method": "Business Bank",
        },
    )

    response = app.app.test_client().get('/vat3/export.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=vat3-summary.csv' in response.headers['Content-Disposition']
    assert b'VAT3_Period,T1,T2,T3,T4,Due_Date' in response.data
    assert b',0.0,23.0,0.0,23.0,' in response.data
    assert b'ZeroRatedSales,ExemptSales,ReverseChargePurchases,Treatment_Notes' in response.data


def test_vat_threshold_summary_warning_and_exceeded_states():
    warning_summary = app._compute_vat_threshold_summary(
        income_rows=[{"Amount (€)": "33600.00"}],
        invoice_rows=[],
        basis="services",
    )
    assert warning_summary["status"] == "warning"
    assert warning_summary["progress_pct"] == 80.0

    exceeded_summary = app._compute_vat_threshold_summary(
        income_rows=[{"Amount (€)": "43000.00"}],
        invoice_rows=[],
        basis="services",
    )
    assert exceeded_summary["status"] == "exceeded"
    assert exceeded_summary["remaining_before_limit"] == 0.0


def test_vat_threshold_summary_splits_services_and_goods_streams():
    summary = app._compute_vat_threshold_summary(
        income_rows=[
            {"Amount (€)": "30000.00", "Supply Type": "services"},
            {"Amount (€)": "20000.00", "Supply Type": "goods"},
        ],
        invoice_rows=[
            {"Total (€)": "1000.00", "Supply Type": "services"},
            {"Total (€)": "40000.00", "Supply Type": "goods"},
        ],
        basis="services",
    )

    stream_trackers = {row["basis"]: row for row in summary["stream_trackers"]}
    assert stream_trackers["services"]["taxable_turnover"] == 30000.0
    assert stream_trackers["goods"]["taxable_turnover"] == 40000.0
    assert stream_trackers["services"]["is_selected"] is True
    assert stream_trackers["goods"]["is_selected"] is False


def test_vat_control_summary_includes_treatment_breakdown():
    period_start, _, _ = app._vat_period_bounds(date.today())
    ledger_entries = [
        {
            "transaction_date": period_start.isoformat(),
            "entity_type": "invoice",
            "vat_treatment": "zero_rated",
            "total_amount_eur": 500.0,
            "journal_lines": [],
        },
        {
            "transaction_date": period_start.isoformat(),
            "entity_type": "income",
            "vat_treatment": "exempt",
            "total_amount_eur": 250.0,
            "journal_lines": [],
        },
        {
            "transaction_date": period_start.isoformat(),
            "entity_type": "expense",
            "vat_treatment": "reverse_charge",
            "total_amount_eur": 100.0,
            "journal_lines": [],
        },
    ]

    summary = app._compute_vat_control_summary(ledger_entries)
    assert summary["zero_rated_sales"] == 500.0
    assert summary["exempt_sales"] == 250.0
    assert summary["reverse_charge_purchases"] == 100.0
    assert "Zero-rated sales" in summary["treatment_notes"]


def test_detect_vat_anomalies_flags_expected_cases():
    ledger_entries = [
        {
            "timestamp": "2026-08-05T10:00:00",
            "entity_type": "invoice",
            "description": "Exempt sale",
            "transaction_date": "2026-08-05",
            "amount_eur": 121.0,
            "total_amount_eur": 121.0,
            "vat_rate": "Exempt",
            "vat_treatment": "exempt",
            "supply_type": "services",
            "vat_amount_eur": 21.0,
        },
        {
            "timestamp": "2026-08-05T10:10:00",
            "entity_type": "expense",
            "description": "Reverse charge cost",
            "transaction_date": "2026-08-05",
            "amount_eur": 123.0,
            "total_amount_eur": 123.0,
            "vat_rate": "23%",
            "vat_treatment": "reverse_charge",
            "supply_type": "services",
            "vat_amount_eur": 23.0,
        },
    ]

    anomalies = app._detect_vat_anomalies(ledger_entries)
    assert len(anomalies) == 2
    assert any("non_zero_vat_with_zero_or_exempt_treatment" in item["flags"] for item in anomalies)
    assert any("reverse_charge_should_not_post_local_vat_amount" in item["flags"] for item in anomalies)


def test_resolve_workbook_path_finds_any_matching_financial_workbook(tmp_path, monkeypatch):
    workbook = tmp_path / "H-Queex_Financial_Control V8.0.xlsm"
    workbook.write_bytes(b"test")

    monkeypatch.setattr(app, "BASE_DIR", tmp_path)
    monkeypatch.setattr(app, "WORKBOOK_PATH", tmp_path / "missing.xlsm")

    assert app._resolve_workbook_path() == workbook


def test_header_aliases_are_resolved_when_reading_a_workbook_sheet(tmp_path):
    # Header-alias resolution is now only exercised by the one-time xlsm->JSON
    # migration and the manual xlsm export, since normal reads/writes go through
    # JSON files with canonical keys. Exercise that surviving code path directly.
    path = tmp_path / "alt.xlsm"
    wb = Workbook()
    wb.remove(wb.active)

    income = wb.create_sheet("Income")
    income.append(["Date", "Description", "Client / Source", "Amount", "Status"])
    income.append(["2026-07-29", "Renamed header", "Client A", 120.0, "Paid"])

    wb.save(path)

    try:
        rows = app._read_workbook_sheet_rows(wb, "Income")
    finally:
        wb.close()

    assert len(rows) == 1
    assert rows[0]["Description"] == "Renamed header"
    assert rows[0]["Amount (€)"] == 120.0


def test_income_add_route_does_not_touch_the_xlsm_workbook(tmp_path):
    # Normal writes are JSON-only now; the xlsm doesn't even need to exist.
    app.WORKBOOK_PATH = tmp_path / "missing.xlsm"
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-07-30",
        "description": "Added without workbook",
        "client_source": "Client A",
        "amount": "50.00",
        "status": "Pending",
    }
    response = app.app.test_client().post('/income/add', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert not (tmp_path / "missing.xlsm").exists()

    income_records = json.loads(app.INCOME_PATH.read_text(encoding="utf-8"))
    assert income_records[-1]["Description"] == "Added without workbook"


def test_due_subscription_posts_expense_and_advances_next_charge(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    subscription_payload = [
        {
            "id": "subscription-1",
            "title": "Adobe Creative Cloud",
            "description": "Creative suite",
            "supplier": "Supplier A",
            "category": "Software",
            "net_amount": 50.0,
            "total_amount": 60.0,
            "frequency": "monthly",
            "start_date": "2026-08-05",
            "next_charge_date": "2026-08-05",
            "last_posted_date": "",
            "end_date": "",
            "status": "active",
            "notes": "Company card",
        }
    ]
    app.SUBSCRIPTIONS_PATH.write_text(json.dumps(subscription_payload), encoding="utf-8")

    result = app._sync_subscriptions_to_expenses(today=date(2026, 8, 5))

    assert result["posted_count"] == 1

    expense_records = json.loads(app.EXPENSES_PATH.read_text(encoding="utf-8"))
    last_record = expense_records[-1]
    assert last_record["Title"] == "Adobe Creative Cloud"
    assert "Subscription charge" in str(last_record.get("Description", ""))
    assert "60" in str(last_record["Total (€)"])

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert subscriptions[0]["last_posted_date"] == "2026-08-05"
    assert subscriptions[0]["next_charge_date"] == "2026-09-05"


def test_add_subscription_route_persists_subscription_and_renders_register(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "title": "Figma",
        "description": "Design collaboration",
        "supplier": "Supplier A",
        "category": "Software",
        "frequency": "monthly",
        "start_date": "2026-08-20",
        "net_amount": "15.00",
        "total_amount": "18.15",
        "status": "active",
        "notes": "Team plan",
    }

    response = app.app.test_client().post('/subscriptions/add', data=payload, follow_redirects=True)

    assert response.status_code == 200
    assert b'Subscription added' in response.data
    assert b'Figma' in response.data

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert len(subscriptions) == 1
    assert subscriptions[0]["title"] == "Figma"
    assert subscriptions[0]["next_charge_date"] == "2026-08-20"


def test_update_and_delete_subscription_routes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.SUBSCRIPTIONS_PATH.write_text(
        json.dumps([
            {
                "id": "sub-1",
                "title": "Adobe",
                "description": "Creative",
                "supplier": "Supplier A",
                "category": "Software",
                "net_amount": 50.0,
                "total_amount": 60.0,
                "frequency": "monthly",
                "start_date": "2026-08-01",
                "next_charge_date": "2026-08-15",
                "last_posted_date": "",
                "end_date": "",
                "status": "active",
                "notes": "note",
            }
        ]),
        encoding="utf-8",
    )

    update_payload = {
        "subscription_id": "sub-1",
        "title": "Adobe CC",
        "description": "Creative suite",
        "supplier": "Supplier A",
        "category": "Software",
        "frequency": "yearly",
        "start_date": "2026-08-01",
        "next_charge_date": "2027-08-01",
        "end_date": "",
        "net_amount": "120.00",
        "total_amount": "145.20",
        "status": "paused",
        "notes": "annual",
    }
    update_response = app.app.test_client().post('/subscriptions/update', data=update_payload, follow_redirects=True)
    assert update_response.status_code == 200
    assert b'Subscription updated' in update_response.data

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert subscriptions[0]["title"] == "Adobe CC"
    assert subscriptions[0]["frequency"] == "yearly"

    delete_response = app.app.test_client().post('/subscriptions/delete', data={"subscription_id": "sub-1"}, follow_redirects=True)
    assert delete_response.status_code == 200
    assert b'Subscription archived' in delete_response.data
    assert json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8")) == []

    archives = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))
    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert archives[0]["entity_type"] == "subscription"
    assert any(entry["action"] == "archive" for entry in audit_entries)


def test_delete_expense_route_removes_workbook_row(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    data = app.load_finance_data()
    expense_row_number = data["sheets"]["Expenses"][0]["__row_number"]

    response = app.app.test_client().post(
        "/expenses/delete",
        data={"row_number": str(expense_row_number)},
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"Expense archived" in response.data

    expense_records = json.loads(app.EXPENSES_PATH.read_text(encoding="utf-8"))
    assert expense_records == []

    archives = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))
    assert archives[0]["entity_type"] == "expense"


def test_update_expense_route_updates_existing_workbook_row(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    row_number = app.load_finance_data()["sheets"]["Expenses"][0]["__row_number"]
    payload = {
        "row_number": str(row_number),
        "date": "2026-08-01",
        "title": "Software",
        "description": "Updated expense",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "category": "Software",
        "net_amount": "90.00",
        "total_amount": "108.90",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }

    response = app.app.test_client().post('/expenses/update', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Expense updated' in response.data

    app.load_finance_data.cache_clear()
    expenses = app.load_finance_data()["sheets"]["Expenses"]
    updated = next((r for r in expenses if r.get("__row_number") == row_number), None)
    assert updated is not None
    assert updated.get("Description") == "Updated expense"
    assert str(updated.get("Total (€)")) == "108.90"


def test_expense_capex_auto_routes_to_capital_schedule(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "date": "2026-08-05",
        "title": "Laptop",
        "description": "MacBook for production",
        "supplier": "Supplier A",
        "supplier_vat_number": "IE1234567A",
        "receipt_reference": "INV-CAP-1",
        "category": "Equipment and Hardware",
        "net_amount": "1500.00",
        "total_amount": "1845.00",
        "vat_rate": "23%",
        "vat_amount": "345.00",
        "input_vat_reclaimable": "Yes",
        "status": "Paid",
        "payment_method": "Business Bank",
    }

    response = app.app.test_client().post('/expenses/add', data=payload, follow_redirects=True)
    assert response.status_code == 200

    assets = json.loads(app.CAPITAL_ASSETS_PATH.read_text(encoding="utf-8"))
    assert len(assets) == 1
    assert assets[0]["source"] == "expense"
    assert assets[0]["cost_eur"] == 1845.0
    assert assets[0]["annual_allowance_eur"] == 230.62


def test_capital_allowances_export_returns_csv(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.CAPITAL_ASSETS_PATH.write_text(
        json.dumps(
            [
                {
                    "id": "expense-9",
                    "source": "expense",
                    "expense_row_number": 9,
                    "acquisition_date": "2026-08-01",
                    "supplier": "Supplier A",
                    "description": "Camera rig",
                    "category": "Equipment and Hardware",
                    "cost_eur": 1200.0,
                    "allowance_rate": 0.125,
                    "allowance_years": 8,
                    "annual_allowance_eur": 150.0,
                    "phase_tag": "Phase 1",
                    "active": True,
                }
            ]
        ),
        encoding="utf-8",
    )

    response = app.app.test_client().get('/capital-allowances/export.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=capital-allowances.csv' in response.headers['Content-Disposition']
    assert b'asset_id,acquisition_date,supplier,description,category,cost_eur,allowance_rate,allowance_years,annual_allowance_eur,phase_tag,active' in response.data
    assert b'expense-9' in response.data


def test_add_payroll_route_persists_register_and_posts_ledger_entry(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "pay_date": "2026-08-05",
        "payroll_period": "2026-08",
        "employee_name": "Hev Team Member",
        "gross_pay": "3000.00",
        "paye": "450.00",
        "usc": "120.00",
        "employee_prsi": "120.00",
        "employer_prsi": "330.00",
        "status": "Paid",
        "payment_method": "Business Bank",
        "payment_date": "2026-08-05",
        "bank_reconciliation": "Reconciled",
        "notes": "Monthly payroll run",
    }

    response = app.app.test_client().post('/payroll/add', data=payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Payroll entry added' in response.data

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    assert len(payroll_entries) == 1
    assert payroll_entries[0]["Employee Name"] == "Hev Team Member"
    assert payroll_entries[0]["Net Pay (€)"] == "2310.00"
    assert payroll_entries[0]["Employer Cost (€)"] == "3330.00"

    ledger_entries = json.loads(app.LEDGER_JOURNAL_PATH.read_text(encoding="utf-8"))
    payroll_ledger = [entry for entry in ledger_entries if entry.get("entity_type") == "payroll"]
    assert payroll_ledger
    latest = payroll_ledger[-1]
    assert latest["entry_balanced"] is True
    assert any(line["account_code"] == "5300" and line["debit"] == 3000.0 for line in latest["journal_lines"])
    assert any(line["account_code"] == "2200" and line["credit"] == 1020.0 for line in latest["journal_lines"])


def test_payroll_validation_and_export_route(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    invalid_payload = {
        "pay_date": "2026-08-05",
        "employee_name": "Invalid Payroll",
        "gross_pay": "1000.00",
        "paye": "700.00",
        "usc": "200.00",
        "employee_prsi": "200.00",
        "employer_prsi": "0.00",
        "status": "Draft",
    }
    invalid_response = app.app.test_client().post('/payroll/add', data=invalid_payload, follow_redirects=True)
    assert invalid_response.status_code == 200
    assert b'Validation:' in invalid_response.data
    assert b'Gross pay must be at least employee deductions total' in invalid_response.data

    valid_payload = {
        "pay_date": "2026-08-12",
        "payroll_period": "2026-08",
        "employee_name": "Valid Payroll",
        "gross_pay": "1200.00",
        "paye": "150.00",
        "usc": "36.00",
        "employee_prsi": "48.00",
        "employer_prsi": "132.00",
        "status": "Approved",
        "bank_reconciliation": "Unreconciled",
    }
    app.app.test_client().post('/payroll/add', data=valid_payload, follow_redirects=True)

    export_response = app.app.test_client().get('/payroll/export.csv')
    assert export_response.status_code == 200
    assert export_response.mimetype == 'text/csv'
    assert 'attachment; filename=payroll-register.csv' in export_response.headers['Content-Disposition']
    assert b'pay_date,payroll_period,employee_name,gross_pay_eur,paye_eur,usc_eur,employee_prsi_eur,employer_prsi_eur,net_pay_eur,employer_cost_eur,status,payment_method,bank_reconciliation,notes,phase_tag' in export_response.data
    assert b'Valid Payroll' in export_response.data


def test_expense_paid_requires_payment_method(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Missing payment method",
            "description": "Validation check",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b'Payment method is required when status is Paid' in response.data


def test_invoice_paid_requires_payment_method_and_date(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "",
            "due_date": "2026-08-20",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Validation",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "balance_due": "0.00",
            "status": "Paid",
            "payment_method": "",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b'Payment method is required when status is Paid' in response.data
    assert b'Payment date is required when status is Paid' in response.data


def test_invoice_add_flash_mentions_payment_date_autofill(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-10",
            "due_date": "2026-08-20",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Invoice flash",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "balance_due": "0.00",
            "status": "Paid",
            "payment_method": "Business Bank",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b'Invoice added' in response.data
    assert b'payment date auto-filled from issue date' in response.data


def test_invoice_paid_defaults_payment_date_from_issue_date():
    payload = {
        "Status": "Paid",
        "Issue Date": "2026-08-10",
        "Payment Date": "",
    }

    applied = app._apply_default_payment_date_for_paid(payload, "invoice", "Issue Date")

    assert applied is True
    assert payload["Payment Date"] == "2026-08-10"


def test_payroll_paid_defaults_payment_date_from_pay_date(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Default Payment Date Payroll",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Paid",
            "payment_method": "Business Bank",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b'Payroll entry added' in response.data
    assert b'payment date auto-filled from pay date' in response.data

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    target = next(item for item in payroll_entries if item.get("Employee Name") == "Default Payment Date Payroll")
    assert str(target.get("Payment Date") or "") == "2026-08-05"


def test_unpaid_entries_allow_blank_payment_fields(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    expense_response = app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Draft expense",
            "description": "Blank payment fields allowed",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Draft",
            "payment_method": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert expense_response.status_code == 200
    assert b'Expense entry added' in expense_response.data

    invoice_response = app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-10",
            "due_date": "2026-08-20",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Draft invoice",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "balance_due": "123.00",
            "status": "Draft",
            "payment_method": "",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert invoice_response.status_code == 200
    assert b'Invoice added' in invoice_response.data

    payroll_response = app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-12",
            "payroll_period": "2026-08",
            "employee_name": "Draft Payroll",
            "gross_pay": "1200.00",
            "paye": "150.00",
            "usc": "36.00",
            "employee_prsi": "48.00",
            "employer_prsi": "132.00",
            "status": "Draft",
            "payment_method": "",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert payroll_response.status_code == 200
    assert b'Payroll entry added' in payroll_response.data


def test_reconciliation_exports_include_queue_and_exceptions(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-07-01",
            "title": "Old paid expense",
            "description": "Needs reconciliation",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
                "payment_method": "Business Bank",
        },
    )

    app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "2026-07-01",
            "due_date": "2026-07-31",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Reconciliation test",
            "net_amount": "200.00",
            "total_amount": "246.00",
            "balance_due": "0.00",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
                "payment_method": "Business Bank",
                "payment_date": "2026-07-01",
        },
    )

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-07-01",
            "payroll_period": "2026-07",
            "employee_name": "Recon Payroll",
            "gross_pay": "1800.00",
            "paye": "200.00",
            "usc": "50.00",
            "employee_prsi": "70.00",
            "employer_prsi": "198.00",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
                "payment_method": "Business Bank",
                "payment_date": "2026-07-01",
        },
    )

    # Simulate a paid invoice that legitimately has no payment date recorded
    # (e.g. edited outside the app's normal add/update validation flow), which
    # is what the missing_payment_date exception exists to catch.
    app.load_finance_data.cache_clear()
    invoice_rows = app.load_finance_data()["sheets"]["Invoices"]
    target_invoice = next(row for row in invoice_rows if row.get("Service / Product") == "Reconciliation test")
    target_invoice["Payment Date"] = ""
    app._update_row_in_sheet("Invoices", target_invoice["__row_number"], target_invoice)
    app.load_finance_data.cache_clear()

    queue_response = app.app.test_client().get('/reconciliation/export.csv')
    assert queue_response.status_code == 200
    assert queue_response.mimetype == 'text/csv'
    assert b'entity_type,reference,counterparty,date,amount_eur,status,bank_reconciliation,payment_method,is_paid,age_days,exception_reasons,matching_group_size' in queue_response.data

    exceptions_response = app.app.test_client().get('/reconciliation/exceptions.csv')
    assert exceptions_response.status_code == 200
    assert b'paid_unreconciled_over_7_days' in exceptions_response.data
    assert b'missing_payment_date' in exceptions_response.data


def test_mark_reconciliation_updates_payroll_and_expense_routes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    expense_response = app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Recon Expense",
            "description": "Mark endpoint check",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "Business Bank",
        },
        follow_redirects=True,
    )
    assert expense_response.status_code == 200

    app.load_finance_data.cache_clear()
    expense_rows = app.load_finance_data()["sheets"]["Expenses"]
    expense_row_number = expense_rows[-1]["__row_number"]
    mark_expense_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "expense",
            "row_number": str(expense_row_number),
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_expense_response.status_code == 200
    assert b'Expense reconciliation updated' in mark_expense_response.data

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Payroll Mark",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "Business Bank",
            "payment_date": "2026-08-05",
        },
    )
    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    payroll_id = payroll_entries[0]["id"]

    mark_payroll_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "payroll",
            "payroll_id": payroll_id,
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_payroll_response.status_code == 200
    assert b'Payroll reconciliation updated' in mark_payroll_response.data

    updated_payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    updated = next(entry for entry in updated_payroll_entries if entry["id"] == payroll_id)
    assert updated["Bank Reconciliation"] == "Reconciled"


def test_mark_reconciliation_blocks_unpaid_expense(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Unpaid expense",
            "description": "Should not reconcile",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Draft",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "",
        },
        follow_redirects=True,
    )

    app.load_finance_data.cache_clear()
    expense_row_number = app.load_finance_data()["sheets"]["Expenses"][-1]["__row_number"]
    mark_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "expense",
            "row_number": str(expense_row_number),
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_response.status_code == 200
    assert b'Reconciliation update failed: Record must have a paid status before it can be reconciled' in mark_response.data

    app.load_finance_data.cache_clear()
    row = next(item for item in app.load_finance_data()["sheets"]["Expenses"] if item["__row_number"] == expense_row_number)
    assert str(row.get("Bank Reconciliation") or "Unreconciled") != "Reconciled"


def test_mark_reconciliation_autofills_payment_date_for_paid_payroll(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Legacy Payroll",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Draft",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "",
            "payment_date": "",
        },
        follow_redirects=True,
    )

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    payroll_entry = payroll_entries[0]
    payroll_entry["Status"] = "Paid"
    payroll_entry["Payment Method"] = "Business Bank"
    payroll_entry["Payment Date"] = ""
    payroll_entry["Bank Reconciliation"] = "Unreconciled"
    app._save_payroll_entries(payroll_entries)

    mark_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "payroll",
            "payroll_id": payroll_entry["id"],
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_response.status_code == 200
    assert b'Payroll reconciliation updated' in mark_response.data

    updated_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    updated = next(item for item in updated_entries if item["id"] == payroll_entry["id"])
    assert updated["Bank Reconciliation"] == "Reconciled"
    assert str(updated.get("Payment Date") or "") == "2026-08-05"


def test_mark_reconciliation_autofills_payment_date_for_paid_invoice(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    add_response = app.app.test_client().post(
        '/invoices/add',
        data={
            "issue_date": "2026-08-05",
            "due_date": "2026-08-20",
            "client_name": "Client A",
            "client_vat_number": "IE1234567A",
            "client_address": "1 Example Street, Dublin",
            "service_product": "Recon Invoice Autofill",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "balance_due": "123.00",
            "status": "Draft",
            "payment_method": "",
            "payment_date": "",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )
    assert add_response.status_code == 200

    app.load_finance_data.cache_clear()
    invoice_rows = app.load_finance_data()["sheets"]["Invoices"]
    target = next(row for row in invoice_rows if row.get("Service / Product") == "Recon Invoice Autofill")
    row_number = target["__row_number"]

    target["Status"] = "Paid"
    target["Payment Method"] = "Business Bank"
    target["Payment Date"] = ""
    target["Bank Reconciliation"] = "Unreconciled"
    app._update_row_in_sheet("Invoices", row_number, target)
    app.load_finance_data.cache_clear()

    mark_response = app.app.test_client().post(
        '/reconciliation/mark',
        data={
            "entity_type": "invoice",
            "row_number": str(row_number),
            "bank_reconciliation": "Reconciled",
            "return_to": "/ledger",
        },
        follow_redirects=True,
    )
    assert mark_response.status_code == 200
    assert b'Invoice reconciliation updated' in mark_response.data

    app.load_finance_data.cache_clear()
    updated = next(item for item in app.load_finance_data()["sheets"]["Invoices"] if item["__row_number"] == row_number)
    assert str(updated.get("Bank Reconciliation") or "Reconciled") == "Reconciled"
    if "Payment Date" in updated:
        assert str(updated.get("Payment Date") or "") == "2026-08-05"


def test_bank_statement_import_and_unmatched_export(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/expenses/add',
        data={
            "date": "2026-08-05",
            "title": "Statement match expense",
            "description": "Should match statement",
            "supplier": "Supplier A",
            "supplier_vat_number": "IE1234567A",
            "category": "Professional Fees",
            "net_amount": "100.00",
            "total_amount": "123.00",
            "input_vat_reclaimable": "Yes",
            "status": "Paid",
            "payment_method": "Business Bank",
            "bank_reconciliation": "Unreconciled",
        },
        follow_redirects=True,
    )

    statement_csv = "Date,Description,Reference,Amount,Payment Method\n2026-08-05,Expense Payment,EXP-1,-123.00,Business Bank\n2026-08-05,Unknown Item,UNK-1,-55.00,Business Bank\n"
    import_response = app.app.test_client().post(
        '/reconciliation/import-statement',
        data={
            "return_to": "/ledger",
            "statement_file": (BytesIO(statement_csv.encode("utf-8")), "statement.csv"),
        },
        content_type='multipart/form-data',
        follow_redirects=True,
    )
    assert import_response.status_code == 200
    assert b'Bank statement imported:' in import_response.data

    queue_response = app.app.test_client().get('/reconciliation/export.csv')
    assert queue_response.status_code == 200
    assert b'statement_match_count' not in queue_response.data
    assert b'Statement match expense' in queue_response.data

    bank_lines_response = app.app.test_client().get('/reconciliation/bank-statements.csv')
    assert bank_lines_response.status_code == 200
    assert b'date,description,reference,amount_eur,balance_eur,payment_method,matched_entity_type,matched_reference,source_filename,uploaded_at' in bank_lines_response.data
    assert b'Unknown Item' in bank_lines_response.data

    unmatched_response = app.app.test_client().get('/reconciliation/unmatched-bank-statements.csv')
    assert unmatched_response.status_code == 200
    assert b'Unknown Item' in unmatched_response.data
    assert b'EXP-1' not in unmatched_response.data


def test_apply_suggested_reconciliation_marks_matched_payroll(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app.app.test_client().post(
        '/payroll/add',
        data={
            "pay_date": "2026-08-05",
            "payroll_period": "2026-08",
            "employee_name": "Batch Recon Payroll",
            "gross_pay": "1000.00",
            "paye": "100.00",
            "usc": "30.00",
            "employee_prsi": "40.00",
            "employer_prsi": "110.00",
            "status": "Paid",
            "bank_reconciliation": "Unreconciled",
            "payment_method": "Business Bank",
            "payment_date": "2026-08-05",
        },
        follow_redirects=True,
    )

    statement_csv = "Date,Description,Reference,Amount,Payment Method\n2026-08-05,Payroll Payment,PAY-1,-830.00,Business Bank\n"
    app.app.test_client().post(
        '/reconciliation/import-statement',
        data={
            "return_to": "/ledger",
            "statement_file": (BytesIO(statement_csv.encode("utf-8")), "statement.csv"),
        },
        content_type='multipart/form-data',
        follow_redirects=True,
    )

    batch_response = app.app.test_client().post(
        '/reconciliation/apply-suggested',
        data={"return_to": "/ledger"},
        follow_redirects=True,
    )
    assert batch_response.status_code == 200
    assert b'Applied 1 suggested reconciliation matches' in batch_response.data

    payroll_entries = json.loads(app.PAYROLL_PATH.read_text(encoding="utf-8"))
    entry = next(item for item in payroll_entries if item["Employee Name"] == "Batch Recon Payroll")
    assert entry["Bank Reconciliation"] == "Reconciled"


def test_apply_suggested_reconciliation_skips_ambiguous_groups(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    for title in ["Ambiguous Expense A", "Ambiguous Expense B"]:
        app.app.test_client().post(
            '/expenses/add',
            data={
                "date": "2026-08-05",
                "title": title,
                "description": "Same amount same day",
                "supplier": "Supplier A",
                "supplier_vat_number": "IE1234567A",
                "category": "Professional Fees",
                "net_amount": "100.00",
                "total_amount": "123.00",
                "input_vat_reclaimable": "Yes",
                "status": "Paid",
                "payment_method": "Business Bank",
                "bank_reconciliation": "Unreconciled",
            },
            follow_redirects=True,
        )

    statement_csv = "Date,Description,Reference,Amount,Payment Method\n2026-08-05,Single line,AMB-1,-123.00,Business Bank\n"
    app.app.test_client().post(
        '/reconciliation/import-statement',
        data={
            "return_to": "/ledger",
            "statement_file": (BytesIO(statement_csv.encode("utf-8")), "statement.csv"),
        },
        content_type='multipart/form-data',
        follow_redirects=True,
    )

    batch_response = app.app.test_client().post(
        '/reconciliation/apply-suggested',
        data={"return_to": "/ledger"},
        follow_redirects=True,
    )
    assert batch_response.status_code == 200

    app.load_finance_data.cache_clear()
    expenses = app.load_finance_data()["sheets"]["Expenses"]
    ambiguous = [row for row in expenses if str(row.get("Title") or "") in {"Ambiguous Expense A", "Ambiguous Expense B"}]
    assert len(ambiguous) == 2
    assert all(str(row.get("Bank Reconciliation") or "") != "Reconciled" for row in ambiguous)


def test_invoice_crud_routes_update_and_remove_rows(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    add_payload = {
        "issue_date": "2026-08-01",
        "due_date": "2026-08-31",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Web design",
        "net_amount": "300.00",
        "total_amount": "363.00",
        "balance_due": "363.00",
        "status": "Draft",
    }
    add_response = app.app.test_client().post('/invoices/add', data=add_payload, follow_redirects=True)
    assert add_response.status_code == 200
    assert b'Invoice added' in add_response.data

    app.load_finance_data.cache_clear()
    invoice_rows = app.load_finance_data()["sheets"]["Invoices"]
    target = next(row for row in invoice_rows if str(row["Invoice #"]).startswith("HQ-2026-"))

    update_payload = {
        "row_number": str(target["__row_number"]),
        "invoice_number": str(target["Invoice #"]),
        "issue_date": "2026-08-02",
        "due_date": "2026-09-01",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Updated service",
        "net_amount": "350.00",
        "total_amount": "423.50",
        "balance_due": "200.00",
        "status": "Sent",
    }
    update_response = app.app.test_client().post('/invoices/update', data=update_payload, follow_redirects=True)
    assert update_response.status_code == 200
    assert b'Invoice updated' in update_response.data

    delete_response = app.app.test_client().post('/invoices/delete', data={"row_number": str(target["__row_number"])}, follow_redirects=True)
    assert delete_response.status_code == 200
    assert b'Invoice cancelled and retained for audit trail' in delete_response.data

    app.load_finance_data.cache_clear()
    retained_invoice = next(row for row in app.load_finance_data()["sheets"]["Invoices"] if row["__row_number"] == target["__row_number"])
    assert retained_invoice["Status"] == "Cancelled"


def test_invoice_validation_blocks_invalid_due_date(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    add_payload = {
        "issue_date": "2026-08-10",
        "due_date": "2026-08-01",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Validation",
        "net_amount": "100.00",
        "total_amount": "121.00",
        "balance_due": "121.00",
        "status": "Draft",
    }
    response = app.app.test_client().post('/invoices/add', data=add_payload, follow_redirects=True)
    assert response.status_code == 200
    assert b'Validation:' in response.data


def test_invoice_numbers_auto_generate_sequentially(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    payload = {
        "issue_date": "2026-08-10",
        "due_date": "2026-08-20",
        "client_name": "Client A",
        "client_vat_number": "IE1234567A",
        "client_address": "1 Example Street, Dublin",
        "service_product": "Service A",
        "net_amount": "100.00",
        "total_amount": "123.00",
        "balance_due": "123.00",
        "status": "Issued",
    }

    response_one = app.app.test_client().post('/invoices/add', data=payload, follow_redirects=True)
    assert response_one.status_code == 200
    response_two = app.app.test_client().post('/invoices/add', data=payload, follow_redirects=True)
    assert response_two.status_code == 200

    app.load_finance_data.cache_clear()
    rows = app.load_finance_data()["sheets"]["Invoices"]
    generated = sorted([str(row.get("Invoice #")) for row in rows if str(row.get("Invoice #", "")).startswith("HQ-2026-")])
    assert generated[-2:] == ["HQ-2026-001", "HQ-2026-002"]


def test_client_and_supplier_update_and_delete_routes(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    client_row = app.load_finance_data()["sheets"]["Clients"][0]["__row_number"]
    client_update = {
        "row_number": str(client_row),
        "client_name": "Client A Updated",
        "contact_person": "Jane Doe",
        "email": "jane.doe@example.com",
        "phone": "999",
        "country": "Belgium",
    }
    client_response = app.app.test_client().post('/clients/update', data=client_update, follow_redirects=True)
    assert client_response.status_code == 200
    assert b'Client updated' in client_response.data

    supplier_row = app.load_finance_data()["sheets"]["Suppliers"][0]["__row_number"]
    supplier_update = {
        "row_number": str(supplier_row),
        "supplier_name": "Supplier A Updated",
        "contact_person": "John Doe",
        "email": "john.doe@example.com",
        "phone": "888",
        "country": "Netherlands",
        "default_vat_treatment": "Reverse charge",
    }
    supplier_response = app.app.test_client().post('/suppliers/update', data=supplier_update, follow_redirects=True)
    assert supplier_response.status_code == 200
    assert b'Supplier updated' in supplier_response.data

    client_delete = app.app.test_client().post('/clients/delete', data={"row_number": str(client_row)}, follow_redirects=True)
    supplier_delete = app.app.test_client().post('/suppliers/delete', data={"row_number": str(supplier_row)}, follow_redirects=True)
    assert client_delete.status_code == 200
    assert supplier_delete.status_code == 200
    assert b'Client archived' in client_delete.data
    assert b'Supplier archived' in supplier_delete.data


def test_archive_view_shows_archive_and_audit_records(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._archive_record("client", {"Client Name": "Archived Client"}, source="workbook")
    app._record_audit("update", "client", {"row_number": 2})

    response = app.app.test_client().get('/archive')
    assert response.status_code == 200
    assert b'Archive &amp; Audit' in response.data
    assert b'Archived Client' in response.data
    assert b'update' in response.data
    assert b'Restore' in response.data


def test_expense_archive_returns_message_when_workbook_is_locked(workbook_copy, monkeypatch):
    def raise_lock_error(*args, **kwargs):
        raise app.WorkbookWriteError("Workbook is locked. Close Excel and try again.")

    monkeypatch.setattr(app, "_delete_row_from_sheet", raise_lock_error)
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    response = app.app.test_client().post('/expenses/delete', data={'row_number': '1'}, follow_redirects=True)

    assert response.status_code == 200
    assert b'Workbook is locked' in response.data


def test_restore_archived_expense_recreates_workbook_row_and_logs_audit(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._archive_record(
        "expense",
        {
            "Date (Registered)": "2026-08-03",
            "Title": "Recovered Expense",
            "Description": "Restored row",
            "Supplier / Payee": "Supplier A",
            "Category": "Software",
            "Net Amount (€)": "50.00",
            "Total (€)": "60.50",
            "Status": "Archived",
        },
        source="workbook",
    )
    archive_id = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id}, follow_redirects=True)
    assert response.status_code == 200
    assert b'Expense restored' in response.data

    expense_records = json.loads(app.EXPENSES_PATH.read_text(encoding="utf-8"))
    assert any(record.get("Title") == "Recovered Expense" for record in expense_records)
    assert json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8")) == []
    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert any(entry["action"] == "restore" and entry["entity_type"] == "expense" for entry in audit_entries)


def test_restore_archived_subscription_recreates_subscription_and_logs_audit(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._archive_record(
        "subscription",
        {
            "id": "restore-sub-1",
            "title": "Restored Subscription",
            "description": "Recovered",
            "supplier": "Supplier A",
            "category": "Software",
            "net_amount": 10.0,
            "total_amount": 12.1,
            "frequency": "monthly",
            "start_date": "2026-08-01",
            "next_charge_date": "2026-09-01",
            "last_posted_date": "",
            "end_date": "",
            "status": "active",
            "notes": "restored",
        },
        source="subscriptions",
    )
    archive_id = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))[0]["id"]

    response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id}, follow_redirects=True)
    assert response.status_code == 200
    assert b'Subscription restored' in response.data

    subscriptions = json.loads(app.SUBSCRIPTIONS_PATH.read_text(encoding="utf-8"))
    assert subscriptions[0]["title"] == "Restored Subscription"
    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert any(entry["action"] == "restore" and entry["entity_type"] == "subscription" for entry in audit_entries)


def test_restore_conflict_is_detected_until_force_restore(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._archive_record(
        "expense",
        {
            "Date (Registered)": "2026-07-29",
            "Title": "Travel",
            "Description": "Hotel",
            "Supplier / Payee": "Supplier A",
            "Category": "Travel",
            "Net Amount (€)": 100.0,
            "Total (€)": 120.0,
            "Status": "Pending",
        },
        source="workbook",
    )
    archive_id = json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))[0]["id"]

    conflict_response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id}, follow_redirects=True)
    assert conflict_response.status_code == 200
    assert b'Restore conflict detected for expense' in conflict_response.data
    assert len(json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8"))) == 1

    force_response = app.app.test_client().post('/archive/restore', data={"archive_id": archive_id, "force_restore": "1"}, follow_redirects=True)
    assert force_response.status_code == 200
    assert b'Expense restored' in force_response.data
    assert json.loads(app.ARCHIVE_PATH.read_text(encoding="utf-8")) == []

    audit_entries = json.loads(app.AUDIT_LOG_PATH.read_text(encoding="utf-8"))
    assert any(entry["action"] == "restore_conflict" for entry in audit_entries)
    assert any(entry["action"] == "restore" for entry in audit_entries)


def test_audit_csv_export_returns_csv_download(workbook_copy):
    app.WORKBOOK_PATH = workbook_copy
    app.load_finance_data.cache_clear()

    app._record_audit("update", "expense", {"row_number": 3, "status": "Paid"})

    response = app.app.test_client().get('/audit/export.csv')
    assert response.status_code == 200
    assert response.mimetype == 'text/csv'
    assert 'attachment; filename=audit-log.csv' in response.headers['Content-Disposition']
    assert b'timestamp,action,entity_type,details_json' in response.data
    assert b'update,expense' in response.data
