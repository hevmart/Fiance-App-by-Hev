from pathlib import Path
import tempfile

from openpyxl import load_workbook
import app

with tempfile.TemporaryDirectory() as tmpdir:
    temp_path = Path(tmpdir) / app.WORKBOOK_PATH.name
    wb = load_workbook(app.WORKBOOK_PATH, data_only=False)
    wb.save(temp_path)
    wb.close()

    app.WORKBOOK_PATH = temp_path
    app.load_finance_data.cache_clear()

    client = app.app.test_client()
    payloads = [
        ('/income/add', {'date': '2026-07-29', 'description': 'Test income', 'client_source': 'Client A', 'amount': '150.00', 'status': 'Paid'}),
        ('/expenses/add', {'date': '2026-07-29', 'title': 'Test expense', 'description': 'Test', 'supplier': 'Supplier A', 'total_amount': '25.00', 'status': 'Pending'}),
        ('/clients/add', {'client_name': 'Test Client', 'contact_person': 'Jane', 'email': 'jane@example.com'}),
        ('/suppliers/add', {'supplier_name': 'Test Supplier', 'contact_person': 'John', 'email': 'john@example.com'})
    ]

    for path, data in payloads:
        response = client.post(path, data=data, follow_redirects=True)
        print(path, response.status_code)

    wb = load_workbook(temp_path, data_only=True)
    for sheet_name in ['Income', 'Expenses', 'Clients', 'Suppliers']:
        ws = wb[sheet_name]
        rows = [row for row in ws.iter_rows(values_only=True) if any(v not in (None, '') for v in row)]
        print(sheet_name, 'rows', len(rows))
        if rows:
            print('last_row', rows[-1])
    wb.close()
